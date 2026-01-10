#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
レポートデータ品質検証スクリプト

生成されたExcelレポートの指数整合性を詳細に確認する
"""

import pandas as pd
import numpy as np
import openpyxl
from pathlib import Path
import warnings
warnings.filterwarnings('ignore')

def load_report_sheets(excel_path: str) -> dict:
    """Excelファイルからレースシートを読み込む"""
    wb = openpyxl.load_workbook(excel_path)
    sheets = {}
    for sheet_name in wb.sheetnames:
        # Race_01, Race_02, ... 形式のシートを読み込む
        if sheet_name.startswith('Race_'):
            ws = wb[sheet_name]
            data = []
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = row
                else:
                    data.append(row)
            if headers and data:
                df = pd.DataFrame(data, columns=headers)
                sheets[sheet_name] = df
    return sheets

def analyze_index_consistency(sheets: dict, venue_name: str) -> dict:
    """指数の整合性を分析"""
    issues = []
    stats = {
        'venue': venue_name,
        'total_races': len(sheets),
        'total_horses': 0,
        'T指数_stats': {'min': np.inf, 'max': -np.inf, 'missing': 0, 'outliers': []},
        'L指数_stats': {'min': np.inf, 'max': -np.inf, 'missing': 0, 'outliers': []},
        'issues': issues
    }
    
    # カラム名のマッピング（Excelでは「タイム指数」「上り指数」）
    T_COL = 'タイム指数'
    L_COL = '上り指数'
    
    for race_name, df in sheets.items():
        if df.empty:
            issues.append(f"{race_name}: データが空です")
            continue
            
        # 馬番と枠番の妥当性確認
        for col in ['出走枠番', '出走馬番']:
            if col in df.columns:
                try:
                    vals = pd.to_numeric(df[col], errors='coerce')
                    invalid = df[vals.isna() | (vals <= 0)][col].tolist()
                    if invalid:
                        issues.append(f"{race_name}: {col}に異常値あり: {invalid[:5]}")
                except:
                    pass
        
        # T指数(タイム指数)の確認
        if T_COL in df.columns:
            for idx, row in df.iterrows():
                t_idx = row.get(T_COL)
                if pd.isna(t_idx) or t_idx == '' or t_idx is None:
                    stats['T指数_stats']['missing'] += 1
                else:
                    try:
                        t_val = float(t_idx)
                        stats['T指数_stats']['min'] = min(stats['T指数_stats']['min'], t_val)
                        stats['T指数_stats']['max'] = max(stats['T指数_stats']['max'], t_val)
                        # 異常値チェック (20未満または80超)
                        if t_val < 20 or t_val > 80:
                            stats['T指数_stats']['outliers'].append({
                                'race': race_name,
                                'horse': row.get('馬名', 'Unknown'),
                                'value': t_val
                            })
                    except (ValueError, TypeError):
                        stats['T指数_stats']['missing'] += 1
        
        # L指数(上り指数)の確認
        if L_COL in df.columns:
            for idx, row in df.iterrows():
                l_idx = row.get(L_COL)
                if pd.isna(l_idx) or l_idx == '' or l_idx is None:
                    stats['L指数_stats']['missing'] += 1
                else:
                    try:
                        l_val = float(l_idx)
                        stats['L指数_stats']['min'] = min(stats['L指数_stats']['min'], l_val)
                        stats['L指数_stats']['max'] = max(stats['L指数_stats']['max'], l_val)
                        # 異常値チェック
                        if l_val < 20 or l_val > 80:
                            stats['L指数_stats']['outliers'].append({
                                'race': race_name,
                                'horse': row.get('馬名', 'Unknown'),
                                'value': l_val
                            })
                    except (ValueError, TypeError):
                        stats['L指数_stats']['missing'] += 1
        
        # タイム・上がりとの整合性チェック
        if 'タイム秒' in df.columns and T_COL in df.columns:
            for idx, row in df.iterrows():
                time_val = row.get('タイム秒')
                t_idx = row.get(T_COL)
                if time_val and t_idx:
                    try:
                        time_f = float(time_val)
                        t_f = float(t_idx)
                        # 極端な相関ずれを検出 (速いタイムなのに低い指数など)
                        # 通常: 速い(低)タイム → 高いT指数
                    except:
                        pass
        
        stats['total_horses'] += len(df)
    
    return stats

def compare_venues(stats_list: list) -> list:
    """競馬場間の指数分布を比較"""
    comparisons = []
    
    for i, s1 in enumerate(stats_list):
        for s2 in stats_list[i+1:]:
            t1_range = s1['T指数_stats']['max'] - s1['T指数_stats']['min'] if s1['T指数_stats']['max'] != -np.inf else 0
            t2_range = s2['T指数_stats']['max'] - s2['T指数_stats']['min'] if s2['T指数_stats']['max'] != -np.inf else 0
            
            if abs(t1_range - t2_range) > 20:
                comparisons.append(f"⚠️ T指数のレンジが大きく異なる: {s1['venue']}={t1_range:.1f}, {s2['venue']}={t2_range:.1f}")
            
            # 欠損率の比較
            miss1 = s1['T指数_stats']['missing'] / max(s1['total_horses'], 1) * 100
            miss2 = s2['T指数_stats']['missing'] / max(s2['total_horses'], 1) * 100
            if abs(miss1 - miss2) > 30:
                comparisons.append(f"⚠️ T指数の欠損率が大きく異なる: {s1['venue']}={miss1:.1f}%, {s2['venue']}={miss2:.1f}%")
    
    return comparisons

def check_horse_data_quality(sheets: dict, venue_name: str) -> list:
    """馬ごとのデータ品質をチェック"""
    issues = []
    
    for race_name, df in sheets.items():
        if df.empty:
            continue
            
        # 馬名の重複チェック
        if '馬名' in df.columns:
            duplicates = df[df['馬名'].duplicated(keep=False)]['馬名'].unique()
            if len(duplicates) > 0:
                issues.append(f"{race_name}: 馬名重複あり: {list(duplicates)[:3]}")
        
        # 過去成績データの確認
        history_cols = ['日付', '場所', 'ｺｰｽ', '距離', '着順']
        found_cols = [c for c in history_cols if c in df.columns]
        
        for idx, row in df.iterrows():
            horse_name = row.get('馬名', f'Horse_{idx}')
            
            # 着順が0または異常値
            finish = row.get('着順')
            if finish:
                try:
                    f = int(finish)
                    if f <= 0 or f > 28:
                        issues.append(f"{race_name}/{horse_name}: 着順異常値 ({f})")
                except:
                    pass
            
            # タイムの妥当性
            time_val = row.get('タイム')
            if time_val:
                try:
                    t = float(time_val)
                    dist = row.get('距離')
                    if dist:
                        d = int(dist)
                        # 大まかな妥当性: 1200m=70秒、1600m=95秒、2000m=120秒
                        expected = d / 1200 * 70
                        if abs(t - expected) > expected * 0.2:  # 20%以上のずれ
                            issues.append(f"{race_name}/{horse_name}: タイム異常の可能性 ({t}秒 @ {d}m)")
                except:
                    pass
    
    return issues[:20]  # 最初の20件のみ

def analyze_metric_calculation(sheets: dict, venue_name: str) -> dict:
    """指数計算の詳細分析"""
    analysis = {
        'venue': venue_name,
        'races': {}
    }
    
    for race_name, df in sheets.items():
        race_analysis = {
            'num_horses': len(df),
            'T指数': {'values': [], 'mean': None, 'std': None},
            'L指数': {'values': [], 'mean': None, 'std': None},
            'correlations': {}
        }
        
        t_values = []
        l_values = []
        times = []
        agari = []
        
        for idx, row in df.iterrows():
            # T指数収集 (Excelカラム名: タイム指数)
            t_idx = row.get('タイム指数')
            if t_idx and t_idx != '':
                try:
                    t_values.append(float(t_idx))
                except:
                    pass
            
            # L指数収集 (Excelカラム名: 上り指数)
            l_idx = row.get('上り指数')
            if l_idx and l_idx != '':
                try:
                    l_values.append(float(l_idx))
                except:
                    pass
                    
            # タイム収集 (Excelカラム名: タイム秒)
            time_val = row.get('タイム秒')
            if time_val:
                try:
                    times.append(float(time_val))
                except:
                    pass
                    
            # 上がり収集
            agari_val = row.get('上がり') or row.get('上り秒')
            if agari_val:
                try:
                    agari.append(float(agari_val))
                except:
                    pass
        
        if t_values:
            race_analysis['T指数']['values'] = t_values
            race_analysis['T指数']['mean'] = np.mean(t_values)
            race_analysis['T指数']['std'] = np.std(t_values)
            
        if l_values:
            race_analysis['L指数']['values'] = l_values
            race_analysis['L指数']['mean'] = np.mean(l_values)
            race_analysis['L指数']['std'] = np.std(l_values)
        
        # 相関分析
        if len(times) >= 3 and len(t_values) >= 3:
            min_len = min(len(times), len(t_values))
            if min_len >= 3:
                try:
                    corr = np.corrcoef(times[:min_len], t_values[:min_len])[0, 1]
                    race_analysis['correlations']['タイム_vs_T指数'] = corr
                except:
                    pass
                    
        if len(agari) >= 3 and len(l_values) >= 3:
            min_len = min(len(agari), len(l_values))
            if min_len >= 3:
                try:
                    corr = np.corrcoef(agari[:min_len], l_values[:min_len])[0, 1]
                    race_analysis['correlations']['上がり_vs_L指数'] = corr
                except:
                    pass
        
        analysis['races'][race_name] = race_analysis
    
    return analysis

def main():
    print("=" * 70)
    print("レポートデータ品質検証")
    print("=" * 70)
    
    reports_dir = Path(r"C:\Users\zk-ht\Keiba\Keiba_AI_v2\outputs\reports")
    
    # 2025/12/28のレポートを分析
    files = [
        ("中山", reports_dir / "race_analysis_20251228_中山.xlsx"),
        ("阪神", reports_dir / "race_analysis_20251228_阪神.xlsx"),
    ]
    
    all_stats = []
    all_analyses = []
    
    for venue_name, file_path in files:
        if not file_path.exists():
            print(f"\n❌ {venue_name}: ファイルが見つかりません: {file_path}")
            continue
            
        print(f"\n{'='*60}")
        print(f"📊 {venue_name}競馬場 分析")
        print(f"{'='*60}")
        
        # シート読み込み
        sheets = load_report_sheets(str(file_path))
        print(f"レース数: {len(sheets)}")
        
        # 指数整合性分析
        stats = analyze_index_consistency(sheets, venue_name)
        all_stats.append(stats)
        
        print(f"\n■ 基本統計:")
        print(f"  総馬数: {stats['total_horses']}")
        
        if stats['T指数_stats']['min'] != np.inf:
            print(f"  T指数 範囲: {stats['T指数_stats']['min']:.1f} ~ {stats['T指数_stats']['max']:.1f}")
            print(f"  T指数 欠損: {stats['T指数_stats']['missing']}件")
            if stats['T指数_stats']['outliers']:
                print(f"  T指数 外れ値:")
                for o in stats['T指数_stats']['outliers'][:5]:
                    print(f"    - {o['race']}/{o['horse']}: {o['value']:.1f}")
        else:
            print(f"  T指数: データなし")
                    
        if stats['L指数_stats']['min'] != np.inf:
            print(f"  L指数 範囲: {stats['L指数_stats']['min']:.1f} ~ {stats['L指数_stats']['max']:.1f}")
            print(f"  L指数 欠損: {stats['L指数_stats']['missing']}件")
            if stats['L指数_stats']['outliers']:
                print(f"  L指数 外れ値:")
                for o in stats['L指数_stats']['outliers'][:5]:
                    print(f"    - {o['race']}/{o['horse']}: {o['value']:.1f}")
        else:
            print(f"  L指数: データなし")
        
        # 問題点
        if stats['issues']:
            print(f"\n■ 検出された問題:")
            for issue in stats['issues'][:10]:
                print(f"  ⚠️ {issue}")
        
        # 詳細分析
        analysis = analyze_metric_calculation(sheets, venue_name)
        all_analyses.append(analysis)
        
        print(f"\n■ レース別指数分析:")
        for race_name, race_data in analysis['races'].items():
            t_mean = race_data['T指数'].get('mean')
            t_std = race_data['T指数'].get('std')
            l_mean = race_data['L指数'].get('mean')
            l_std = race_data['L指数'].get('std')
            
            print(f"\n  {race_name} ({race_data['num_horses']}頭):")
            if t_mean is not None:
                print(f"    T指数: 平均={t_mean:.1f}, 標準偏差={t_std:.1f}")
            else:
                print(f"    T指数: データ不足")
                
            if l_mean is not None:
                print(f"    L指数: 平均={l_mean:.1f}, 標準偏差={l_std:.1f}")
            else:
                print(f"    L指数: データ不足")
                
            corrs = race_data.get('correlations', {})
            if corrs:
                print(f"    相関:")
                for k, v in corrs.items():
                    marker = "✅" if abs(v) > 0.3 else "⚠️"
                    print(f"      {marker} {k}: {v:.3f}")
        
        # 馬データ品質
        horse_issues = check_horse_data_quality(sheets, venue_name)
        if horse_issues:
            print(f"\n■ 馬データの品質問題:")
            for issue in horse_issues[:10]:
                print(f"  ⚠️ {issue}")
    
    # 競馬場間比較
    print(f"\n{'='*60}")
    print("📈 競馬場間比較")
    print(f"{'='*60}")
    
    comparisons = compare_venues(all_stats)
    if comparisons:
        for c in comparisons:
            print(f"  {c}")
    else:
        print("  ✅ 競馬場間で大きな指数の乖離は検出されませんでした")
    
    # サマリー
    print(f"\n{'='*60}")
    print("📋 総合評価")
    print(f"{'='*60}")
    
    total_issues = sum(len(s['issues']) for s in all_stats)
    total_t_missing = sum(s['T指数_stats']['missing'] for s in all_stats)
    total_l_missing = sum(s['L指数_stats']['missing'] for s in all_stats)
    total_horses = sum(s['total_horses'] for s in all_stats)
    
    print(f"  分析レース総数: {sum(s['total_races'] for s in all_stats)}")
    print(f"  分析馬総数: {total_horses}")
    print(f"  検出問題数: {total_issues}")
    print(f"  T指数欠損率: {total_t_missing}/{total_horses} ({total_t_missing/max(total_horses,1)*100:.1f}%)")
    print(f"  L指数欠損率: {total_l_missing}/{total_horses} ({total_l_missing/max(total_horses,1)*100:.1f}%)")
    
    # 判定
    if total_issues == 0 and total_t_missing / max(total_horses, 1) < 0.3:
        print(f"\n  ✅ データ品質: 良好")
    elif total_issues < 5:
        print(f"\n  ⚠️ データ品質: 軽微な問題あり")
    else:
        print(f"\n  ❌ データ品質: 要確認")

if __name__ == "__main__":
    main()
