import openpyxl
import sys

# NAR Excel読み込み
path = r'outputs/reports/nar/race_analysis_20260505_船橋.xlsx'
wb = openpyxl.load_workbook(path)

for sn in wb.sheetnames:
    if sn.startswith('Race_'):
        ws = wb[sn]
        headers = [c.value for c in ws[1]]
        
        # 重要カラムのインデックスを探す
        target_cols = ['上がり', '上り秒', '通過', '1C', '2C', '3C', '4C', 'RPCI', 'タイム指数', '上り指数', '馬名']
        col_indices = {}
        for t in target_cols:
            if t in headers:
                col_indices[t] = headers.index(t)
        
        print(f'\n=== {sn} ===')
        print(f'Headers found: {list(col_indices.keys())}')
        
        # 最初の5行のデータを表示
        count = 0
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, 20), values_only=True):
            if count >= 5:
                break
            data = {}
            for col_name, idx in col_indices.items():
                val = row[idx] if idx < len(row) else None
                data[col_name] = val
            if data.get('馬名'):  # 馬名があるデータ行のみ
                print(f'  {data}')
                count += 1
        
        break  # 最初のRaceシートのみ

print('\n--- Done ---')
