import pandas as pd
import requests
from bs4 import BeautifulSoup
import argparse
import os
from datetime import datetime, timedelta
import re
import time
import logging
from typing import Dict, List, Optional, Tuple, Any
import sys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pickle
import socket
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import ScatterChart, Reference, Series, StockChart
from openpyxl.chart.axis import DateAxis, ChartLines
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.updown_bars import UpDownBars
from openpyxl.drawing.line import LineProperties

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)

# Constants
DATA_DIR = r"C:\Users\zk-ht\Keiba\Keiba_AI_v2\keibaai\data"
HTML_DIR = os.path.join(DATA_DIR, "raw", "html")
PARQUET_PATH = os.path.join(DATA_DIR, "parsed", "parquet", "races", "races.parquet")

# Venue Mapping (Netkeiba Code -> Name)
VENUE_MAP = {
    "01": "札幌", "02": "函館", "03": "福島", "04": "新潟", "05": "東京", 
    "06": "中山", "07": "中京", "08": "京都", "09": "阪神", "10": "小倉"
}
VENUE_NAME_TO_CODE = {v: k for k, v in VENUE_MAP.items()}

import numpy as np

class MetricCalculator:
    def __init__(self, reference_df: pd.DataFrame):
        self.reference_df = reference_df
        self.avg_time_lookup = {}
        self.avg_cond_time_lookup = {}
        self.std_3f_lookup = {} # Kept for fallback or reference if needed
        self.std_cond_3f_lookup = {} 
        self.grouped_data = {} # (venue, dist, surf) -> DataFrame
        self.std_time_cache = {} # (venue, dist, surf, date_str) -> val
        self.std_cond_time_cache = {} # (venue, dist, surf, cond, date_str) -> val
        
        self.regression_models = {} # (venue, dist, surf) -> (slope, intercept)
        self.regression_cond_models = {} # (venue, dist, surf, cond) -> (slope, intercept)
        
        self._prepare_lookups()

    def _prepare_lookups(self):
        """Pre-calculates static lookups and groups data."""
        logger.info("Preparing metric lookups...")
        
        # Filter for 1-3rd place once
        df_top3 = self.reference_df[self.reference_df['finish_position'].isin([1, 2, 3])].copy()
        
        # 1. Static Averages (Venue, Dist, Surf)
        avg_time_series = df_top3.groupby(['venue', 'distance_m', 'track_surface'])['finish_time_seconds'].mean()
        self.avg_time_lookup = avg_time_series.to_dict()
        
        # 2. Static Condition Averages (Venue, Dist, Surf, Cond)
        avg_cond_time_series = df_top3.groupby(['venue', 'distance_m', 'track_surface', 'track_condition'])['finish_time_seconds'].mean()
        self.avg_cond_time_lookup = avg_cond_time_series.to_dict()
        
        # 3. Group Data for Time-Dependent Queries & Regression
        for name, group in df_top3.groupby(['venue', 'distance_m', 'track_surface']):
            self.grouped_data[name] = group.sort_values('race_date')
            
            # Regression for Standard 3F
            # X = (Distance - 600) / (Time - 3F)
            # Y = 3F
            self._fit_regression(name, group, is_condition_specific=False)
            
            # Regression for Standard Condition 3F
            for cond, sub_group in group.groupby('track_condition'):
                cond_name = name + (cond,)
                self._fit_regression(cond_name, sub_group, is_condition_specific=True)
            
        logger.info("Lookups prepared.")

    def _fit_regression(self, key, group, is_condition_specific):
        """Fits a linear regression model for 3F prediction."""
        if len(group) < 2:
            return

        try:
            # Calculate X: Average speed of the first part
            # (Distance - 600) / (Finish Time - Last 3F)
            # Ensure no division by zero
            dist_minus_600 = group['distance_m'] - 600
            time_minus_3f = group['finish_time_seconds'] - group['last_3f_time']
            
            valid_mask = (time_minus_3f > 0) & (dist_minus_600 > 0)
            if not valid_mask.any():
                return

            X = (dist_minus_600[valid_mask] / time_minus_3f[valid_mask]).values
            Y = group.loc[valid_mask, 'last_3f_time'].values

            if len(X) < 2:
                return

            # Linear Regression: Y = aX + b
            slope, intercept = np.polyfit(X, Y, 1)
            
            if is_condition_specific:
                self.regression_cond_models[key] = (slope, intercept)
            else:
                self.regression_models[key] = (slope, intercept)
        except Exception as e:
            logger.warning(f"Regression failed for {key}: {e}")

    def get_avg_time(self, venue, dist, surf):
        return self.avg_time_lookup.get((venue, dist, surf))

    def get_avg_cond_time(self, venue, dist, surf, cond):
        return self.avg_cond_time_lookup.get((venue, dist, surf, cond))

    def predict_std_3f(self, venue, dist, surf, finish_time, last_3f):
        """Predicts Standard 3F based on regression."""
        key = (venue, dist, surf)
        model = self.regression_models.get(key)
        if not model:
            return None
        
        slope, intercept = model
        try:
            # Calculate X for the target horse
            dist_minus_600 = dist - 600
            time_minus_3f = finish_time - last_3f
            if time_minus_3f <= 0:
                return None
            
            X = dist_minus_600 / time_minus_3f
            predicted_3f = slope * X + intercept
            return predicted_3f
        except:
            return None

    def predict_std_cond_3f(self, venue, dist, surf, cond, finish_time, last_3f):
        """Predicts Standard Condition 3F based on regression."""
        key = (venue, dist, surf, cond)
        model = self.regression_cond_models.get(key)
        if not model:
            return None
        
        slope, intercept = model
        try:
            dist_minus_600 = dist - 600
            time_minus_3f = finish_time - last_3f
            if time_minus_3f <= 0:
                return None
            
            X = dist_minus_600 / time_minus_3f
            predicted_3f = slope * X + intercept
            return predicted_3f
        except:
            return None

    def get_std_time(self, venue, dist, surf, date_obj):
        """Calculates Standard Time (Avg of +/- 3 days) with caching."""
        date_str = date_obj.strftime('%Y%m%d')
        key = (venue, dist, surf, date_str)
        if key in self.std_time_cache:
            return self.std_time_cache[key]
        
        group = self.grouped_data.get((venue, dist, surf))
        if group is None or group.empty:
            self.std_time_cache[key] = None
            return None
            
        # Filter by date range
        start_date = date_obj - timedelta(days=3)
        end_date = date_obj + timedelta(days=3)
        
        # Since group is sorted, we could use searchsorted, but boolean mask is fast enough on small groups
        mask = (group['race_date'] >= start_date) & (group['race_date'] <= end_date)
        sub_group = group[mask]
        
        if sub_group.empty:
            val = None
        else:
            val = sub_group['finish_time_seconds'].mean()
        
        self.std_time_cache[key] = val
        return val

    def get_std_cond_time(self, venue, dist, surf, cond, date_obj):
        """Calculates Standard Condition Time with caching."""
        date_str = date_obj.strftime('%Y%m%d')
        key = (venue, dist, surf, cond, date_str)
        if key in self.std_cond_time_cache:
            return self.std_cond_time_cache[key]
        
        # We can reuse the group from get_std_time logic, but we need to filter by condition too.
        # It's cleaner to just do the filter here.
        group = self.grouped_data.get((venue, dist, surf))
        if group is None or group.empty:
            self.std_cond_time_cache[key] = None
            return None
            
        start_date = date_obj - timedelta(days=3)
        end_date = date_obj + timedelta(days=3)
        
        mask = (group['race_date'] >= start_date) & (group['race_date'] <= end_date) & (group['track_condition'] == cond)
        sub_group = group[mask]
        
        if sub_group.empty:
            val = None
        else:
            val = sub_group['finish_time_seconds'].mean()
            
        self.std_cond_time_cache[key] = val
        return val

class NetkeibaAnalyzer:
    def __init__(self, target_date: str, venue_name: str):
        self.target_date = target_date  # YYYYMMDD
        self.venue_name = venue_name
        self.venue_code = VENUE_NAME_TO_CODE.get(venue_name)
        if not self.venue_code:
            raise ValueError(f"Unknown venue name: {venue_name}")
        
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        })
        
        self.driver = None
        self.reference_df = None
        self.metric_calculator = None
        self.load_reference_data()
        self.load_stats_lookup()

    def init_driver(self):
        """Initializes Selenium WebDriver."""
        if self.driver:
            return
        logger.info("Initializing Selenium Driver...")
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--disable-extensions')
        options.add_argument("--log-level=3")
        # Use eager strategy to return as soon as DOM is interactive
        options.page_load_strategy = 'eager' 
        self.driver = webdriver.Chrome(options=options)
        self.driver.set_page_load_timeout(15) # Reduce timeout to 15s

    def close_driver(self):
        if self.driver:
            self.driver.quit()
            self.driver = None

    def load_reference_data(self):
        """Loads races.parquet for metric calculations."""
        logger.info(f"Loading reference data from {PARQUET_PATH}...")
        if not os.path.exists(PARQUET_PATH):
            logger.warning(f"Reference data not found at {PARQUET_PATH}. Metrics requiring it will be NaN.")
            return

        try:
            # Optimization: Load only required columns
            required_cols = [
                'race_id', 'horse_id', 'trainer_name', # Added for Stable lookup
                'race_date', 'venue', 'distance_m', 'track_surface', 
                'track_condition', 'finish_position', 'finish_time_seconds', 'last_3f_time'
            ]
            
            df = pd.read_parquet(PARQUET_PATH, columns=required_cols)
            df['race_date'] = pd.to_datetime(df['race_date'])
            # Ensure IDs are strings for lookup
            df['race_id'] = df['race_id'].astype(str)
            df['horse_id'] = df['horse_id'].astype(str)
            
            self.reference_df = df
            logger.info(f"Loaded {len(df)} rows of reference data.")
            
            # Create Trainer Lookup: (race_id, horse_id) -> trainer_name
            self.trainer_lookup = df.set_index(['race_id', 'horse_id'])['trainer_name'].to_dict()
            
            # Initialize Metric Calculator
            self.metric_calculator = MetricCalculator(df)
            
        except Exception as e:
            logger.error(f"Failed to load reference data: {e}")

    def scrape_race_list(self) -> List[str]:
        """Scrapes the race list for the target date and venue to get race_ids using Selenium."""
        self.init_driver()
        url = f"https://race.netkeiba.com/top/race_list.html?kaisai_date={self.target_date}"
        logger.info(f"Scraping race list from {url}")
        
        try:
            self.driver.get(url)
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "RaceList_Data"))
            )
            
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            race_ids = []
            prefix = f"{self.target_date[:4]}{self.venue_code}"
            
            links = soup.find_all('a', href=True)
            for link in links:
                href = link['href']
                rid = None
                if 'race_id=' in href:
                    rid = href.split('race_id=')[1].split('&')[0]
                elif '/race/' in href:
                    parts = href.split('/')
                    for p in parts:
                        if p.isdigit() and len(p) == 12:
                            rid = p
                            break
                
                if rid and rid.startswith(prefix):
                    # Extract Race Name
                    name = link.get_text(strip=True)
                    
                    # Check if already added
                    if not any(r['id'] == rid for r in race_ids):
                        race_ids.append({'id': rid, 'name': name})
            
            race_ids.sort(key=lambda x: x['id'])
            logger.info(f"Found {len(race_ids)} races for {self.venue_name} ({self.venue_code}) on {self.target_date}")
            return race_ids
            
        except Exception as e:
            logger.error(f"Error scraping race list: {e}")
            return []

    def scrape_shutuba(self, race_id: str) -> List[Dict]:
        """Scrapes the Shutuba table for a given race ID using requests (Static HTML)."""
        url = f"https://race.netkeiba.com/race/shutuba.html?race_id={race_id}"
        logger.info(f"Scraping shutuba for {race_id}")
        
        try:
            # Use requests instead of Selenium for stability
            resp = self.session.get(url, timeout=(5, 15))
            resp.encoding = 'euc-jp'
            soup = BeautifulSoup(resp.content, 'html.parser')
            
            race_name_elem = soup.select_one('.RaceName')
            race_name = race_name_elem.text.strip() if race_name_elem else "Unknown Race"
            
            # Check for Shinba (New Horse) immediately
            if '新馬' in race_name:
                logger.info(f"Skipping Race {race_id} ({race_name}) - New Horse Race")
                return []

            horses = []
            rows = soup.select('tr.HorseList')
            
            race_meta = soup.select_one('.RaceData01')
            race_meta_text = race_meta.text.strip() if race_meta else ""
            course_info = ""
            if race_meta_text:
                course_info = race_meta_text.split('/')[0].strip()

            for row in rows:
                horse_data = {
                    'race_id': race_id,
                    'race_name': race_name,
                    'course_info': course_info,
                    'date': self.target_date
                }
                
                waku_elem = row.select_one('td[class^="Waku"]')
                horse_data['枠番'] = waku_elem.text.strip() if waku_elem else ""
                
                umaban_elem = row.select_one('td[class^="Umaban"]')
                horse_data['馬番'] = umaban_elem.text.strip() if umaban_elem else ""
                
                name_elem = row.select_one('.HorseName a')
                if name_elem:
                    horse_data['馬名'] = name_elem.text.strip()
                    href = name_elem['href']
                    if '/horse/' in href:
                            horse_data['horse_id'] = href.split('/horse/')[-1].replace('/', '')
                
                jockey_elem = row.select_one('.Jockey a')
                horse_data['騎手'] = jockey_elem.text.strip() if jockey_elem else ""
                
                trainer_elem = row.select_one('.Trainer a')
                horse_data['厩舎'] = trainer_elem.text.strip() if trainer_elem else ""
                
                tds = row.find_all('td')
                if len(tds) >= 6:
                    horse_data['性齢'] = tds[4].text.strip()
                    horse_data['斤量'] = tds[5].text.strip()

                horses.append(horse_data)
            
            time.sleep(1) # Polite delay
            return horses

        except Exception as e:
            logger.error(f"Error scraping shutuba for {race_id}: {e}")
            return []

    def get_horse_history(self, horse_id: str) -> List[Dict]:
        """Gets past performance for a horse."""
        local_path = os.path.join(HTML_DIR, "horse", f"{horse_id}_perf.bin")
        html_content = None
        
        if os.path.exists(local_path):
            # --- Smart Cache Invalidation ---
            # If cache is older than 7 days, ignore it to force re-scrape
            try:
                mtime = os.path.getmtime(local_path)
                file_time = datetime.fromtimestamp(mtime)
                if datetime.now() - file_time > timedelta(days=7):
                    logger.info(f"Cache for {horse_id} is stale ({file_time.strftime('%Y-%m-%d')}). Re-scraping...")
                    html_content = None # Force re-scrape
                else:
                    with open(local_path, 'rb') as f:
                        html_content = f.read()
            except Exception as e:
                logger.error(f"Error checking/reading local file {local_path}: {e}")
                html_content = None
        
        if not html_content:
            url = f"https://db.netkeiba.com/horse/result/{horse_id}/"
            try:
                # logger.debug(f"Requesting {url}...")
                # Add explicit connect/read timeout
                resp = self.session.get(url, timeout=(5, 15))
                resp.encoding = 'euc-jp'
                html_content = resp.content
                # logger.debug(f"Received {len(html_content)} bytes.")
                time.sleep(0.5)
            except Exception as e:
                logger.error(f"Error scraping history for {horse_id}: {e}")
                return []

        return self.parse_horse_history(html_content, horse_id)

    def parse_horse_history(self, html_content: bytes, horse_id: str) -> List[Dict]:
        """Parses the horse result table with advanced column extraction."""
        try:
            soup = BeautifulSoup(html_content, 'html.parser')
            table = soup.find('table', class_='db_h_race_results')
            if not table:
                return []
            
            history = []
            rows = table.find('tbody').find_all('tr')
            for i, row in enumerate(rows):
                cols = row.find_all('td')
                if len(cols) < 28: # Ensure enough columns
                    # logger.debug(f"Row {i} skipped: Not enough columns ({len(cols)})")
                    continue
                
                data = [c.text.strip() for c in cols]
                # logger.debug(f"Row {i} Date: {data[0]}, Race: {data[4]}")
                
                race_id_full = ""
                race_link = cols[4].find('a')
                if race_link and 'race/' in race_link['href']:
                     race_id_full = race_link['href'].split('/race/')[-1].replace('/', '')

                kai = ""
                day = ""
                race_no = data[3] # R
                if race_id_full and len(race_id_full) == 12:
                    kai = str(int(race_id_full[6:8]))
                    day = str(int(race_id_full[8:10]))
                
                # Corrected Indices based on debug output
                # 16: 馬場, 18: タイム, 19: 着差, 21: 通過, 22: ペース, 23: 上り, 24: 馬体重
                
                passing = data[21]
                p1, p2, p3, p4 = "", "", "", ""
                if passing:
                    parts = passing.split('-')
                    if len(parts) >= 1: p1 = parts[0]
                    if len(parts) >= 2: p2 = parts[1]
                    if len(parts) >= 3: p3 = parts[2]
                    if len(parts) >= 4: p4 = parts[3]
                
                pace = data[22]
                pace1, pace2 = "", ""
                if pace and '-' in pace:
                    pp = pace.split('-')
                    if len(pp) == 2:
                        pace1, pace2 = pp[0], pp[1]
                
                weight_raw = data[24]
                weight = ""
                change = ""
                if weight_raw:
                    match = re.match(r'(\d+)\((.+)\)', weight_raw)
                    if match:
                        weight = match.group(1)
                        change = match.group(2)
                    else:
                        weight = weight_raw

                # Lookup Trainer
                trainer = ""
                if hasattr(self, 'trainer_lookup') and race_id_full:
                     key = (str(race_id_full), str(horse_id))
                     trainer = self.trainer_lookup.get(key, "")
                     # Debug logging for first few lookups
                     if not trainer and len(history) < 3:
                         logger.info(f"Lookup failed for {key}. Sample keys: {list(self.trainer_lookup.keys())[:3]}")

                race_data = {
                    '日付': data[0],
                    '場所': "".join([c for c in data[1] if not c.isdigit()]),
                    '回': kai,
                    '日': day,
                    'ｺｰｽ': "芝" if "芝" in data[14] else "ダート" if "ダ" in data[14] else "障害" if "障" in data[14] else "",
                    '距離': re.search(r'\d+', data[14]).group() if re.search(r'\d+', data[14]) else "",
                    'R': race_no,
                    '馬場': data[16],
                    '天気': data[2],
                    '頭数': data[6],
                    '枠番': data[7],
                    '馬番': data[8],
                    '馬名': "",
                    '斤量': data[13],
                    '騎手': data[12],
                    '厩舎': trainer,
                    'ﾀｲﾑ': data[18],
                    '着差': data[19],
                    '人気': data[10],
                    'ｵｯｽﾞ': data[9],
                    '上り': data[23],
                    'ﾍﾟｰｽ1': pace1,
                    'ﾍﾟｰｽ2': pace2,
                    '通過': passing,
                    '1C': p1, '2C': p2, '3C': p3, '4C': p4,
                    '着順': data[11],
                    '馬体重': weight,
                    '増減': change,
                    'レース名': data[4],
                    '性': "",
                    '年齢': "",
                    'horse_id': "",
                    'race_id': race_id_full
                }
                history.append(race_data)
                
            return history
        except Exception as e:
            logger.error(f"Error parsing history: {e}")
            return []

    def load_stats_lookup(self):
        """Load pre-calculated statistics lookup table."""
        path = r"keibaai\data\processed\stats_lookup.pkl"
        if os.path.exists(path):
            try:
                with open(path, 'rb') as f:
                    self.stats_lookup = pickle.load(f)
                logger.info(f"Loaded stats lookup with {len(self.stats_lookup)} conditions.")
            except Exception as e:
                logger.error(f"Failed to load stats lookup: {e}")
                self.stats_lookup = {}
        else:
            logger.warning("stats_lookup.pkl not found. Statistical metrics will be empty.")
            self.stats_lookup = {}

    def calculate_metrics(self, history: List[Dict]) -> List[Dict]:
        """Calculates statistical metrics for each race in history."""
        if not hasattr(self, 'stats_lookup'):
            self.load_stats_lookup()

        for race in history:
            venue = race.get('場所', '')
            dist_str = race.get('距離', '')
            cond = race.get('馬場', '')
            time_str = race.get('ﾀｲﾑ', '')
            l3f_str = race.get('上り', '')
            pace_str = race.get('ﾍﾟｰｽ1', '')
            
            # Parse Distance and Surface
            surface = race.get('ｺｰｽ', '')
            
            try:
                distance = int(dist_str) if dist_str.isdigit() else 0
            except:
                distance = 0
            
            # Parse Time
            def parse_time(t_str):
                try:
                    if not t_str: return None
                    parts = t_str.split(':')
                    if len(parts) == 2:
                        return float(parts[0]) * 60 + float(parts[1])
                    return float(t_str)
                except:
                    return None
            
            time_sec = parse_time(time_str)
            l3f_sec = parse_time(l3f_str)
            
            # Initialize metrics with None
            for col in ['タイム指数', '上り指数', '馬場差', 'RPCI',
                       '平均t', '平均場t', '基準t', '基準場t', '基準3F', '基準場3F', 
                       '平t差', '平場t差', '基t差', '基場t差', '基3F差', '基場3F差']:
                race[col] = None
            
            # Parse Date
            try:
                date_obj = pd.to_datetime(race.get('日付', ''))
            except:
                date_obj = None

            # Map abbreviated condition to full string
            cond_map = {"良": "良", "稍": "稍重", "重": "重", "不": "不良"}
            full_cond = cond_map.get(cond, cond)

            # --- Calculate Standard/Average Metrics using MetricCalculator ---
            if self.metric_calculator and date_obj and distance > 0 and time_sec:
                # 1. Average Time (Static)
                avg_t = self.metric_calculator.get_avg_time(venue, distance, surface)
                if avg_t:
                    race['平均t'] = round(avg_t, 1)
                    race['平t差'] = round(time_sec - avg_t, 1)

                # 2. Average Condition Time (Static)
                avg_cond_t = self.metric_calculator.get_avg_cond_time(venue, distance, surface, full_cond)
                if avg_cond_t:
                    race['平均場t'] = round(avg_cond_t, 1)
                    race['平場t差'] = round(time_sec - avg_cond_t, 1)

                # 3. Standard Time (Dynamic - +/- 3 days)
                std_t = self.metric_calculator.get_std_time(venue, distance, surface, date_obj)
                if std_t:
                    race['基準t'] = round(std_t, 1)
                    race['基t差'] = round(time_sec - std_t, 1)
                
                # Calculate Track Bias (馬場差) = Standard Time - Average Time
                if avg_t and std_t:
                    race['馬場差'] = round(std_t - avg_t, 1)

                # 4. Standard Condition Time (Dynamic)
                std_cond_t = self.metric_calculator.get_std_cond_time(venue, distance, surface, full_cond, date_obj)
                if std_cond_t:
                    race['基準場t'] = round(std_cond_t, 1)
                    race['基場t差'] = round(time_sec - std_cond_t, 1)
                
                # 5. Standard 3F (Regression)
                if l3f_sec:
                    std_3f = self.metric_calculator.predict_std_3f(venue, distance, surface, time_sec, l3f_sec)
                    if std_3f:
                        race['基準3F'] = round(std_3f, 1)
                        race['基3F差'] = round(l3f_sec - std_3f, 1)
                        
                    std_cond_3f = self.metric_calculator.predict_std_cond_3f(venue, distance, surface, full_cond, time_sec, l3f_sec)
                    if std_cond_3f:
                        race['基準場3F'] = round(std_cond_3f, 1)
                        race['基場3F差'] = round(l3f_sec - std_cond_3f, 1)

            # Lookup Stats (Time Index, L3F Index, RPCI)
            key = (venue, distance, surface, full_cond)
            stats = self.stats_lookup.get(key)
            
            if stats and time_sec:
                # Time Index: 50 + 10 * (Mean - Time) / Std
                mean = stats['time_mean']
                std = stats['time_std']
                if std > 0:
                    idx = 50 + 10 * (mean - time_sec) / std
                    race['タイム指数'] = round(idx, 1)
                    
            if stats and l3f_sec:
                # L3F Index
                mean_3f = stats['l3f_mean']
                std_3f = stats['l3f_std']
                if std_3f > 0:
                    idx_3f = 50 + 10 * (mean_3f - l3f_sec) / std_3f
                    race['上り指数'] = round(idx_3f, 1)

            # RPCI: (First 3F / Last 3F) * 50
            if pace_str and l3f_sec:
                try:
                    first_3f = float(pace_str)
                    if first_3f > 0 and l3f_sec > 0:
                        rpci = (first_3f / l3f_sec) * 50
                        race['RPCI'] = round(rpci, 1)
                except:
                    pass
                    
        return history

    def run(self):
        try:
            races = self.scrape_race_list()
            if not races:
                logger.error("No races found.")
                return

            all_races_data = {}

            for race_info in races:
                race_id = race_info['id']
                race_name = race_info['name']
                
                # Skip Shinba (New Horse) races EARLY
                if '新馬' in race_name:
                    logger.info(f"Skipping Race {race_id} ({race_name}) - New Horse Race (Skipped before scraping)")
                    continue
                
                logger.info(f"Processing Race {race_id} ({race_name})...")
                shutuba = self.scrape_shutuba(race_id)
                
                if not shutuba:
                    continue
                
                # --- 1. Collect all history first for Grouping ---
                race_horses_data = [] 
                all_history_entries = [] 
                
                total_horses = len(shutuba)
                for idx, horse in enumerate(shutuba, 1):
                    horse_id = horse.get('horse_id')
                    horse_name = horse.get('馬名', 'Unknown')
                    if not horse_id:
                        continue
                    
                    logger.info(f"[{idx}/{total_horses}] Fetching history for {horse_name} ({horse_id})...")
                    history = self.get_horse_history(horse_id)
                    
                    # --- FILTER: Only keep history from 2024/10/1 onwards ---
                    filter_date = pd.to_datetime("2024-10-01")
                    filtered_history = []
                    for h in history:
                        try:
                            h_date = pd.to_datetime(h.get('日付', ''))
                            if h_date >= filter_date:
                                filtered_history.append(h)
                        except:
                            pass
                    history = filtered_history
                    # --------------------------------------------------------

                    history = self.calculate_metrics(history)
                    
                    race_horses_data.append({
                        'horse_info': horse,
                        'history': history
                    })
                    
                    for i, h_row in enumerate(history):
                        all_history_entries.append({
                            'horse_id': horse_id,
                            'row': h_row,
                            'date': pd.to_datetime(h_row['日付']),
                            'venue': h_row['場所'],
                            'dist': h_row['距離'],
                            'surf': h_row['馬場']
                        })

                # --- 2. Calculate GroupNo (Clustering) ---
                grouped_entries = {}
                for entry in all_history_entries:
                    key = (entry['venue'], entry['dist'])
                    if key not in grouped_entries:
                        grouped_entries[key] = []
                    grouped_entries[key].append(entry)
                
                next_group_id = 1
                
                for key, entries in grouped_entries.items():
                    entries.sort(key=lambda x: x['date'])
                    
                    if not entries:
                        continue
                        
                    current_cluster = [entries[0]]
                    
                    for i in range(1, len(entries)):
                        prev = entries[i-1]
                        curr = entries[i]
                        diff = (curr['date'] - prev['date']).days
                        
                        if diff <= 3:
                            current_cluster.append(curr)
                        else:
                            unique_horses = set(e['horse_id'] for e in current_cluster)
                            if len(unique_horses) >= 2:
                                for e in current_cluster:
                                    e['row']['グループNo'] = next_group_id
                                next_group_id += 1
                            current_cluster = [curr]
                    
                    unique_horses = set(e['horse_id'] for e in current_cluster)
                    if len(unique_horses) >= 2:
                        for e in current_cluster:
                            e['row']['グループNo'] = next_group_id
                        next_group_id += 1

                # --- 3. Finalize Data & Jockey Change ---
                race_data_list = []
                target_date_obj = pd.to_datetime(self.target_date)
                
                for item in race_horses_data:
                    horse = item['horse_info']
                    history = item['history']
                    horse_id = horse.get('horse_id')
                    horse_name = horse.get('馬名', 'Unknown')
                    current_jockey = horse.get('騎手', '')
                    
                    prev_jockey = ""
                    for h_row in history:
                        h_date = pd.to_datetime(h_row['日付'])
                        if h_date < target_date_obj:
                            prev_jockey = h_row.get('騎手', '')
                            break
                    
                    jockey_change = "〇" if prev_jockey and current_jockey != prev_jockey else "-"
                    
                    for h_row in history:
                        h_row['馬名'] = horse_name
                        h_row['horse_id'] = horse_id
                        h_row['出走枠番'] = horse.get('枠番', '')
                        h_row['出走馬番'] = horse.get('馬番', '')
                        h_row['乗り替わり'] = jockey_change
                        if 'グループNo' not in h_row:
                            h_row['グループNo'] = ""
                        
                        current_sex_age = horse.get('性齢', '')
                        if current_sex_age and len(current_sex_age) >= 2:
                            sex = current_sex_age[0]
                            try:
                                current_age = int(re.search(r'\d+', current_sex_age).group())
                                race_date = pd.to_datetime(h_row['日付'])
                                year_diff = target_date_obj.year - race_date.year
                                hist_age = current_age - year_diff
                                h_row['性'] = sex
                                h_row['年齢'] = str(hist_age) if hist_age > 0 else ""
                            except:
                                pass

                    race_data_list.append(item)
                
                all_races_data[race_id] = race_data_list
            
            self.generate_excel(all_races_data)
        finally:
            self.close_driver()

    def generate_excel(self, all_data: Dict):
        output_file = f"race_analysis_{self.target_date}_{self.venue_name}.xlsx"
        logger.info(f"Generating Excel: {output_file}")
        
        columns = [
            '出走枠番', '出走馬番', '乗り替わり', 'グループNo',
            '日付', '場所', '回', '日', 'ｺｰｽ', '距離', 'R', '馬場', '天気', '頭数', 
            '枠番', '馬番', '馬名', '斤量', '騎手', '厩舎', 'ﾀｲﾑ', '着差', '人気', 'ｵｯｽﾞ', 
            '上り', 'ﾍﾟｰｽ1', 'ﾍﾟｰｽ2', '通過', '1C', '2C', '3C', '4C', '着順', '馬体重', '増減', 
            'レース名', '性', '年齢', 
            '平均t', '平均場t', '基準t', '基準場t', '基準3F', '基準場3F', 
            '平t差', '平場t差', '基t差', '基場t差', '基3F差', '基場3F差', 
            'タイム指数', '上り指数', '馬場差', 'RPCI',
            'horse_id', 'race_id'
        ]
        
        # Waku Color Definitions (Standard JRA Colors)
        WAKU_COLORS = {
            1: {'bg': 'FFFFFF', 'fg': '000000'}, # White
            2: {'bg': '000000', 'fg': 'FFFFFF'}, # Black
            3: {'bg': 'FF0000', 'fg': 'FFFFFF'}, # Red
            4: {'bg': '0000FF', 'fg': 'FFFFFF'}, # Blue
            5: {'bg': 'FFFF00', 'fg': '000000'}, # Yellow
            6: {'bg': '008000', 'fg': 'FFFFFF'}, # Green
            7: {'bg': 'FFA500', 'fg': '000000'}, # Orange
            8: {'bg': 'FFC0CB', 'fg': '000000'}  # Pink
        }

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            for race_id, horses_data in all_data.items():
                sheet_name = f"Race_{race_id[-2:]}"
                
                # Sort horses by Umaban for consistent chart order
                try:
                    horses_data.sort(key=lambda x: int(x['horse_info'].get('馬番', 999)))
                except:
                    pass

                rows = []
                # Header Row
                rows.append(columns)
                
                for h_data in horses_data:
                    for hist in h_data['history']:
                        row_vals = [hist.get(col, '') for col in columns]
                        rows.append(row_vals)
                
                df_out = pd.DataFrame(rows[1:], columns=rows[0])
                
                # Convert numeric columns to numbers (float/int) for sorting/filtering
                numeric_cols = [
                    '枠番', '馬番', '斤量', '着順', '馬体重', '増減', '年齢',
                    '平均t', '平均場t', '基準t', '基準場t', '基準3F', '基準場3F',
                    '平t差', '平場t差', '基t差', '基場t差', '基3F差', '基場3F差',
                    'タイム指数', '上り指数', '馬場差', 'RPCI', '上り', 'ﾀｲﾑ', 'ｵｯｽﾞ', '人気',
                    'ﾍﾟｰｽ1', 'ﾍﾟｰｽ2', '4C'
                ]
                
                for col in numeric_cols:
                    if col in df_out.columns:
                        df_out[col] = pd.to_numeric(df_out[col], errors='coerce')

                df_out.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # --- Formatting & Charts ---
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # --- Create Hidden Chart Data Sheet ---
                # This sheet ensures charts remain stable even if the main sheet is sorted/filtered.
                chart_sheet_name = f"ChartData_{race_id[-2:]}"
                chart_sheet = workbook.create_sheet(title=chart_sheet_name)
                chart_sheet.sheet_state = 'hidden' # Hide from user
                
                # Write Header
                chart_cols = ['馬番', '馬名', 'タイム指数', '上り指数', '馬場差', 'RPCI', 'ﾍﾟｰｽ1', '4C', '枠番']
                chart_sheet.append(chart_cols)
                
                # Write Data (Sorted by Umaban)
                chart_data_start_row = 2
                current_chart_row = chart_data_start_row
                
                # Track row ranges for each horse to create series
                horse_row_ranges = {} # {horse_index: (start_row, end_row)}
                
                for h_idx, h_data in enumerate(horses_data):
                    horse_info = h_data['horse_info']
                    h_name = horse_info.get('馬名', 'Unknown')
                    umaban = horse_info.get('馬番', '')
                    waku = horse_info.get('枠番', '')
                    
                    start_r = current_chart_row
                    
                    for hist in h_data['history']:
                        # Extract values, ensuring they are numeric
                        def get_val(key):
                            val = hist.get(key, '')
                            try:
                                return float(val)
                            except:
                                return None
                        
                        row_vals = [
                            umaban,
                            h_name,
                            get_val('タイム指数'),
                            get_val('上り指数'),
                            get_val('馬場差'),
                            get_val('RPCI'),
                            get_val('ﾍﾟｰｽ1'),
                            get_val('4C'),
                            waku
                        ]
                        chart_sheet.append(row_vals)
                        current_chart_row += 1
                    
                    end_r = current_chart_row - 1
                    if end_r >= start_r:
                        horse_row_ranges[h_idx] = (start_r, end_r)

                # Helper to find column index in Chart Sheet (1-based)
                def get_chart_col_idx(name):
                    try:
                        return chart_cols.index(name) + 1
                    except ValueError:
                        return None
                
                idx_c_time = get_chart_col_idx('タイム指数')
                idx_c_l3f = get_chart_col_idx('上り指数')
                idx_c_bias = get_chart_col_idx('馬場差')
                idx_c_rpci = get_chart_col_idx('RPCI')
                idx_c_pace = get_chart_col_idx('ﾍﾟｰｽ1')
                idx_c_4c = get_chart_col_idx('4C')

                # Helper to find column index in Main Sheet (for Waku coloring on main sheet)
                def get_col_idx(name):
                    try:
                        return columns.index(name) + 1
                    except ValueError:
                        return None

                idx_waku = get_col_idx('出走枠番')
                idx_umaban = get_col_idx('出走馬番')
                
                # --- Apply Waku Colors (Main Sheet) ---
                if idx_waku:
                    for row in range(2, len(df_out) + 2):
                        cell = worksheet.cell(row=row, column=idx_waku)
                        try:
                            waku_val = int(cell.value)
                            if waku_val in WAKU_COLORS:
                                style = WAKU_COLORS[waku_val]
                                cell.fill = PatternFill(start_color=style['bg'], end_color=style['bg'], fill_type='solid')
                                cell.font = Font(color=style['fg'], bold=True)
                                cell.alignment = Alignment(horizontal='center')
                                
                                # Apply to Umaban as well if present
                                if idx_umaban:
                                    cell_umaban = worksheet.cell(row=row, column=idx_umaban)
                                    cell_umaban.fill = PatternFill(start_color=style['bg'], end_color=style['bg'], fill_type='solid')
                                    cell_umaban.font = Font(color=style['fg'], bold=True)
                                    cell_umaban.alignment = Alignment(horizontal='center')
                        except:
                            pass

                # --- Add Charts to Separate Sheet ---
                # Check if we have enough data columns in ChartData sheet
                if idx_c_time and idx_c_l3f and idx_c_bias:
                    analysis_sheet_name = f"Analysis_{race_id[-2:]}"
                    analysis_sheet = workbook.create_sheet(title=analysis_sheet_name)
                    
                    # --- Explanatory Text ---
                    # Helper to add text box
                    def add_explanation(row, col, title, text):
                        cell = analysis_sheet.cell(row=row, column=col)
                        cell.value = title
                        cell.font = Font(bold=True, size=11)
                        
                        desc_cell = analysis_sheet.cell(row=row+1, column=col)
                        desc_cell.value = text
                        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')
                        
                        # Merge for description
                        analysis_sheet.merge_cells(start_row=row+1, start_column=col, end_row=row+5, end_column=col+8)
                        
                    # Common Chart Setup
                    def setup_chart(title, x_title, y_title):
                        chart = ScatterChart()
                        chart.title = title
                        chart.style = 2
                        chart.x_axis.title = x_title
                        chart.y_axis.title = y_title
                        chart.height = 10 # cm (Reduced from 15)
                        chart.width = 16  # cm
                        chart.legend = None # Remove Legend
                        
                        # Gridlines Style (Light Gray)
                        sp_pr = GraphicalProperties(ln=LineProperties(solidFill="D9D9D9"))
                        chart.x_axis.majorGridlines = ChartLines(spPr=sp_pr)
                        chart.y_axis.majorGridlines = ChartLines(spPr=sp_pr)
                        
                        return chart

                    # 1. Speed vs Stamina (X=L3F Index, Y=Time Index)
                    chart1 = setup_chart("スピード vs スタミナ", "上り指数", "タイム指数")
                    
                    # Place Text BELOW Chart 1 (Row 23)
                    add_explanation(23, 2, "【スピード vs スタミナ】", 
                                    "・縦軸：走破タイムの優秀さ（上が速い）\n"
                                    "・横軸：末脚の鋭さ（右が速い）\n"
                                    "★右上に位置する馬ほど、総合的な能力が高いと言えます。")

                    # 2. Track Suitability (X=Track Bias, Y=Time Index)
                    chart2 = setup_chart("馬場適性", "馬場差", "タイム指数")
                    
                    # Place Text BELOW Chart 2 (Row 51)
                    add_explanation(51, 2, "【馬場適性】", 
                                    "・縦軸：走破タイムの優秀さ\n"
                                    "・横軸：馬場の重さ（右が重い/タフ、左が軽い/高速）\n"
                                    "★今回の馬場状態（予想）に近い位置で好走している馬を探します。")

                    # 3. Pace Suitability (X=Pace, Y=Time Index)
                    chart3 = setup_chart("ペース適性", "前半3F (秒)", "タイム指数")
                    
                    # Place Text BELOW Chart 3 (Row 23, Col M=13)
                    add_explanation(23, 13, "【ペース適性】", 
                                    "・縦軸：走破タイムの優秀さ\n"
                                    "・横軸：前半3Fタイム（左がハイペース、右がスロー）\n"
                                    "★今回の展開予想（H/M/S）に近いペースで好走している馬が狙い目です。")

                    # 4. Running Style (X=4C Pos, Y=L3F Index)
                    chart4 = setup_chart("脚質・末脚", "4コーナー通過順", "上り指数")
                    
                    # Place Text BELOW Chart 4 (Row 51, Col M=13)
                    add_explanation(51, 13, "【脚質・末脚】", 
                                    "・縦軸：末脚の鋭さ（上が速い）\n"
                                    "・横軸：4コーナー位置取り（左が前、右が後）\n"
                                    "★左上：逃げて速い上がり（最強）\n"
                                    "★右上：追い込み一気\n"
                                    "★左下：粘り込み")

                    # 5. Pace Aptitude (X=RPCI, Y=Time Index)
                    chart5 = setup_chart("展開適性 (RPCI vs タイム)", "RPCI", "タイム指数")
                    
                    # Place Text BELOW Chart 5 (Row 79)
                    add_explanation(79, 2, "【展開適性 (RPCI)】", 
                                    "・縦軸：走破タイムの優秀さ\n"
                                    "・横軸：ペース配分（右がスロー/瞬発戦、左がハイ/消耗戦）\n"
                                    "★50が平均。50以上はハイペース、50以下はスローペース。\n"
                                    "★今回の展開予想（H/M/S）に合った位置で好走している馬を探します。")

                    # 6. Pace vs L3F (X=RPCI, Y=L3F Index)
                    chart6 = setup_chart("展開別末脚 (RPCI vs 上り)", "RPCI", "上り指数")
                    
                    # Place Text BELOW Chart 6 (Row 79, Col M=13)
                    add_explanation(79, 13, "【展開別末脚】", 
                                    "・縦軸：末脚の鋭さ\n"
                                    "・横軸：ペース配分（右がスロー、左がハイ）\n"
                                    "★スローなら切れるがハイペースだと不発、などの特性を見抜きます。")

                    # Create Series for each horse using ChartData sheet
                    for h_idx, h_data in enumerate(horses_data):
                        if h_idx in horse_row_ranges:
                            start_row, end_row = horse_row_ranges[h_idx]
                            
                            horse_info = h_data['horse_info']
                            horse_name = horse_info.get('馬名', 'Unknown')
                            waku = horse_info.get('枠番', '')
                            umaban = horse_info.get('馬番', '')
                            
                            # Legend Name: (Umaban) HorseName
                            legend_name = f"({umaban}) {horse_name}" if umaban else horse_name
                            
                            # Determine Color
                            marker_color = "808080" # Default Gray
                            try:
                                waku_int = int(waku)
                                if waku_int in WAKU_COLORS:
                                    marker_color = WAKU_COLORS[waku_int]['bg']
                            except:
                                pass

                            # Common Series Props
                            def create_series(x_col, y_col, title, color):
                                # Reference ChartData sheet
                                vals_x = Reference(chart_sheet, min_col=x_col, min_row=start_row, max_row=end_row)
                                vals_y = Reference(chart_sheet, min_col=y_col, min_row=start_row, max_row=end_row)
                                s = Series(vals_y, vals_x, title=title)
                                s.marker.symbol = "circle"
                                s.marker.size = 10
                                s.graphicalProperties.line.noFill = True 
                                s.marker.graphicalProperties.solidFill = color
                                s.marker.graphicalProperties.line.solidFill = "000000"
                                return s

                            # Chart 1: L3F vs Time
                            chart1.series.append(create_series(idx_c_l3f, idx_c_time, legend_name, marker_color))
                            
                            # Chart 2: Bias vs Time
                            chart2.series.append(create_series(idx_c_bias, idx_c_time, legend_name, marker_color))
                            
                            # Chart 3: Pace vs Time
                            if idx_c_pace:
                                chart3.series.append(create_series(idx_c_pace, idx_c_time, legend_name, marker_color))
                                
                            # Chart 4: 4C vs L3F
                            if idx_c_4c:
                                chart4.series.append(create_series(idx_c_4c, idx_c_l3f, legend_name, marker_color))

                            # Chart 5: RPCI vs Time
                            if idx_c_rpci:
                                chart5.series.append(create_series(idx_c_rpci, idx_c_time, legend_name, marker_color))
                                
                            # Chart 6: RPCI vs L3F
                            if idx_c_rpci:
                                chart6.series.append(create_series(idx_c_rpci, idx_c_l3f, legend_name, marker_color))
                    
                    # Position Charts on Analysis Sheet
                    if len(chart1.series) > 0: analysis_sheet.add_chart(chart1, "B2")
                    if len(chart3.series) > 0: analysis_sheet.add_chart(chart3, "M2")
                    if len(chart2.series) > 0: analysis_sheet.add_chart(chart2, "B30")
                    if len(chart4.series) > 0: analysis_sheet.add_chart(chart4, "M30")
                    if len(chart5.series) > 0: analysis_sheet.add_chart(chart5, "B58")
                    if len(chart6.series) > 0: analysis_sheet.add_chart(chart6, "M58")

                    # --- Chart 7: Time Index Distribution (Box Plot Proxy using Stacked Bar + Error Bars) ---
                    
                    # Write Data for Chart (Hidden area, e.g., Col Z)
                    stock_data_col = 26 # Z
                    stock_data_row = 2
                    
                    # Headers
                    analysis_sheet.cell(row=stock_data_row, column=stock_data_col, value="Horse")
                    analysis_sheet.cell(row=stock_data_row, column=stock_data_col+1, value="Base(Q1)")
                    analysis_sheet.cell(row=stock_data_row, column=stock_data_col+2, value="Box(IQR)")
                    analysis_sheet.cell(row=stock_data_row, column=stock_data_col+3, value="ErrPlus(Max-Q3)")
                    analysis_sheet.cell(row=stock_data_row, column=stock_data_col+4, value="ErrMinus(Q1-Min)")
                    
                    stock_data_row += 1
                    start_data_row = stock_data_row
                    
                    # Iterate through ALL horses to ensure X-axis is complete
                    for h_data in horses_data:
                        horse_info = h_data['horse_info']
                        horse_name = horse_info.get('馬名', 'Unknown')
                        umaban = horse_info.get('馬番', '')
                        label = f"{umaban}" if umaban else horse_name[:3]
                        
                        time_indices = []
                        for h in h_data['history']:
                            if h.get('タイム指数'):
                                try:
                                    time_indices.append(float(h['タイム指数']))
                                except:
                                    pass
                        
                        q1 = 0
                        iqr = 0
                        err_plus = 0
                        err_minus = 0
                        
                        if len(time_indices) >= 3:
                            import statistics
                            try:
                                quantiles = statistics.quantiles(time_indices, n=4)
                                q1 = quantiles[0]
                                q3 = quantiles[2]
                                min_val = min(time_indices)
                                max_val = max(time_indices)
                                
                                iqr = q3 - q1
                                err_plus = max_val - q3
                                err_minus = q1 - min_val
                            except:
                                pass
                        
                        # Write to sheet (Always write, even if 0)
                        analysis_sheet.cell(row=stock_data_row, column=stock_data_col, value=label)
                        analysis_sheet.cell(row=stock_data_row, column=stock_data_col+1, value=q1)      # Base
                        analysis_sheet.cell(row=stock_data_row, column=stock_data_col+2, value=iqr)     # Box
                        analysis_sheet.cell(row=stock_data_row, column=stock_data_col+3, value=err_plus) # Top Whisker
                        analysis_sheet.cell(row=stock_data_row, column=stock_data_col+4, value=err_minus)# Bottom Whisker
                        
                        stock_data_row += 1

                    end_data_row = stock_data_row - 1
                    
                    from openpyxl.chart import BarChart
                    from openpyxl.chart.error_bar import ErrorBars
                    from openpyxl.chart.data_source import NumDataSource, NumRef
                    from openpyxl.utils import get_column_letter
                    
                    # Create Stacked Bar Chart
                    chart7 = BarChart()
                    chart7.type = "col"
                    chart7.grouping = "stacked"
                    chart7.title = "タイム指数 分布 (箱ひげ図)"
                    chart7.height = 10
                    chart7.width = 30
                    
                    # Categories (X)
                    cats = Reference(analysis_sheet, min_col=stock_data_col, min_row=start_data_row, max_row=end_data_row)
                    chart7.set_categories(cats)
                    
                    # Series 1: Base (Q1)
                    base_data = Reference(analysis_sheet, min_col=stock_data_col+1, min_row=start_data_row, max_row=end_data_row)
                    s1 = Series(base_data, title="Base")
                    s1.graphicalProperties.noFill = True # Invisible
                    s1.graphicalProperties.line.noFill = True
                    
                    # Add Minus Error Bar to Base (Bottom Whisker)
                    minus_col_letter = get_column_letter(stock_data_col+4) # AD
                    minus_ref = f"'Analysis_{race_id}'!${minus_col_letter}${start_data_row}:${minus_col_letter}${end_data_row}"
                    
                    # Create Zero Column for Plus direction (to avoid unwanted whiskers)
                    zero_col = stock_data_col + 5
                    analysis_sheet.cell(row=start_data_row-1, column=zero_col, value="Zero")
                    for r in range(start_data_row, end_data_row+1):
                        analysis_sheet.cell(row=r, column=zero_col, value=0)
                    zero_col_letter = get_column_letter(zero_col)
                    zero_ref = f"'Analysis_{race_id}'!${zero_col_letter}${start_data_row}:${zero_col_letter}${end_data_row}"

                    s1.errBars = ErrorBars(errDir='y', errValType='cust', 
                                           minus=NumDataSource(numRef=NumRef(f=minus_ref)),
                                           plus=NumDataSource(numRef=NumRef(f=zero_ref)))
                    # Explicitly set line color to Black
                    s1.errBars.spPr = GraphicalProperties(ln=LineProperties(solidFill="000000", w=12700)) # 1pt width
                    
                    chart7.series.append(s1)
                    
                    # Series 2: Box (IQR)
                    box_data = Reference(analysis_sheet, min_col=stock_data_col+2, min_row=start_data_row, max_row=end_data_row)
                    s2 = Series(box_data, title="Box")
                    s2.graphicalProperties.solidFill = "CCCCCC" # Gray Box
                    s2.graphicalProperties.line.solidFill = "000000" # Black Border
                    
                    # S2 Error Bars (Plus Only)
                    plus_col_letter = get_column_letter(stock_data_col+3) # AC
                    plus_ref = f"'Analysis_{race_id}'!${plus_col_letter}${start_data_row}:${plus_col_letter}${end_data_row}"
                    
                    s2.errBars = ErrorBars(errDir='y', errValType='cust',
                                           plus=NumDataSource(numRef=NumRef(f=plus_ref)),
                                           minus=NumDataSource(numRef=NumRef(f=zero_ref)))
                    # Explicitly set line color to Black
                    s2.errBars.spPr = GraphicalProperties(ln=LineProperties(solidFill="000000", w=12700))
                    
                    chart7.series.append(s2)
                    
                    # Axis Titles
                    chart7.y_axis.title = "タイム指数"
                    chart7.x_axis.title = "馬番"
                    
                    # Add to sheet
                    analysis_sheet.add_chart(chart7, "B86")
                    
                    add_explanation(107, 2, "【タイム指数分布】", 
                                    "・各馬のタイム指数のバラつきを表示（箱ひげ図）。\n"
                                    "・箱（グレー）: 指数の中心的な範囲（Q1〜Q3）。\n"
                                    "・ヒゲ（上下の線）: 最大値と最小値。\n"
                                    "★箱が高い位置にあり、かつ小さい（安定している）馬が信頼できます。")

                    # --- 3. Time Matchup Table (Grouped by Condition & Period) ---
                    # Goal: Compare horses in similar conditions (Same Venue, Dist, Surf, Date +/- 3 days)
                    
                    # 1. Collect all history entries
                    all_entries = []
                    for h_data in horses_data:
                        horse_info = h_data['horse_info']
                        h_name = horse_info.get('馬名', 'Unknown')
                        umaban = horse_info.get('馬番', '')
                        display_name = f"({umaban}) {h_name}" if umaban else h_name
                        waku = horse_info.get('枠番', '')
                        
                        for hist in h_data['history']:
                            # Parse Date
                            try:
                                d_str = hist.get('日付', '')
                                # Format: 2024/10/20
                                date_obj = datetime.strptime(d_str, '%Y/%m/%d')
                            except:
                                continue
                                
                            entry = {
                                'horse_name': display_name,
                                'waku': waku,
                                'date': date_obj,
                                'date_str': d_str,
                                'place': hist.get('場所', ''),
                                'course': hist.get('ｺｰｽ', ''), # e.g. 芝
                                'dist': hist.get('距離', ''),   # e.g. 2000
                                'cond': hist.get('馬場', ''),
                                'race_name': hist.get('レース名', ''),
                                'rank': hist.get('着順', 99),
                                'time': hist.get('ﾀｲﾑ', ''),
                                'diff': hist.get('着差', ''),
                                '3f': hist.get('上り', ''),
                                'passing': hist.get('通過', '')
                            }
                            try:
                                entry['rank'] = int(entry['rank'])
                            except:
                                entry['rank'] = 99
                                
                            # Parse Time to Seconds
                            try:
                                t_str = entry['time']
                                if t_str:
                                    parts = t_str.split(':')
                                    if len(parts) == 2:
                                        entry['time_sec'] = float(parts[0]) * 60 + float(parts[1])
                                    else:
                                        entry['time_sec'] = float(t_str)
                                else:
                                    entry['time_sec'] = None
                            except:
                                entry['time_sec'] = None
                                
                            all_entries.append(entry)
                    
                    # 2. Group by (Place, Dist, Course)
                    grouped_candidates = {}
                    for e in all_entries:
                        key = (e['place'], e['dist'], e['course'])
                        if key not in grouped_candidates:
                            grouped_candidates[key] = []
                        grouped_candidates[key].append(e)
                        
                    # 3. Cluster by Date (within 3 days) -> Assign GroupNo
                    valid_groups = []
                    
                    for key, entries in grouped_candidates.items():
                        # Sort by date
                        entries.sort(key=lambda x: x['date'])
                        
                        current_cluster = []
                        
                        for e in entries:
                            if not current_cluster:
                                current_cluster.append(e)
                            else:
                                # Check date diff with first element of cluster
                                # (Assuming cluster spans small range)
                                diff = (e['date'] - current_cluster[0]['date']).days
                                if diff <= 3:
                                    current_cluster.append(e)
                                else:
                                    # Finalize current cluster
                                    # Check if >= 2 unique horses
                                    unique_horses = set(x['horse_name'] for x in current_cluster)
                                    if len(unique_horses) >= 2:
                                        valid_groups.append({
                                            'key': key,
                                            'entries': current_cluster
                                        })
                                    # Start new cluster
                                    current_cluster = [e]
                        
                        # Check last cluster
                        if current_cluster:
                            unique_horses = set(x['horse_name'] for x in current_cluster)
                            if len(unique_horses) >= 2:
                                valid_groups.append({
                                    'key': key,
                                    'entries': current_cluster
                                })

                    # Sort groups by Date (Newest first)
                    valid_groups.sort(key=lambda x: x['entries'][0]['date'], reverse=True)

                    if valid_groups:
                        matchup_sheet_name = f"Matchups_{race_id[-2:]}"
                        matchup_sheet = workbook.create_sheet(title=matchup_sheet_name)
                        current_row = 1
                        
                        # --- A. Win/Loss Matrix ---
                        # Identify all horses involved in valid groups
                        involved_horses = set()
                        for g in valid_groups:
                            for e in g['entries']:
                                involved_horses.add(e['horse_name'])
                        
                        # Sort by Umaban (extract number from "(N) Name")
                        def get_umaban_sort(name):
                            try:
                                return int(name.split(')')[0].replace('(', ''))
                            except:
                                return 999
                        
                        sorted_horses = sorted(list(involved_horses), key=get_umaban_sort)
                        
                        # Calculate Wins/Losses
                        # matrix[A][B] = {'win': 0, 'loss': 0}
                        matrix_scores = {h: {h2: {'win': 0, 'loss': 0} for h2 in sorted_horses} for h in sorted_horses}
                        
                        for g in valid_groups:
                            ents = g['entries']
                            # Pairwise comparison in this group
                            for i in range(len(ents)):
                                for j in range(i + 1, len(ents)):
                                    h1 = ents[i]
                                    h2 = ents[j]
                                    
                                    if h1['horse_name'] == h2['horse_name']: continue
                                    
                                    # Compare Time (Faster is better)
                                    t1 = h1.get('time_sec')
                                    t2 = h2.get('time_sec')
                                    
                                    if t1 is not None and t2 is not None:
                                        if t1 < t2:
                                            # h1 wins (faster)
                                            matrix_scores[h1['horse_name']][h2['horse_name']]['win'] += 1
                                            matrix_scores[h2['horse_name']][h1['horse_name']]['loss'] += 1
                                        elif t1 > t2:
                                            # h2 wins (faster)
                                            matrix_scores[h1['horse_name']][h2['horse_name']]['loss'] += 1
                                            matrix_scores[h2['horse_name']][h1['horse_name']]['win'] += 1
                                    else:
                                        # Fallback to Rank if Time is missing (shouldn't happen often)
                                        if h1['rank'] < h2['rank']:
                                            matrix_scores[h1['horse_name']][h2['horse_name']]['win'] += 1
                                            matrix_scores[h2['horse_name']][h1['horse_name']]['loss'] += 1
                                        elif h1['rank'] > h2['rank']:
                                            matrix_scores[h1['horse_name']][h2['horse_name']]['loss'] += 1
                                            matrix_scores[h2['horse_name']][h1['horse_name']]['win'] += 1
                        
                        # Draw Matrix Header
                        matchup_sheet.cell(row=current_row, column=1, value="対戦マトリクス (表\\裏)")
                        for col_idx, h_name in enumerate(sorted_horses, 2):
                            # Vertical Text for Header
                            c = matchup_sheet.cell(row=current_row, column=col_idx, value=h_name)
                            c.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
                            c.font = Font(bold=True)
                        
                        current_row += 1
                        
                        # Draw Matrix Rows
                        for h_row in sorted_horses:
                            # Row Header
                            c = matchup_sheet.cell(row=current_row, column=1, value=h_row)
                            c.font = Font(bold=True)
                            
                            for col_idx, h_col in enumerate(sorted_horses, 2):
                                cell = matchup_sheet.cell(row=current_row, column=col_idx)
                                cell.alignment = Alignment(horizontal='center')
                                
                                if h_row == h_col:
                                    cell.value = "-"
                                    cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
                                else:
                                    record = matrix_scores[h_row][h_col]
                                    wins = record['win']
                                    losses = record['loss']
                                    
                                    if wins == 0 and losses == 0:
                                        cell.value = ""
                                    else:
                                        cell.value = f"{wins}勝{losses}敗"
                                        if wins > losses:
                                            cell.font = Font(color="FF0000", bold=True) # Red for Winning Record
                                        elif losses > wins:
                                            cell.font = Font(color="0000FF") # Blue for Losing Record
                                        else:
                                            cell.font = Font(color="000000") # Black for Even

                            current_row += 1
                        
                        current_row += 2 # Spacer
                        
                        # --- B. Detailed List by Group ---
                        header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                        bold_font = Font(bold=True)
                        
                        for g_idx, group in enumerate(valid_groups, 1):
                            entries = group['entries']
                            first = entries[0]
                            # Header: [Date] [Place] [Course] [Dist] [Cond] (Group N)
                            # Note: Date might be a range if clustered, but usually same day or close.
                            d_str = first['date_str']
                            header_text = f"Group {g_idx}: {d_str} {first['place']} {first['course']}{first['dist']} ({first['cond']})"
                            
                            cell = matchup_sheet.cell(row=current_row, column=1, value=header_text)
                            cell.font = bold_font
                            cell.fill = header_fill
                            matchup_sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
                            current_row += 1
                            
                            # Cols
                            cols = ["着順", "馬名", "レース名", "タイム", "着差", "上り", "通過"]
                            for c_idx, col_name in enumerate(cols, 1):
                                cell = matchup_sheet.cell(row=current_row, column=c_idx, value=col_name)
                                cell.font = bold_font
                                cell.alignment = Alignment(horizontal='center')
                            current_row += 1
                            
                            # Sort by Rank
                            entries.sort(key=lambda x: x['rank'])
                            
                            for e in entries:
                                matchup_sheet.cell(row=current_row, column=1, value=e['rank'])
                                
                                name_cell = matchup_sheet.cell(row=current_row, column=2, value=e['horse_name'])
                                try:
                                    w_int = int(e['waku'])
                                    if w_int in WAKU_COLORS:
                                        s = WAKU_COLORS[w_int]
                                        name_cell.font = Font(color=s['bg'], bold=True)
                                except:
                                    pass
                                
                                matchup_sheet.cell(row=current_row, column=3, value=e['race_name'])
                                matchup_sheet.cell(row=current_row, column=4, value=e['time'])
                                matchup_sheet.cell(row=current_row, column=5, value=e['diff'])
                                matchup_sheet.cell(row=current_row, column=6, value=e['3f'])
                                matchup_sheet.cell(row=current_row, column=7, value=e['passing'])
                                current_row += 1
                            
                            current_row += 1
        
        logger.info(f"Excel report saved to {output_file}")

def main():
    parser = argparse.ArgumentParser(description="Generate Netkeiba Race Analysis Excel")
    parser.add_argument("--date", required=True, help="Target Date (YYYYMMDD), e.g., 20251130")
    parser.add_argument("--venue", required=True, help="Venue Name, e.g., 東京")
    
    args = parser.parse_args()
    
    analyzer = NetkeibaAnalyzer(args.date, args.venue)
    analyzer.run()

if __name__ == "__main__":
    main()
