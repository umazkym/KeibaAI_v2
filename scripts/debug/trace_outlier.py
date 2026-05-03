#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""外れ値レコードの計算過程を詳細に追跡"""
import openpyxl
import pickle

# データ読み込み
with open('keibaai/data/processed/global_stats_lookup.pkl', 'rb') as f:
    global_stats = pickle.load(f)
with open('keibaai/data/processed/venue_adjustment.pkl', 'rb') as f:
    venue_adjustment = pickle.load(f)

# レポート読み込み
wb = openpyxl.load_workbook('outputs/reports/race_analysis_20251228_中山.xlsx')
ws = wb['Race_01']
headers = [cell.value for cell in ws[1]]

# 外れ値を持つ行を詳細分析
print("=" * 80)
print("外れ値レコードの詳細追跡")
print("=" * 80)

for row in ws.iter_rows(min_row=2, max_row=50, values_only=True):
    data = dict(zip(headers, row))
    
    t_idx = data.get('タイム指数')
    if t_idx is None:
        continue
    
    try:
        t = float(t_idx)
    except:
        continue
    
    if t < 20 or t > 80:
        horse = data.get('馬名')
        venue = data.get('場所')
        dist_raw = data.get('距離')
        surf = data.get('ｺｰｽ')
        cond = data.get('馬場')
        time_sec = data.get('タイム秒')
        
        print(f"\n■ 馬名: {horse} (指数={t:.1f})")
        print(f"  場所: {venue}, 距離: {dist_raw}, コース: {surf}, 馬場: {cond}")
        print(f"  タイム秒: {time_sec}")
        
        # 距離をパース
        try:
            dist = int(dist_raw) if dist_raw else 0
        except:
            dist = 0
        
        # 馬場状態マッピング
        cond_map = {"良": "良", "稍": "稍重", "重": "重", "不": "不良"}
        full_cond = cond_map.get(cond, cond) if cond else None
        
        # キーを構築
        venue_key = (venue, dist, surf, full_cond) if venue and dist and surf and full_cond else None
        global_key = (dist, surf, full_cond) if dist and surf and full_cond else None
        
        print(f"  venue_key: {venue_key}")
        print(f"  global_key: {global_key}")
        
        # 補正データ取得
        adj_data = venue_adjustment.get(venue_key) if venue_key else None
        if not adj_data and venue_key:
            # フォールバック
            fallback_key = (venue, dist, surf)
            adj_data = venue_adjustment.get(fallback_key)
            print(f"  → フォールバックキー使用: {fallback_key}")
        
        print(f"  補正データ: {adj_data}")
        
        # グローバル統計取得
        g_stats = global_stats.get(global_key) if global_key else None
        print(f"  グローバル統計: {g_stats}")
        
        # 計算シミュレーション
        if adj_data and g_stats and time_sec:
            time_adj = adj_data.get('time_adjustment', 0.0)
            adj_time = float(time_sec) - time_adj
            g_mean = g_stats['time_mean']
            g_std = g_stats['time_std']
            calculated_idx = 50 + 10 * (g_mean - adj_time) / g_std
            
            print(f"\n  【計算シミュレーション】")
            print(f"    time_adjustment: {time_adj:.3f}")
            print(f"    adj_time = {time_sec} - {time_adj:.3f} = {adj_time:.2f}")
            print(f"    global_mean: {g_mean:.2f}, global_std: {g_std:.2f}")
            print(f"    指数 = 50 + 10 * ({g_mean:.2f} - {adj_time:.2f}) / {g_std:.2f}")
            print(f"         = 50 + 10 * {g_mean - adj_time:.2f} / {g_std:.2f}")
            print(f"         = 50 + {10 * (g_mean - adj_time) / g_std:.1f}")
            print(f"         = {calculated_idx:.1f}")
            
            if abs(calculated_idx - t) > 0.5:
                print(f"  ⚠️ 計算結果とレポート値が不一致! レポート値={t:.1f}")
        else:
            print(f"  ⚠️ データ不足で計算できず")
