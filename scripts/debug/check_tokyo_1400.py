#!/usr/bin/env python3
import pickle
import openpyxl

# 補正値データ
with open('keibaai/data/processed/venue_adjustment.pkl', 'rb') as f:
    venue_adjustment = pickle.load(f)
with open('keibaai/data/processed/global_stats_lookup.pkl', 'rb') as f:
    global_stats = pickle.load(f)

# 芝1400m東京が欠損している理由を調査
key = ('東京', 1400, '芝', '良')
print(f'東京 芝1400m 良:')
print(f'  venue_adjustment 完全キー: {venue_adjustment.get(key)}')
print(f'  venue_adjustment フォールバック: {venue_adjustment.get(("東京", 1400, "芝"))}')
print(f'  グローバル統計: {global_stats.get((1400, "芝", "良"))}')
print()

# 芝1400m のすべての競馬場を確認
print('芝1400m の全競馬場補正値:')
for venue in ['札幌', '函館', '福島', '新潟', '東京', '中山', '中京', '京都', '阪神', '小倉']:
    key = (venue, 1400, '芝', '良')
    adj = venue_adjustment.get(key)
    if adj:
        print(f'  {venue}: {adj["time_adjustment"]:+.2f}秒 (count={adj["count"]})')
    else:
        # フォールバック確認
        fb = venue_adjustment.get((venue, 1400, '芝'))
        if fb:
            print(f'  {venue}: {fb["time_adjustment"]:+.2f}秒 (フォールバック)')
        else:
            print(f'  {venue}: なし')
print()

# 実際のレポートで東京芝1400mの指数がどうなっているか確認
wb = openpyxl.load_workbook('outputs/reports/race_analysis_20251228_中山.xlsx')
ws = wb['Race_12']
headers = [cell.value for cell in ws[1]]

print('東京芝1400mの過去成績サンプル:')
found = False
for row in ws.iter_rows(min_row=2, max_row=200, values_only=True):
    data = dict(zip(headers, row))
    venue = data.get('場所')
    dist = data.get('距離')
    surf = data.get('ｺｰｽ')
    if venue == '東京' and str(dist) == '1400' and surf == '芝':
        print(f'  馬名: {data.get("馬名")}')
        print(f'  タイム: {data.get("タイム秒")}秒')
        print(f'  タイム指数: {data.get("タイム指数")}')
        print(f'  馬場: {data.get("馬場")}')
        found = True
        break

if not found:
    print('  該当データなし')
