
import pandas as pd
import os
import sys
from openpyxl import Workbook

# Add project root to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '../../')))

from scripts.reports.create_interactive_charts import create_interactive_charts

def create_mock_excel(filename):
    wb = Workbook()
    
    # 1. AI Analysis Sheet
    ws_ai = wb.create_sheet("AI分析_01R")
    ws_ai.append(["01R Mock Race"]) # Row 1: Title
    ws_ai.append(["馬番", "馬名", "騎手", "現オッズ", "人気"]) # Row 2: Header
    
    horses = [
        [1, "Mock Horse 1", "Jockey A", 2.5, 1],
        [2, "Mock Horse 2", "Jockey B", 5.0, 2],
        [3, "Mock Horse 3", "Jockey C", 8.0, 3],
        [4, "Mock Horse 4", "Jockey D", 15.0, 4],
        [5, "Mock Horse 5", "Jockey E", 50.0, 5],
        [6, "Mock Horse 6", "Jockey F", 100.0, 6],
    ]
    
    for h in horses:
        ws_ai.append(h)

    # 2. Race Data Sheet
    ws_race = wb.create_sheet("Race_01")
    headers = [
        '日付', '場所', 'R', 'レース名', 'ｺｰｽ', '距離', '馬場', '頭数', 
        '枠', '番', '馬名', '騎手', '着順', 'タイム', '上がり', 
        'T指数', 'L指数', 'RPCI', '1角', '2角', '3角', '4角', '通過'
    ]
    ws_race.append(headers)
    
    # Generate some dummy history data
    import random
    venues = ["中山", "東京", "阪神", "京都", "船橋", "大井"]
    courses = ["芝", "ダート"]
    
    for i, h in enumerate(horses):
        umaban = h[0]
        name = h[1]
        jockey = h[2]
        
        for _ in range(10): # 10 history records each
            row = {
                '日付': f"2025/01/{random.randint(1, 20):02d}",
                '場所': random.choice(venues),
                'R': f"{random.randint(1, 12)}R",
                'レース名': "Dummy Race",
                'ｺｰｽ': random.choice(courses),
                '距離': random.choice(["1200", "1600", "2000", "2400"]),
                '馬場': "良",
                '頭数': 16,
                '枠': (umaban - 1) // 2 + 1,
                '番': umaban, # Usually this is '出走馬番' but code maps '番' too
                '出走馬番': umaban,
                '出走枠番': (umaban - 1) // 2 + 1,
                '馬名': name,
                '騎手': jockey,
                '着順': random.randint(1, 16),
                'タイム': "1:35.0",
                '上がり': 35.0,
                'T指数': random.uniform(40, 80),
                'L指数': random.uniform(40, 80),
                'RPCI': random.uniform(40, 60),
                '1角': random.randint(1, 16),
                '2角': random.randint(1, 16),
                '3角': random.randint(1, 16),
                '4角': random.randint(1, 16)
            }
            # Add required columns for code to work
            data = [row.get(col, '') for col in headers]
            ws_race.append(data)
            
    ws_race.cell(row=1, column=10).value = "出走馬番" # Fix header for mapping logic check
    ws_race.cell(row=1, column=9).value = "出走枠番"

    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
        
    wb.save(filename)
    print(f"Created {filename}")
    return filename

if __name__ == "__main__":
    excel_file = "C:/Users/zk-ht/Keiba/Keiba_AI_v2/outputs/reports/race_analysis_mock_interactive.xlsx"
    os.makedirs(os.path.dirname(excel_file), exist_ok=True)
    create_mock_excel(excel_file)
    
    html_file = create_interactive_charts(excel_file)
    print(f"Generated HTML: {html_file}")
