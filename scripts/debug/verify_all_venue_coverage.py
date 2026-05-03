#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
全競馬場×距離×馬場での補正値網羅性を検証

1. 実際のレポートで過去成績の競馬場分布を確認
2. 各競馬場×距離×馬場の補正値存在を確認
3. 補正値が欠損しているケースを特定
"""

import openpyxl
import pickle
from collections import defaultdict
from pathlib import Path

def load_data():
    """データ読み込み"""
    with open('keibaai/data/processed/global_stats_lookup.pkl', 'rb') as f:
        global_stats = pickle.load(f)
    with open('keibaai/data/processed/venue_adjustment.pkl', 'rb') as f:
        venue_adjustment = pickle.load(f)
    return global_stats, venue_adjustment

def analyze_past_venues_in_report():
    """レポート内の過去成績で使われている競馬場を分析"""
    print("=" * 80)
    print("1. 実際のレポート内の過去成績で使われている競馬場を分析")
    print("=" * 80)
    
    reports = [
        Path('outputs/reports/race_analysis_20251228_中山.xlsx'),
        Path('outputs/reports/race_analysis_20251228_阪神.xlsx'),
    ]
    
    venue_dist_cond_count = defaultdict(int)
    
    for report in reports:
        if not report.exists():
            continue
            
        wb = openpyxl.load_workbook(report)
        
        for sheet_name in wb.sheetnames:
            if not sheet_name.startswith('Race_'):
                continue
            
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                
                venue = data.get('場所')
                dist = data.get('距離')
                surf = data.get('ｺｰｽ')
                cond = data.get('馬場')
                
                if venue and dist and surf and cond:
                    try:
                        d = int(dist)
                        cond_map = {"良": "良", "稍": "稍重", "重": "重", "不": "不良"}
                        full_cond = cond_map.get(cond, cond)
                        key = (venue, d, surf, full_cond)
                        venue_dist_cond_count[key] += 1
                    except:
                        pass
    
    return venue_dist_cond_count

def check_coverage(venue_dist_cond_count, global_stats, venue_adjustment):
    """補正値のカバレッジを確認"""
    print("\n" + "=" * 80)
    print("2. 過去成績条件に対する補正値カバレッジ")
    print("=" * 80)
    
    # 競馬場ごとの統計
    venue_stats = defaultdict(lambda: {'total': 0, 'covered': 0, 'missing': []})
    
    for key, count in sorted(venue_dist_cond_count.items(), key=lambda x: -x[1]):
        venue, dist, surf, cond = key
        venue_stats[venue]['total'] += count
        
        # 補正データ存在確認
        adj = venue_adjustment.get(key)
        if adj:
            venue_stats[venue]['covered'] += count
        else:
            # フォールバックキー確認
            fallback = venue_adjustment.get((venue, dist, surf))
            if fallback:
                venue_stats[venue]['covered'] += count
            else:
                venue_stats[venue]['missing'].append((key, count))
    
    print(f"\n{'競馬場':<8} {'総レコード':>10} {'カバー':>10} {'カバー率':>10} {'欠損条件数':>10}")
    print("-" * 60)
    
    total_records = 0
    covered_records = 0
    
    for venue in ['札幌', '函館', '福島', '新潟', '東京', '中山', '中京', '京都', '阪神', '小倉']:
        s = venue_stats.get(venue, {'total': 0, 'covered': 0, 'missing': []})
        rate = s['covered'] / s['total'] * 100 if s['total'] > 0 else 0
        missing_count = len(s['missing'])
        print(f"{venue:<8} {s['total']:>10} {s['covered']:>10} {rate:>9.1f}% {missing_count:>10}")
        total_records += s['total']
        covered_records += s['covered']
    
    print("-" * 60)
    overall_rate = covered_records / total_records * 100 if total_records > 0 else 0
    print(f"{'合計':<8} {total_records:>10} {covered_records:>10} {overall_rate:>9.1f}%")
    
    return venue_stats

def show_missing_details(venue_stats):
    """欠損している条件の詳細を表示"""
    print("\n" + "=" * 80)
    print("3. 補正値が欠損している条件の詳細（上位20件）")
    print("=" * 80)
    
    all_missing = []
    for venue, stats in venue_stats.items():
        for key, count in stats['missing']:
            all_missing.append((key, count))
    
    all_missing.sort(key=lambda x: -x[1])
    
    if not all_missing:
        print("欠損なし！全ての条件がカバーされています。")
        return
    
    print(f"\n{'条件':<40} {'レコード数':>10}")
    print("-" * 55)
    for key, count in all_missing[:20]:
        venue, dist, surf, cond = key
        condition_str = f"{venue} {surf}{dist}m {cond}"
        print(f"{condition_str:<40} {count:>10}")

def analyze_distance_coverage(venue_adjustment):
    """距離ごとの補正値カバレッジを分析"""
    print("\n" + "=" * 80)
    print("4. 距離別の補正値存在状況")
    print("=" * 80)
    
    venues = ['札幌', '函館', '福島', '新潟', '東京', '中山', '中京', '京都', '阪神', '小倉']
    distances = [1000, 1200, 1400, 1600, 1800, 2000, 2200, 2400]
    surfaces = ['芝', 'ダート']
    conditions = ['良', '稍重', '重', '不良']
    
    print("\n■ 芝・良 の補正値存在マトリックス")
    print(f"{'距離':<8}", end="")
    for venue in venues:
        print(f"{venue:<6}", end="")
    print()
    print("-" * 70)
    
    for dist in distances:
        print(f"{dist}m", end="   ")
        for venue in venues:
            key = (venue, dist, '芝', '良')
            if key in venue_adjustment:
                adj = venue_adjustment[key]['time_adjustment']
                print(f"{adj:+5.2f}", end=" ")
            else:
                print(f"{'---':>5}", end=" ")
        print()
    
    print("\n■ ダート・良 の補正値存在マトリックス")
    print(f"{'距離':<8}", end="")
    for venue in venues:
        print(f"{venue:<6}", end="")
    print()
    print("-" * 70)
    
    for dist in [1000, 1200, 1400, 1600, 1700, 1800, 2000, 2100]:
        print(f"{dist}m", end="   ")
        for venue in venues:
            key = (venue, dist, 'ダート', '良')
            if key in venue_adjustment:
                adj = venue_adjustment[key]['time_adjustment']
                print(f"{adj:+5.2f}", end=" ")
            else:
                print(f"{'---':>5}", end=" ")
        print()

def trace_specific_horse():
    """特定の馬の過去成績で補正を追跡"""
    print("\n" + "=" * 80)
    print("5. 具体例: 特定の馬の過去成績補正を追跡")
    print("=" * 80)
    
    # データ読み込み
    global_stats, venue_adjustment = load_data()
    
    report = Path('outputs/reports/race_analysis_20251228_中山.xlsx')
    if not report.exists():
        print("レポートが見つかりません")
        return
    
    wb = openpyxl.load_workbook(report)
    ws = wb['Race_12']  # 有馬記念などの重賞
    headers = [cell.value for cell in ws[1]]
    
    # 一人の馬の過去成績を追跡（最初の馬）
    current_horse = None
    horse_records = []
    
    for row in ws.iter_rows(min_row=2, max_row=20, values_only=True):
        data = dict(zip(headers, row))
        
        horse = data.get('馬名')
        if not current_horse:
            current_horse = horse
        
        if horse == current_horse:
            horse_records.append(data)
    
    if not horse_records:
        print("馬のデータが見つかりません")
        return
    
    print(f"\n馬名: {current_horse}")
    print(f"過去走数: {len(horse_records)}")
    print()
    
    for i, rec in enumerate(horse_records[:8]):  # 最大8件
        venue = rec.get('場所')
        dist = rec.get('距離')
        surf = rec.get('ｺｰｽ')
        cond = rec.get('馬場')
        time_idx = rec.get('タイム指数')
        time_sec = rec.get('タイム秒')
        
        cond_map = {"良": "良", "稍": "稍重", "重": "重", "不": "不良"}
        full_cond = cond_map.get(cond, cond) if cond else None
        
        try:
            d = int(dist) if dist else 0
        except:
            d = 0
        
        key = (venue, d, surf, full_cond)
        adj = venue_adjustment.get(key)
        if not adj:
            adj = venue_adjustment.get((venue, d, surf))
        
        print(f"[{i+1}] {venue} {surf}{d}m {cond}")
        print(f"    タイム: {time_sec}秒, タイム指数: {time_idx}")
        if adj:
            print(f"    補正値: {adj.get('time_adjustment', 0):+.2f}秒 ✓")
        else:
            print(f"    補正値: 欠損 ⚠️")
        print()

def main():
    global_stats, venue_adjustment = load_data()
    
    print(f"グローバル統計: {len(global_stats)}条件")
    print(f"競馬場補正値: {len(venue_adjustment)}条件")
    
    # 1. 実際のレポートで使われている条件を収集
    venue_dist_cond_count = analyze_past_venues_in_report()
    print(f"\n使用されている条件数: {len(venue_dist_cond_count)}")
    print(f"総レコード数: {sum(venue_dist_cond_count.values())}")
    
    # 2. カバレッジ確認
    venue_stats = check_coverage(venue_dist_cond_count, global_stats, venue_adjustment)
    
    # 3. 欠損詳細
    show_missing_details(venue_stats)
    
    # 4. 距離マトリックス
    analyze_distance_coverage(venue_adjustment)
    
    # 5. 具体例
    trace_specific_horse()

if __name__ == "__main__":
    main()
