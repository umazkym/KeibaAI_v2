#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
グローバル指数の実装を深く検証するスクリプト

問題: 競馬場間で指数が本当に整合しているか？
- 同じ能力の馬が異なる競馬場で走った場合、同じ指数になるか？
- 補正値計算のロジックに問題はないか？
- 外れ値の原因は何か？
"""

import pandas as pd
import numpy as np
import pickle
import openpyxl
from pathlib import Path

def load_data():
    """データ読み込み"""
    global_stats = {}
    venue_adjustment = {}
    
    gpath = Path(r"keibaai\data\processed\global_stats_lookup.pkl")
    vpath = Path(r"keibaai\data\processed\venue_adjustment.pkl")
    
    if gpath.exists():
        with open(gpath, 'rb') as f:
            global_stats = pickle.load(f)
    if vpath.exists():
        with open(vpath, 'rb') as f:
            venue_adjustment = pickle.load(f)
    
    return global_stats, venue_adjustment

def analyze_calculation_logic(global_stats, venue_adjustment):
    """計算ロジックの妥当性を検証"""
    print("=" * 70)
    print("1. 補正値の分布検証")
    print("=" * 70)
    
    time_adjs = []
    l3f_adjs = []
    
    for key, data in venue_adjustment.items():
        if len(key) == 4:  # (venue, dist, surf, cond)
            time_adjs.append(data['time_adjustment'])
            l3f_adj = data.get('l3f_adjustment', 0)
            l3f_adjs.append(l3f_adj)
    
    print(f"タイム補正値: {len(time_adjs)}件")
    print(f"  平均: {np.mean(time_adjs):.3f}秒")
    print(f"  標準偏差: {np.std(time_adjs):.3f}秒")
    print(f"  最小: {np.min(time_adjs):.3f}秒")
    print(f"  最大: {np.max(time_adjs):.3f}秒")
    
    print(f"\nL3F補正値: {len(l3f_adjs)}件")
    print(f"  平均: {np.mean(l3f_adjs):.3f}秒")
    print(f"  標準偏差: {np.std(l3f_adjs):.3f}秒")
    print(f"  最小: {np.min(l3f_adjs):.3f}秒")
    print(f"  最大: {np.max(l3f_adjs):.3f}秒")
    
    # 極端な補正値を持つ条件をリスト
    print("\n■ 極端なタイム補正値 (|adj| > 2.0秒):")
    for key, data in venue_adjustment.items():
        if len(key) == 4 and abs(data['time_adjustment']) > 2.0:
            print(f"  {key}: {data['time_adjustment']:+.2f}秒 (count={data['count']})")
    
    print("\n■ 極端なL3F補正値 (|adj| > 1.0秒):")
    for key, data in venue_adjustment.items():
        if len(key) == 4:
            l3f = data.get('l3f_adjustment', 0)
            if abs(l3f) > 1.0:
                print(f"  {key}: {l3f:+.2f}秒 (count={data['count']})")

def simulate_index_calculation(global_stats, venue_adjustment):
    """具体的なケースで指数計算をシミュレート"""
    print("\n" + "=" * 70)
    print("2. 同一条件での競馬場間比較シミュレーション")
    print("=" * 70)
    
    # 芝2000m良での比較
    test_cases = [
        ("芝2000m良", (2000, "芝", "良"), 121.0),  # 仮想タイム 2:01.0
        ("ダート1800m良", (1800, "ダート", "良"), 114.0),  # 仮想タイム 1:54.0
    ]
    
    venues = ["東京", "阪神", "中山", "京都"]
    
    for case_name, global_key, base_time in test_cases:
        if global_key not in global_stats:
            print(f"\n{case_name}: グローバル統計なし")
            continue
            
        g = global_stats[global_key]
        print(f"\n■ {case_name}: グローバル平均={g['time_mean']:.2f}秒, 標準偏差={g['time_std']:.2f}秒")
        print(f"  仮想タイム: {base_time:.1f}秒（各競馬場での走破タイムが同一と仮定）")
        print()
        
        indices = {}
        for venue in venues:
            key = (venue, global_key[0], global_key[1], global_key[2])
            adj_data = venue_adjustment.get(key)
            
            if adj_data:
                time_adj = adj_data['time_adjustment']
                adj_time = base_time - time_adj
                idx = 50 + 10 * (g['time_mean'] - adj_time) / g['time_std']
                indices[venue] = idx
                print(f"  {venue}: 補正={time_adj:+.2f}秒 → 補正後タイム={adj_time:.2f}秒 → 指数={idx:.1f}")
            else:
                print(f"  {venue}: 補正データなし")
        
        if len(indices) > 1:
            vals = list(indices.values())
            print(f"\n  → 指数のレンジ: {max(vals) - min(vals):.1f} (理想は0に近いこと)")

def check_edge_cases(global_stats, venue_adjustment):
    """エッジケース：補正値がないケースの挙動確認"""
    print("\n" + "=" * 70)
    print("3. 補正値が欠損するケースの分析")
    print("=" * 70)
    
    # グローバル統計にあるが、特定の競馬場補正がないケース
    missing_count = 0
    partial_count = 0
    
    venues = ["札幌", "函館", "福島", "新潟", "東京", "中山", "中京", "京都", "阪神", "小倉"]
    
    for g_key in global_stats.keys():
        dist, surf, cond = g_key
        venue_count = 0
        for venue in venues:
            v_key = (venue, dist, surf, cond)
            if v_key in venue_adjustment:
                venue_count += 1
        
        if venue_count == 0:
            missing_count += 1
        elif venue_count < 10:
            partial_count += 1
    
    print(f"全グローバル条件数: {len(global_stats)}")
    print(f"  競馬場補正が完全欠損: {missing_count}件")
    print(f"  競馬場補正が部分的: {partial_count}件 (10場未満)")
    
    # フォールバック（馬場状態なし）の確認
    print("\n■ フォールバックキー（馬場状態なし）の有効性:")
    fallback_keys = [k for k in venue_adjustment.keys() if len(k) == 3]
    print(f"  フォールバックキー数: {len(fallback_keys)}")

def analyze_actual_report_data():
    """実際のレポートデータで外れ値を詳細分析"""
    print("\n" + "=" * 70)
    print("4. 実際のレポートデータの外れ値分析")
    print("=" * 70)
    
    report_path = Path(r"outputs\reports\race_analysis_20251228_中山.xlsx")
    if not report_path.exists():
        print("レポートファイルが見つかりません")
        return
    
    wb = openpyxl.load_workbook(report_path)
    
    # Race_12シート（指数データあり）を分析
    for sheet_name in ['Race_12', 'Race_02', 'Race_03']:
        if sheet_name not in wb.sheetnames:
            continue
            
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        
        # カラムインデックスを取得
        cols = {}
        for i, h in enumerate(headers):
            if h:
                cols[h] = i
        
        print(f"\n■ {sheet_name}シート分析")
        
        # 各行のデータを収集
        outliers = []
        normal = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            
            t_idx = data.get('タイム指数')
            l_idx = data.get('上り指数')
            venue = data.get('場所')
            dist = data.get('距離')
            cond = data.get('馬場')
            time_sec = data.get('タイム秒')
            
            if t_idx is not None:
                try:
                    t = float(t_idx)
                    if t < 20 or t > 80:
                        outliers.append({
                            '馬名': data.get('馬名'),
                            '場所': venue,
                            '距離': dist,
                            '馬場': cond,
                            'タイム秒': time_sec,
                            'タイム指数': t,
                            '上り指数': l_idx
                        })
                    else:
                        normal.append(t)
                except:
                    pass
        
        if normal:
            print(f"  正常範囲の指数: {len(normal)}件, 平均={np.mean(normal):.1f}")
        if outliers:
            print(f"  外れ値: {len(outliers)}件")
            for o in outliers[:5]:
                print(f"    - {o}")
        
        break  # 1シートだけサンプル

def main():
    global_stats, venue_adjustment = load_data()
    
    if not global_stats or not venue_adjustment:
        print("データが読み込めませんでした")
        return
    
    print(f"グローバル統計: {len(global_stats)}条件")
    print(f"競馬場補正値: {len(venue_adjustment)}条件")
    
    analyze_calculation_logic(global_stats, venue_adjustment)
    simulate_index_calculation(global_stats, venue_adjustment)
    check_edge_cases(global_stats, venue_adjustment)
    analyze_actual_report_data()
    
    print("\n" + "=" * 70)
    print("5. 根本的な問題と改善案")
    print("=" * 70)
    print("""
【発見された問題】
1. 補正値の符号: venue_mean - global_mean で「競馬場が遅い場合は正」
   → 計算時に time_sec - adj で「遅い場所のタイムを速くする」は正しい
   
2. 問題: 「過去成績」のレースが少ないサンプルの条件の場合、
   補正値が極端になり、指数も極端になる可能性

3. 根本的な懸念: 
   - グローバル基準の平均・標準偏差は「全競馬場の混合データ」から算出
   - しかし競馬場ごとのレベル差（出走馬の質の違い）は考慮されていない
   - 例: 東京の重賞 vs 地方転戦組が集まるローカル → 同じ距離でも馬の質が違う

【改善案】
1. 補正値に最小サンプル数制限（count>=20など）を追加
2. 極端な補正値（±3秒以上など）にはクリッピング
3. 補正値がない場合のフォールバック精度を向上
""")

if __name__ == "__main__":
    main()
