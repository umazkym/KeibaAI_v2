#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
シンプル分析BOOK生成スクリプト

Race_XX、Analysis_XX、Matchups_XXのみで構成されるシンプルなExcelブックを生成

Analysis_XXには以下のチャートを含める:
① 横軸：上り指数、縦軸：タイム指数
② 横軸：1C/頭数、縦軸：タイム指数  
③ 横軸：1C/頭数、縦軸：上り指数
"""

import argparse
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.legend import Legend
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# 枠色定義
WAKU_COLORS = {
    1: {'bg': 'FFFFFF', 'fg': '000000'},
    2: {'bg': '000000', 'fg': 'FFFFFF'},
    3: {'bg': 'FF0000', 'fg': 'FFFFFF'},
    4: {'bg': '0000FF', 'fg': 'FFFFFF'},
    5: {'bg': 'FFFF00', 'fg': '000000'},
    6: {'bg': '00FF00', 'fg': '000000'},
    7: {'bg': 'FFA500', 'fg': '000000'},
    8: {'bg': 'FFC0CB', 'fg': '000000'},
}

def get_waku_color(umaban: int) -> str:
    """馬番から枠色を取得"""
    if umaban <= 2: waku = 1
    elif umaban <= 4: waku = 2
    elif umaban <= 6: waku = 3
    elif umaban <= 8: waku = 4
    elif umaban <= 10: waku = 5
    elif umaban <= 12: waku = 6
    elif umaban <= 14: waku = 7
    else: waku = 8
    return WAKU_COLORS.get(waku, {}).get('bg', '808080')


def create_simple_analysis_book(source_file: str, output_file: str = None):
    """
    既存のレポートからシンプル分析BOOKを生成
    
    Args:
        source_file: 元のrace_analysis_YYYYMMDD_会場.xlsx
        output_file: 出力ファイル名（省略時は_simple.xlsxを付加）
    """
    if not os.path.exists(source_file):
        logger.error(f"ファイルが見つかりません: {source_file}")
        return None
    
    if not output_file:
        base, ext = os.path.splitext(source_file)
        output_file = f"{base}_simple{ext}"
    
    logger.info(f"シンプルBOOKを生成: {source_file} → {output_file}")
    
    # 元のワークブックを読み込む
    source_wb = load_workbook(source_file)
    
    # 新しいワークブックを作成
    new_wb = Workbook()
    new_wb.remove(new_wb.active)  # デフォルトシートを削除
    
    # Race_, Analysis_, Matchups_ シートのみコピー
    for sheet_name in source_wb.sheetnames:
        if sheet_name.startswith('Race_') or sheet_name.startswith('Matchups_'):
            # シートをコピー
            source_sheet = source_wb[sheet_name]
            new_sheet = new_wb.create_sheet(title=sheet_name)
            
            for row in source_sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.alignment = cell.alignment.copy()
                        new_cell.border = cell.border.copy()
            
            # 列幅をコピー
            for col_letter, dim in source_sheet.column_dimensions.items():
                new_sheet.column_dimensions[col_letter].width = dim.width
    
    # Analysis_シートを新規に作成（指定のチャートを含む）
    for sheet_name in source_wb.sheetnames:
        if sheet_name.startswith('Race_'):
            race_num = sheet_name.replace('Race_', '')
            analysis_name = f"Analysis_{race_num}"
            
            # Race_シートからデータを取得
            race_sheet = source_wb[sheet_name]
            
            # ヘッダー行を取得
            headers = [cell.value for cell in race_sheet[1]]
            
            # 必要な列のインデックスを取得
            def get_col_idx(name):
                try:
                    return headers.index(name) + 1
                except ValueError:
                    return None
            
            idx_time = get_col_idx('タイム指数')
            idx_l3f = get_col_idx('上り指数')
            idx_1c = get_col_idx('1C')
            idx_head = get_col_idx('頭数')
            idx_umaban = get_col_idx('馬番')
            idx_name = get_col_idx('馬名')
            
            if not all([idx_time, idx_l3f, idx_1c, idx_head]):
                logger.warning(f"{sheet_name}: 必要なカラムが見つかりません")
                continue
            
            # Analysis_シートを作成
            analysis_sheet = new_wb.create_sheet(title=analysis_name)
            
            # チャート用データを収集
            chart_data = []
            for row_idx in range(2, race_sheet.max_row + 1):
                time_idx = race_sheet.cell(row=row_idx, column=idx_time).value
                l3f_idx = race_sheet.cell(row=row_idx, column=idx_l3f).value
                c1 = race_sheet.cell(row=row_idx, column=idx_1c).value
                heads = race_sheet.cell(row=row_idx, column=idx_head).value
                umaban = race_sheet.cell(row=row_idx, column=idx_umaban).value if idx_umaban else ''
                name = race_sheet.cell(row=row_idx, column=idx_name).value if idx_name else ''
                
                if time_idx is not None and l3f_idx is not None and c1 is not None and heads is not None:
                    try:
                        c1_ratio = float(c1) / float(heads) if float(heads) > 0 else None
                        chart_data.append({
                            'time_idx': float(time_idx),
                            'l3f_idx': float(l3f_idx),
                            'c1_ratio': c1_ratio,
                            'umaban': int(umaban) if umaban else 0,
                            'name': name or ''
                        })
                    except:
                        pass
            
            if not chart_data:
                logger.warning(f"{analysis_name}: チャートデータがありません")
                continue
            
            # チャート用データをシートに書き込み（右側に配置）
            data_col = 20  # T列から
            analysis_sheet.cell(row=1, column=data_col, value="馬番")
            analysis_sheet.cell(row=1, column=data_col+1, value="馬名")
            analysis_sheet.cell(row=1, column=data_col+2, value="タイム指数")
            analysis_sheet.cell(row=1, column=data_col+3, value="上り指数")
            analysis_sheet.cell(row=1, column=data_col+4, value="1C/頭数")
            
            for i, d in enumerate(chart_data):
                row = i + 2
                analysis_sheet.cell(row=row, column=data_col, value=d['umaban'])
                analysis_sheet.cell(row=row, column=data_col+1, value=d['name'])
                analysis_sheet.cell(row=row, column=data_col+2, value=d['time_idx'])
                analysis_sheet.cell(row=row, column=data_col+3, value=d['l3f_idx'])
                analysis_sheet.cell(row=row, column=data_col+4, value=d['c1_ratio'])
            
            end_row = len(chart_data) + 1
            
            # チャート作成関数
            def create_chart(title, x_title, y_title, x_col, y_col):
                chart = ScatterChart()
                chart.title = title
                chart.style = 2
                chart.x_axis.title = x_title
                chart.y_axis.title = y_title
                chart.height = 15  # 高さを大きく
                chart.width = 18   # 幅を大きく
                chart.legend = None
                
                # 軸の値を表示（重要）
                chart.x_axis.delete = False
                chart.y_axis.delete = False
                chart.x_axis.majorTickMark = "out"
                chart.y_axis.majorTickMark = "out"
                chart.x_axis.minorTickMark = None
                chart.y_axis.minorTickMark = None
                
                # 軸のラベルを表示
                chart.x_axis.number_format = '0.0'
                chart.y_axis.number_format = '0.0'
                
                # グリッドライン（薄いグレー）
                sp_pr = GraphicalProperties(ln=LineProperties(solidFill="E0E0E0"))
                chart.x_axis.majorGridlines = ChartLines(spPr=sp_pr)
                chart.y_axis.majorGridlines = ChartLines(spPr=sp_pr)
                
                # 馬ごとにシリーズを追加
                for i, d in enumerate(chart_data):
                    row = i + 2
                    x_ref = Reference(analysis_sheet, min_col=x_col, min_row=row, max_row=row)
                    y_ref = Reference(analysis_sheet, min_col=y_col, min_row=row, max_row=row)
                    
                    series_title = f"({d['umaban']}) {d['name']}"
                    s = Series(y_ref, x_ref, title=series_title)
                    s.marker.symbol = "circle"
                    s.marker.size = 12  # マーカーを大きく
                    s.graphicalProperties.line.noFill = True
                    
                    # 枠色を適用
                    color = get_waku_color(d['umaban'])
                    s.marker.graphicalProperties.solidFill = color
                    s.marker.graphicalProperties.line.solidFill = "000000"
                    
                    chart.series.append(s)
                
                return chart
            
            # ① 上り指数 vs タイム指数（A2に配置）
            chart1 = create_chart(
                "① スピード vs スタミナ（右上ほど高能力）",
                "上り指数 →", "タイム指数 ↑",
                data_col + 3, data_col + 2  # l3f_idx, time_idx
            )
            analysis_sheet.add_chart(chart1, "A2")
            
            # ② 1C/頭数 vs タイム指数（A34に配置）
            chart2 = create_chart(
                "② 位置取り vs タイム（左=逃げ先行、右=追込）",
                "1C/頭数 →", "タイム指数 ↑",
                data_col + 4, data_col + 2  # c1_ratio, time_idx
            )
            analysis_sheet.add_chart(chart2, "A34")
            
            # ③ 1C/頭数 vs 上り指数（A66に配置）
            chart3 = create_chart(
                "③ 位置取り vs 末脚（脚質タイプ分析）",
                "1C/頭数 →", "上り指数 ↑",
                data_col + 4, data_col + 3  # c1_ratio, l3f_idx
            )
            analysis_sheet.add_chart(chart3, "A66")
            
            logger.info(f"  {analysis_name}: 3つのチャートを作成")
    
    # 保存
    new_wb.save(output_file)
    logger.info(f"シンプルBOOKを保存: {output_file}")
    return output_file


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="シンプル分析BOOK生成")
    parser.add_argument("source", help="元のレポートファイル (race_analysis_YYYYMMDD_会場.xlsx)")
    parser.add_argument("-o", "--output", help="出力ファイル名（省略時は_simple.xlsxを付加）")
    
    args = parser.parse_args()
    
    result = create_simple_analysis_book(args.source, args.output)
    if result:
        print(f"完了: {result}")
    else:
        print("失敗")
