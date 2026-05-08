#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
NAR（地方競馬）レース分析レポート生成スクリプト V2

netkeiba (NAR) からスクレイピングし、中央競馬と同等の分析用Excelを作成する。
create_interactive_charts.py と完全互換のフォーマットで出力。
"""

import pandas as pd
import requests
from bs4 import BeautifulSoup
import argparse
import os
from datetime import datetime, timedelta
import re
import time
import logging
from typing import Dict, List, Optional, Any
import sys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# Constants
# scripts/reports/nar -> scripts/reports -> scripts -> Keiba_AI_v2
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
DATA_DIR = os.path.join(PROJECT_ROOT, "keibaai", "data")
HTML_DIR = os.path.join(DATA_DIR, "raw", "nar_html")
PARQUET_DIR = os.path.join(DATA_DIR, "parsed", "parquet", "nar_races")

# NAR Venue Mapping
NAR_VENUE_MAP = {
    "44": "大井", "42": "浦和", "43": "船橋", "45": "川崎",
    "46": "金沢", "47": "笠松", "48": "名古屋", "50": "園田",
    "51": "姫路", "54": "高知", "55": "佐賀", "30": "門別",
    "35": "盛岡", "36": "水沢"
}
NAR_VENUE_NAME_TO_CODE = {v: k for k, v in NAR_VENUE_MAP.items()}

# Excel出力カラム（中央競馬と同一構造）
EXCEL_COLUMNS = [
    '出走枠番', '出走馬番', '乗り替わり', 'グループNo',
    '日付', '場所', '回', '日', 'ｺｰｽ', '距離', 'R', '馬場', '天気', '頭数',
    '枠番', '馬番', '馬名', '斤量', '騎手', '厩舎', 'ﾀｲﾑ', 'タイム秒', '着差', '人気', 'ｵｯｽﾞ',
    '上り', '上り秒', 'ﾍﾟｰｽ1', 'ﾍﾟｰｽ2', '通過', '1C', '2C', '3C', '4C', '着順', '馬体重', '増減',
    'レース名', '性', '年齢',
    '平均t', '平均場t', '基準t', '基準場t', '基準3F', '基準場3F',
    '平t差', '平場t差', '基t差', '基場t差', '基3F差', '基場3F差',
    'タイム指数', '上り指数', '馬場差', 'RPCI',
    'horse_id', 'race_id'
]


class NarMetricCalculator:
    """NAR基準データからメトリクスを計算するクラス"""
    
    def __init__(self):
        self.ref_data = None
        self.time_stats = {}  # (venue, distance, course) -> (mean, std, base_mean, base_l3f)
        self._load_reference_data()
    
    def _load_reference_data(self):
        """parquetファイルから基準データを読み込み"""
        parquet_path = os.path.join(PARQUET_DIR, "nar_races_combined.parquet")
        
        if not os.path.exists(parquet_path):
            logger.warning(f"Reference data not found: {parquet_path}")
            return
        
        try:
            self.ref_data = pd.read_parquet(parquet_path)
            logger.info(f"Loaded {len(self.ref_data)} reference records from parquet")
            self._compute_statistics()
        except Exception as e:
            logger.warning(f"Failed to load reference data: {e}")
    
    def _compute_statistics(self):
        """条件別の統計を計算"""
        if self.ref_data is None or len(self.ref_data) == 0:
            return
        
        # 有効なレコードのみ使用
        valid = self.ref_data[
            (self.ref_data['time_sec'].notna()) & 
            (self.ref_data['l3f_sec'].notna())
        ].copy()
        
        # 会場・距離・コース種別でグループ化
        for (venue, dist, course), group in valid.groupby(['venue_name', 'distance', 'course_type']):
            if len(group) < 5:  # 最低5レコード必要
                continue
            
            key = (venue, int(dist), course)
            
            # 全体平均
            mean_time = group['time_sec'].mean()
            std_time = group['time_sec'].std()
            mean_l3f = group['l3f_sec'].mean()
            std_l3f = group['l3f_sec'].std()
            
            # 基準タイム（1-3着の平均）
            top3 = group[group['rank'].astype(str).str.isdigit()]
            top3 = top3[top3['rank'].astype(int) <= 3]
            
            base_time = top3['time_sec'].mean() if len(top3) >= 3 else mean_time
            base_l3f = top3['l3f_sec'].mean() if len(top3) >= 3 else mean_l3f
            
            self.time_stats[key] = {
                'mean_time': mean_time,
                'std_time': std_time if std_time > 0 else 1.0,
                'mean_l3f': mean_l3f,
                'std_l3f': std_l3f if std_l3f > 0 else 1.0,
                'base_time': base_time,
                'base_l3f': base_l3f,
                'count': len(group)
            }
        
        logger.info(f"Computed statistics for {len(self.time_stats)} venue/distance/course combinations")
    
    def get_metrics(self, venue: str, distance: int, course_type: str, 
                    time_sec: float, l3f_sec: float, condition: str = None) -> Dict:
        """基準データを使ってメトリクスを計算"""
        result = {
            '平均t': None, '平均場t': None, 
            '基準t': None, '基準場t': None,
            '基準3F': None, '基準場3F': None,
            '平t差': None, '平場t差': None,
            '基t差': None, '基場t差': None,
            '基3F差': None, '基場3F差': None,
            'タイム指数': None, '上り指数': None,
            '馬場差': None
        }
        
        key = (venue, distance, course_type)
        
        if key not in self.time_stats:
            return result
        
        stats = self.time_stats[key]
        
        # 基準値を設定
        result['平均t'] = round(stats['mean_time'], 1)
        result['平均場t'] = round(stats['mean_time'], 1)  # 同じ会場なので
        result['基準t'] = round(stats['base_time'], 1)
        result['基準場t'] = round(stats['base_time'], 1)
        result['基準3F'] = round(stats['base_l3f'], 1)
        result['基準場3F'] = round(stats['base_l3f'], 1)
        
        if time_sec:
            # 差分計算
            result['平t差'] = round(time_sec - stats['mean_time'], 1)
            result['平場t差'] = round(time_sec - stats['mean_time'], 1)
            result['基t差'] = round(time_sec - stats['base_time'], 1)
            result['基場t差'] = round(time_sec - stats['base_time'], 1)
            
            # タイム指数: 50 + 10 * (基準タイム - 実績) / 標準偏差
            # 速いほど高い値
            result['タイム指数'] = round(50 + 10 * (stats['base_time'] - time_sec) / stats['std_time'], 1)
        
        if l3f_sec:
            result['基3F差'] = round(l3f_sec - stats['base_l3f'], 1)
            result['基場3F差'] = round(l3f_sec - stats['base_l3f'], 1)
            
            # 上り指数
            result['上り指数'] = round(50 + 10 * (stats['base_l3f'] - l3f_sec) / stats['std_l3f'], 1)
        
        # 馬場差（今回は基準タイムと平均タイムの差として定義）
        result['馬場差'] = round(stats['base_time'] - stats['mean_time'], 1)
        
        return result

class NarNetkeibaAnalyzer:
    def __init__(self, target_date: str, venue_name: str):
        self.target_date = target_date  # YYYYMMDD
        self.venue_name = venue_name
        self.venue_code = NAR_VENUE_NAME_TO_CODE.get(venue_name)
        
        if not self.venue_code:
            raise ValueError(f"Unknown venue name: {venue_name}. Available: {list(NAR_VENUE_NAME_TO_CODE.keys())}")
        
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
        })
        
        self.driver = None
        self.horse_cache = {}
        
        # 基準データから指数計算
        self.metric_calculator = NarMetricCalculator()
        
        os.makedirs(HTML_DIR, exist_ok=True)
        os.makedirs(os.path.join(HTML_DIR, "horse"), exist_ok=True)

    def init_driver(self):
        if self.driver:
            return
        logger.info("Initializing Selenium Driver...")
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument("--log-level=3")
        options.page_load_strategy = 'eager'
        self.driver = webdriver.Chrome(options=options)
        self.driver.set_page_load_timeout(30)

    def close_driver(self):
        if self.driver:
            self.driver.quit()
            self.driver = None

    def scrape_race_list(self) -> List[Dict]:
        """レース一覧を取得（requestsベース - 高速）"""
        # 1Rにアクセスしてレース番号リンクを取得
        base_race_id = f"{self.target_date[:4]}{self.venue_code}{self.target_date[4:]}01"
        url = f"https://nar.netkeiba.com/race/shutuba.html?race_id={base_race_id}"
        
        logger.info(f"Fetching race list from: {url}")
        
        try:
            resp = self.session.get(url, timeout=15)
            # nar.netkeiba.comはEUC-JP（meta charsetで確認済み）
            resp.encoding = 'euc-jp'
            soup = BeautifulSoup(resp.text, 'html.parser')
            
            race_ids = []
            num_wrap = soup.find('div', class_='RaceNumWrap')
            if num_wrap:
                links = num_wrap.find_all('a', href=True)
                for link in links:
                    href = link['href']
                    if 'race_id=' in href:
                        rid = href.split('race_id=')[1].split('&')[0]
                        rname = link.get_text(strip=True)
                        race_ids.append({'id': rid, 'name': rname})
            
            race_ids.sort(key=lambda x: x['id'])
            logger.info(f"Found {len(race_ids)} races for {self.venue_name}")
            
            if race_ids:
                return race_ids
            
        except Exception as e:
            logger.warning(f"requests failed for race list: {e}")
        
        # Fallback: 1-12R
        logger.info("Using fallback race list (1-12R)")
        return [{'id': f"{self.target_date[:4]}{self.venue_code}{self.target_date[4:]}{i:02d}", 
                 'name': f"{i}R"} for i in range(1, 13)]

    def scrape_shutuba(self, race_id: str, use_selenium_for_odds: bool = False) -> Dict:
        """出馬表をスクレイピング（requestsベース - 高速）
        
        Args:
            race_id: レースID
            use_selenium_for_odds: オッズ取得にSeleniumを使用するか（デフォルト: False）
        """
        url = f"https://nar.netkeiba.com/race/shutuba.html?race_id={race_id}"
        logger.info(f"Scraping shutuba for {race_id}")
        
        try:
            # まずrequestsで基本情報を取得
            resp = self.session.get(url, timeout=15)
            # nar.netkeiba.comはEUC-JP（meta charsetで確認済み）
            resp.encoding = 'euc-jp'
            soup = BeautifulSoup(resp.text, 'html.parser')
            
            # レース情報
            race_name = soup.select_one('.RaceName')
            race_name_text = race_name.text.strip() if race_name else "Unknown"
            
            race_data01 = soup.select_one('.RaceData01')
            race_data_text = race_data01.text.strip() if race_data01 else ""
            
            # コース種別・距離の抽出
            course_type = "ダート"
            distance = ""
            if "芝" in race_data_text:
                course_type = "芝"
            
            dist_match = re.search(r'(\d+)m', race_data_text)
            if dist_match:
                distance = dist_match.group(1)
            
            # 馬場状態
            condition = "良"
            for c in ["良", "稍", "重", "不"]:
                if c in race_data_text:
                    condition = c if c != "不" else "不良"
                    if c == "稍":
                        condition = "稍重"
                    break
            
            # RaceData02からの情報取得
            race_data02 = soup.select_one('.RaceData02')
            heads = ""
            kai = ""
            nichi = ""
            if race_data02:
                spans = race_data02.find_all('span')
                for span in spans:
                    text = span.text.strip()
                    if "回" in text:
                        match = re.search(r'(\d+)回', text)
                        if match:
                            kai = match.group(1)
                    elif "日目" in text:
                        match = re.search(r'(\d+)日目', text)
                        if match:
                            nichi = match.group(1)
                    elif "頭" in text:
                        match = re.search(r'(\d+)頭', text)
                        if match:
                            heads = match.group(1)
            
            race_info = {
                'race_id': race_id,
                'race_name': race_name_text,
                'course': course_type,
                'distance': distance,
                'condition': condition,
                'heads': heads,
                'kai': kai,
                'nichi': nichi
            }
            
            # 馬リストの取得
            horses = []
            rows = soup.select('tr.HorseList')
            
            for row in rows:
                h_data = self._parse_horse_row(row, race_info)
                if h_data:
                    horses.append(h_data)
            
            # オッズが必要でSeleniumを使う場合のみ動的取得
            if use_selenium_for_odds and horses:
                try:
                    self._fetch_odds_with_selenium(race_id, horses)
                except Exception as e:
                    logger.warning(f"Failed to fetch odds with Selenium: {e}")
            
            return {'race_info': race_info, 'horses': horses}
            
        except Exception as e:
            logger.error(f"Error scraping shutuba: {e}")
            return {'race_info': {}, 'horses': []}
    
    def _fetch_odds_with_selenium(self, race_id: str, horses: List[Dict]):
        """Seleniumでオッズを取得（オプション）"""
        url = f"https://nar.netkeiba.com/race/shutuba.html?race_id={race_id}"
        
        self.init_driver()
        self.driver.get(url)
        time.sleep(2)
        
        soup = BeautifulSoup(self.driver.page_source, 'html.parser')
        rows = soup.select('tr.HorseList')
        
        # 馬番をキーにしたマップを作成
        horse_map = {h['出走馬番']: h for h in horses}
        
        for row in rows:
            umaban = row.select_one('td[class^="Umaban"]')
            if not umaban:
                continue
            umaban_text = umaban.text.strip()
            
            if umaban_text in horse_map:
                tds = row.find_all('td')
                for td in tds:
                    classes = td.get('class', [])
                    if 'Popular' in classes and 'Txt_R' in classes:
                        try:
                            horse_map[umaban_text]['現オッズ'] = float(td.text.strip())
                        except:
                            pass
                    elif 'Popular' in classes and 'Txt_C' in classes:
                        try:
                            horse_map[umaban_text]['現人気'] = int(td.text.strip())
                        except:
                            pass

    def _parse_horse_row(self, row, race_info: Dict) -> Optional[Dict]:
        """馬の行をパース"""
        try:
            h_data = {
                '出走枠番': '',
                '出走馬番': '',
                '乗り替わり': '',
                'グループNo': '',
                '馬名': '',
                '性齢': '',
                '斤量': '',
                '騎手': '',
                '厩舎': '',
                '現オッズ': None,
                '現人気': None,
                'horse_id': ''
            }
            
            # 枠番
            waku = row.select_one('td[class^="Waku"]')
            h_data['出走枠番'] = waku.text.strip() if waku else ""
            
            # 馬番
            umaban = row.select_one('td[class^="Umaban"]')
            h_data['出走馬番'] = umaban.text.strip() if umaban else ""
            
            # 馬名・馬ID
            name_link = row.select_one('.HorseName a, .HorseInfo a')
            if name_link:
                h_data['馬名'] = name_link.text.strip()
                href = name_link.get('href', '')
                if '/horse/' in href:
                    h_data['horse_id'] = href.split('/horse/')[-1].replace('/', '')
            
            # 騎手
            jockey = row.select_one('.Jockey a')
            h_data['騎手'] = jockey.text.strip() if jockey else ""
            
            # 調教師
            trainer = row.select_one('.Trainer a')
            trainer_text = trainer.text.strip() if trainer else ""
            # "大井月岡健二" -> "月岡健二" (場所名を除去)
            for venue in NAR_VENUE_MAP.values():
                if trainer_text.startswith(venue):
                    trainer_text = trainer_text[len(venue):]
                    break
            h_data['厩舎'] = trainer_text
            
            # 全セルを取得して詳細を抽出
            tds = row.find_all('td')
            if len(tds) >= 6:
                # 性齢 (index 4)
                h_data['性齢'] = tds[4].text.strip() if len(tds) > 4 else ""
                # 斤量 (index 5)
                h_data['斤量'] = tds[5].text.strip() if len(tds) > 5 else ""
            
            # オッズ・人気（Popular クラスから）
            for td in tds:
                classes = td.get('class', [])
                if 'Popular' in classes and 'Txt_R' in classes:
                    # オッズ
                    try:
                        h_data['現オッズ'] = float(td.text.strip())
                    except:
                        pass
                elif 'Popular' in classes and 'Txt_C' in classes:
                    # 人気
                    try:
                        h_data['現人気'] = int(td.text.strip())
                    except:
                        pass
            
            return h_data
            
        except Exception as e:
            logger.error(f"Error parsing horse row: {e}")
            return None

    def get_horse_history(self, horse_id: str) -> List[Dict]:
        """馬の過去成績を取得"""
        if horse_id in self.horse_cache:
            return self.horse_cache[horse_id]
        
        local_path = os.path.join(HTML_DIR, "horse", f"{horse_id}_perf.bin")
        html_content = None
        
        # キャッシュチェック（7日以内）
        if os.path.exists(local_path):
            mtime = os.path.getmtime(local_path)
            if datetime.now() - datetime.fromtimestamp(mtime) < timedelta(days=7):
                with open(local_path, 'rb') as f:
                    html_content = f.read()
        
        if not html_content:
            url = f"https://db.netkeiba.com/horse/result/{horse_id}/"
            try:
                time.sleep(0.5)  # 礼儀正しい遅延
                resp = self.session.get(url, timeout=10)
                resp.encoding = 'euc-jp'
                html_content = resp.content
                with open(local_path, 'wb') as f:
                    f.write(html_content)
            except Exception as e:
                logger.error(f"Error scraping history for {horse_id}: {e}")
                return []
        
        history = self._parse_horse_history(html_content, horse_id)
        self.horse_cache[horse_id] = history
        return history

    def _parse_horse_history(self, html_content: bytes, horse_id: str) -> List[Dict]:
        """過去成績をパース（完全版）"""
        try:
            soup = BeautifulSoup(html_content, 'html.parser', from_encoding='euc-jp')
            table = soup.find('table', class_='db_h_race_results')
            if not table:
                return []
            
            tbody = table.find('tbody')
            if not tbody:
                return []
            
            history = []
            rows = tbody.find_all('tr')
            
            for row in rows:
                cols = row.find_all('td')
                if len(cols) < 29:
                    continue
                
                data = [c.text.strip() for c in cols]
                
                # race_id の取得
                race_id_full = ""
                race_link = cols[4].find('a')
                if race_link and 'href' in race_link.attrs:
                    href = race_link['href']
                    if '/race/' in href:
                        race_id_full = re.search(r'/race/(\d+)', href)
                        race_id_full = race_id_full.group(1) if race_id_full else ""
                
                # 距離からコース種別を抽出
                dist_raw = data[14] if len(data) > 14 else ""
                course = "ダート"
                distance = ""
                if dist_raw:
                    if "芝" in dist_raw:
                        course = "芝"
                    elif "障" in dist_raw:
                        course = "障害"
                    distance = re.sub(r'\D', '', dist_raw)
                
                # 通過順位の分解
                # db.netkeibaのテーブルにはタイム指数マスター系の隠しカラム(4列)があるため、
                # 実際のインデックスは+4オフセット
                passing = data[25] if len(data) > 25 else ""
                p1, p2, p3, p4 = "", "", "", ""
                if passing:
                    parts = passing.split('-')
                    if len(parts) >= 1: p1 = parts[0]
                    if len(parts) >= 2: p2 = parts[1]
                    if len(parts) >= 3: p3 = parts[2]
                    if len(parts) >= 4: p4 = parts[3]
                
                # ペースの分解
                pace = data[26] if len(data) > 26 else ""
                pace1, pace2 = "", ""
                if pace and '-' in pace:
                    pp = pace.split('-')
                    if len(pp) == 2:
                        pace1, pace2 = pp[0], pp[1]
                
                # RPCI計算（前半 / 全体 * 100）
                rpci_val = None
                if pace1 and pace2:
                    try:
                        p1f, p2f = float(pace1), float(pace2)
                        if p1f + p2f > 0:
                            rpci_val = round(p1f / (p1f + p2f) * 100, 1)
                    except:
                        pass
                
                # 馬体重の分解
                weight_raw = data[28] if len(data) > 28 else ""
                weight, change = "", ""
                if weight_raw:
                    match = re.match(r'(\d+)\(([+-]?\d+)\)', weight_raw)
                    if match:
                        weight = match.group(1)
                        change = match.group(2)
                    else:
                        weight = re.sub(r'\D', '', weight_raw)
                
                # タイムの秒換算
                time_str = data[18] if len(data) > 18 else ""
                time_sec = self._parse_time(time_str)
                
                # 上りの秒換算
                l3f_str = data[27] if len(data) > 27 else ""
                l3f_sec = self._parse_time(l3f_str)
                
                # レース名から回・日を抽出
                kai = ""
                nichi = ""
                if race_id_full and len(race_id_full) >= 10:
                    try:
                        kai = str(int(race_id_full[6:8]))
                        nichi = str(int(race_id_full[8:10]))
                    except:
                        pass
                
                race_data = {
                    '日付': data[0],
                    '場所': re.sub(r'\d', '', data[1]) if len(data) > 1 else "",
                    '回': kai,
                    '日': nichi,
                    'ｺｰｽ': course,
                    '距離': distance,
                    'R': data[3] if len(data) > 3 else "",
                    '馬場': data[16] if len(data) > 16 else "",
                    '天気': data[2] if len(data) > 2 else "",
                    '頭数': data[6] if len(data) > 6 else "",
                    '枠番': data[7] if len(data) > 7 else "",
                    '馬番': data[8] if len(data) > 8 else "",
                    '馬名': "",  # 自分自身なので空
                    '斤量': data[13] if len(data) > 13 else "",
                    '騎手': data[12] if len(data) > 12 else "",
                    '厩舎': "",  # 履歴には含まれない
                    'ﾀｲﾑ': time_str,
                    'タイム秒': time_sec,
                    '着差': data[19] if len(data) > 19 else "",
                    '人気': data[10] if len(data) > 10 else "",
                    'ｵｯｽﾞ': data[9] if len(data) > 9 else "",
                    '上り': l3f_str,
                    '上り秒': l3f_sec,
                    'ﾍﾟｰｽ1': pace1,
                    'ﾍﾟｰｽ2': pace2,
                    '通過': passing,
                    '1C': p1, '2C': p2, '3C': p3, '4C': p4,
                    '着順': data[11] if len(data) > 11 else "",
                    '馬体重': weight,
                    '増減': change,
                    'レース名': data[4] if len(data) > 4 else "",
                    '性': "",
                    '年齢': "",
                    # 指数系（地方競馬用基準データがないためNone）
                    '平均t': None, '平均場t': None, '基準t': None, '基準場t': None,
                    '基準3F': None, '基準場3F': None,
                    '平t差': None, '平場t差': None, '基t差': None, '基場t差': None,
                    '基3F差': None, '基場3F差': None,
                    'タイム指数': None, '上り指数': None, '馬場差': None, 'RPCI': rpci_val,
                    'horse_id': horse_id,
                    'race_id': race_id_full
                }
                
                # RPCIの計算（可能な場合）
                if pace1 and l3f_sec:
                    try:
                        first_3f = float(pace1)
                        if first_3f > 0 and l3f_sec > 0:
                            race_data['RPCI'] = round((first_3f / l3f_sec) * 50, 1)
                    except:
                        pass
                
                history.append(race_data)
            
            return history
            
        except Exception as e:
            logger.error(f"Error parsing history: {e}")
            return []

    def _parse_time(self, time_str: str) -> Optional[float]:
        """タイム文字列を秒に変換"""
        if not time_str:
            return None
        try:
            if ':' in time_str:
                parts = time_str.split(':')
                return float(parts[0]) * 60 + float(parts[1])
            return float(time_str)
        except:
            return None

    def calculate_simple_metrics(self, history: List[Dict]) -> List[Dict]:
        """簡易指数計算（地方競馬用）"""
        if not history:
            return history
        
        # 距離・コース別にグループ化
        time_groups = {}  # タイム用
        l3f_groups = {}   # 上り用
        
        for race in history:
            dist = race.get('距離', '')
            course = race.get('ｺｰｽ', '')
            time_sec = race.get('タイム秒')
            l3f_sec = race.get('上り秒')
            
            if dist and course:
                key = (dist, course)
                
                if time_sec:
                    if key not in time_groups:
                        time_groups[key] = []
                    time_groups[key].append(time_sec)
                
                if l3f_sec:
                    if key not in l3f_groups:
                        l3f_groups[key] = []
                    l3f_groups[key].append(l3f_sec)
        
        # 平均・標準偏差の計算
        def calc_stats(groups):
            stats = {}
            for key, values in groups.items():
                if len(values) >= 2:
                    mean = sum(values) / len(values)
                    variance = sum((v - mean) ** 2 for v in values) / len(values)
                    std = variance ** 0.5
                    if std > 0:
                        stats[key] = (mean, std)
            return stats
        
        time_stats = calc_stats(time_groups)
        l3f_stats = calc_stats(l3f_groups)
        
        # 指数の計算
        for race in history:
            dist = race.get('距離', '')
            course = race.get('ｺｰｽ', '')
            time_sec = race.get('タイム秒')
            l3f_sec = race.get('上り秒')
            
            key = (dist, course)
            
            # タイム指数: 50 + 10 * (平均 - 実績) / 標準偏差
            # 速い（小さい）ほど高い値になる
            if key in time_stats and time_sec:
                mean, std = time_stats[key]
                race['タイム指数'] = round(50 + 10 * (mean - time_sec) / std, 1)
            
            # 上り指数: 同様に計算（上りタイム専用の統計を使用）
            if key in l3f_stats and l3f_sec:
                mean, std = l3f_stats[key]
                race['上り指数'] = round(50 + 10 * (mean - l3f_sec) / std, 1)
        
        return history

    def run(self):
        """メイン実行フロー"""
        try:
            race_list = self.scrape_race_list()
            
            all_data = {}
            for race in race_list:
                rid = race['id']
                rname = race['name']
                
                logger.info(f"Processing {rname} ({rid})...")
                
                race_data = self.scrape_shutuba(rid)
                if not race_data['horses']:
                    logger.warning(f"No horses found for {rname}")
                    continue
                
                # 各馬の履歴を取得
                for horse in race_data['horses']:
                    hid = horse.get('horse_id')
                    if hid:
                        history = self.get_horse_history(hid)
                        history = self.calculate_simple_metrics(history)
                        horse['history'] = history
                    else:
                        horse['history'] = []
                
                all_data[rid] = race_data
            
            self.generate_excel(all_data)
            
        finally:
            self.close_driver()

    def generate_excel(self, all_data: Dict):
        """Excel出力（中央競馬と完全に同じ構造）"""
        output_dir = os.path.join(PROJECT_ROOT, "outputs", "reports", "nar")
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, f"race_analysis_{self.target_date}_{self.venue_name}.xlsx")
        
        logger.info(f"Generating Excel: {output_file}")
        
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        sorted_rids = sorted(all_data.keys())
        
        # === RaceInfoシートを作成（create_interactive_chartsとの互換性確保）===
        ri_ws = wb.create_sheet('RaceInfo')
        ri_ws.append(['race_num', 'race_name', 'course_info', 'venue', 'date'])
        for rid in sorted_rids:
            race_data = all_data[rid]
            race_info = race_data['race_info']
            race_num = rid[-2:]
            course = race_info.get('course', 'ダート')
            distance = race_info.get('distance', '')
            course_info = f"{course}{distance}m" if distance else course
            ri_ws.append([
                race_num,
                race_info.get('race_name', ''),
                course_info,
                self.venue_name,
                f"{self.target_date[:4]}/{self.target_date[4:6]}/{self.target_date[6:]}"
            ])
        
        for rid in sorted_rids:
            race_data = all_data[rid]
            race_info = race_data['race_info']
            horses = race_data['horses']
            
            # シート名（Race_01, Race_02, ...）
            race_num = rid[-2:]
            sheet_name = f"Race_{race_num}"
            
            # === AI分析シートを追加（中央競馬と同じ形式）===
            ai_sheet_name = f"AI分析_{race_num}R"
            ai_ws = wb.create_sheet(ai_sheet_name)
            
            # Row1: レース名
            race_name = race_info.get('race_name', '')
            ai_ws.cell(row=1, column=1, value=f"{race_num}R {race_name}")
            
            # Row2: ヘッダー
            ai_headers = ['馬番', '馬名', '騎手', '現オッズ', '人気', '過去走', '平均着順']
            for col_idx, header in enumerate(ai_headers, 1):
                ai_ws.cell(row=2, column=col_idx, value=header)
            
            # Row3以降: 各馬のデータ
            horses_sorted = sorted(horses, key=lambda x: int(x.get('出走馬番', 999)) if x.get('出走馬番', '').isdigit() else 999)
            for row_idx, horse in enumerate(horses_sorted, 3):
                history = horse.get('history', [])
                past_runs = len(history)
                
                # 平均着順を計算
                avg_rank = ''
                ranks = []
                for h in history:
                    rank_str = h.get('着順', '')
                    if rank_str and rank_str.isdigit():
                        ranks.append(int(rank_str))
                if ranks:
                    avg_rank = round(sum(ranks) / len(ranks), 1)
                
                ai_ws.cell(row=row_idx, column=1, value=horse.get('出走馬番', ''))
                ai_ws.cell(row=row_idx, column=2, value=horse.get('馬名', ''))
                ai_ws.cell(row=row_idx, column=3, value=horse.get('騎手', ''))
                ai_ws.cell(row=row_idx, column=4, value=horse.get('現オッズ'))
                ai_ws.cell(row=row_idx, column=5, value=horse.get('現人気'))
                ai_ws.cell(row=row_idx, column=6, value=past_runs)
                ai_ws.cell(row=row_idx, column=7, value=avg_rank)
            
            # === Race_XX シート ===
            ws = wb.create_sheet(sheet_name)
            
            # ヘッダー行
            ws.append(EXCEL_COLUMNS)
            
            # 各馬の履歴を出力
            for horse in horses:
                history = horse.get('history', [])
                
                if not history:
                    # 履歴がない場合は1行だけ出力
                    row = self._make_row(horse, {}, race_info)
                    ws.append(row)
                else:
                    for hist in history:
                        row = self._make_row(horse, hist, race_info)
                        ws.append(row)
        
        wb.save(output_file)
        logger.info(f"Saved report to {output_file}")

    def _make_row(self, horse: Dict, hist: Dict, race_info: Dict) -> List:
        """Excel行データを生成"""
        # 性齢から性・年齢を分離
        seirei = horse.get('性齢', '')
        sei = seirei[0] if seirei else ""
        age = seirei[1:] if len(seirei) > 1 else ""
        
        # metric_calculatorを使って基準タイムを計算
        venue = hist.get('場所', '')
        distance = hist.get('距離', '')
        course = hist.get('ｺｰｽ', '')
        time_sec = hist.get('タイム秒')
        l3f_sec = hist.get('上り秒')
        
        metrics = {}
        if venue and distance and course:
            try:
                dist_int = int(distance)
                metrics = self.metric_calculator.get_metrics(
                    venue, dist_int, course, time_sec, l3f_sec
                )
            except (ValueError, TypeError):
                pass
        
        row = [
            horse.get('出走枠番', ''),
            horse.get('出走馬番', ''),
            horse.get('乗り替わり', ''),
            horse.get('グループNo', ''),
            hist.get('日付', ''),
            hist.get('場所', ''),
            hist.get('回', ''),
            hist.get('日', ''),
            hist.get('ｺｰｽ', ''),
            hist.get('距離', ''),
            hist.get('R', ''),
            hist.get('馬場', ''),
            hist.get('天気', ''),
            hist.get('頭数', ''),
            hist.get('枠番', ''),
            hist.get('馬番', ''),
            horse.get('馬名', ''),
            hist.get('斤量', '') or horse.get('斤量', ''),
            hist.get('騎手', '') or horse.get('騎手', ''),
            hist.get('厩舎', '') or horse.get('厩舎', ''),
            hist.get('ﾀｲﾑ', ''),
            hist.get('タイム秒'),
            hist.get('着差', ''),
            hist.get('人気', ''),
            hist.get('ｵｯｽﾞ', ''),
            hist.get('上り', ''),
            hist.get('上り秒'),
            hist.get('ﾍﾟｰｽ1', ''),
            hist.get('ﾍﾟｰｽ2', ''),
            hist.get('通過', ''),
            hist.get('1C', ''),
            hist.get('2C', ''),
            hist.get('3C', ''),
            hist.get('4C', ''),
            hist.get('着順', ''),
            hist.get('馬体重', ''),
            hist.get('増減', ''),
            hist.get('レース名', ''),
            sei,
            age,
            # parquetベースの基準タイム
            metrics.get('平均t') or hist.get('平均t'),
            metrics.get('平均場t') or hist.get('平均場t'),
            metrics.get('基準t') or hist.get('基準t'),
            metrics.get('基準場t') or hist.get('基準場t'),
            metrics.get('基準3F') or hist.get('基準3F'),
            metrics.get('基準場3F') or hist.get('基準場3F'),
            metrics.get('平t差') or hist.get('平t差'),
            metrics.get('平場t差') or hist.get('平場t差'),
            metrics.get('基t差') or hist.get('基t差'),
            metrics.get('基場t差') or hist.get('基場t差'),
            metrics.get('基3F差') or hist.get('基3F差'),
            metrics.get('基場3F差') or hist.get('基場3F差'),
            metrics.get('タイム指数') or hist.get('タイム指数'),
            metrics.get('上り指数') or hist.get('上り指数'),
            metrics.get('馬場差') or hist.get('馬場差'),
            hist.get('RPCI'),
            horse.get('horse_id', ''),
            hist.get('race_id', '')
        ]
        return row


def main():
    parser = argparse.ArgumentParser(description="NAR Race Analyzer V2")
    parser.add_argument("--date", required=True, help="YYYYMMDD")
    parser.add_argument("--venue", required=True, help="Venue Name (e.g. 大井)")
    args = parser.parse_args()
    
    analyzer = NarNetkeibaAnalyzer(args.date, args.venue)
    analyzer.run()


if __name__ == "__main__":
    main()
