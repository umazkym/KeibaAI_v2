#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
インタラクティブ競馬分析アプリ（スマホ最適化版 v8.7: Fix Waku Colors）

- Improve:
    1. 【ソート】ボタンを全表示（スクロールなし、Wrap）。
    2. 【比較】対戦マトリクス（早見表）＋詳細アコーディオン。
    3. 【チャート】
        - ③ 末脚(縦) vs 初期位置(横) に変更。
        - ④ 上がり指数(縦) vs RPCI(横) を追加。
    4. 【ロジック】位置取り計算を「最初に取得できる通過順 / 頭数」に変更。
    5. 【データ】比較詳細の「通過」が表示されないデータ不備を修正（1C~4Cから合成）。
    6. 【バグ修正】対戦履歴の馬番識別を「馬名」ベースに変更（過去の馬番混入防止）。
    7. 【バグ修正】枠番の色分けをデータ（出走枠番）準拠に変更（変則頭数等に対応）。
"""

import argparse
import os
from openpyxl import load_workbook
import json
import logging
import itertools
from collections import defaultdict
import re
import statistics
from datetime import datetime

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

WAKU_COLORS = {
    1: '#F0F0F0', 2: '#222222', 3: '#EF5350',
    4: '#42A5F5', 5: '#FBC02D', 6: '#66BB6A',
    7: '#FFA726', 8: '#F48FB1',
}

WAKU_TEXT_COLORS = {
    1: '#333', 2: '#FFF', 3: '#FFF',
    4: '#FFF', 5: '#333', 6: '#FFF',
    7: '#FFF', 8: '#333',
}

DISPLAY_COLUMNS = [
    '枠', '番',
    '日付', '場所', 'ｺｰｽ', '距離', 'R', '馬場', '天気', '頭数',
    '枠番', '馬番', '馬名', '斤量', '騎手',
    'タイム', '着差', '人気', 'ｵｯｽﾞ',
    '上がり', 'ﾍﾟｰｽ1', 'ﾍﾟｰｽ2', '通過', '1角', '2角', '3角', '4角',
    '着順', '体重', '増減', 'レース名',
    '平t差', '基t差', 'T指数', 'L指数', '馬場差', 'RPCI',
    '指数信頼度', '指数N', '補正情報'
]

COL_MAP = {
    '出走枠番': '枠', '出走馬番': '番',
    'タイム秒': 'タイム', '上り': '上がり',
    '1C': '1角', '2C': '2角', '3C': '3角', '4C': '4角',
    '馬体重': '体重', 'タイム指数': 'T指数', '上り指数': 'L指数'
}

def normalize_text(text):
    if not text: return ""
    t = str(text).strip()
    t = t.translate(str.maketrans('０１２３４５６７８９', '0123456789'))
    return t

def parse_report_date(value):
    if value is None or value == "":
        return None
    s = normalize_text(value)
    if s.endswith(".0"):
        s = s[:-2]
    s = s.replace("/", "").replace("-", "")
    try:
        if len(s) == 8 and s.isdigit():
            return datetime.strptime(s, "%Y%m%d")
        return datetime.fromisoformat(str(value).replace("/", "-"))
    except Exception:
        return None

def is_pre_target_history(row, target_dt):
    if target_dt is None:
        return True
    row_dt = parse_report_date(row.get('日付'))
    return row_dt is not None and row_dt < target_dt

def get_waku_fallback(umaban):
    """Fallback calculation if data missing"""
    try:
        u = int(umaban)
        if u <= 2: return 1
        elif u <= 4: return 2
        elif u <= 6: return 3
        elif u <= 8: return 4
        elif u <= 10: return 5
        elif u <= 12: return 6
        elif u <= 14: return 7
        return 8
    except:
        return 1

def read_sheet_data(sheet):
    if sheet.max_row < 2: return [], []
    headers = [cell.value for cell in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        has_data = False
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                key = str(headers[i])
                val_str = str(val) if val is not None else ''
                row_data[key] = val_str
                if val_str: has_data = True
        if has_data:
            rows.append(row_data)
    return headers, rows

def make_pass_string(row):
    pass_vals = []
    for k in ['1角', '2角', '3角', '4角']:
        val = row.get(k, '')
        if val and str(val).strip() != '':
            pass_vals.append(str(val))
    if not pass_vals:
        return row.get('通過', '-')
    return '-'.join(pass_vals)

def calculate_dynamic_matchups(race_rows, name_map):
    """
    Raceデータから対戦履歴を動的生成
    name_map: { '馬名': current_umaban }
    """
    past_races = defaultdict(list)

    for row in race_rows:
        date = normalize_text(row.get('日付'))
        place = normalize_text(row.get('場所', ''))
        red = normalize_text(row.get('R', ''))
        racename = normalize_text(row.get('レース名', ''))

        if not date: continue

        place_clean = place.replace('競馬場', '').replace('中央', '').replace('地方', '')
        red_clean = red.replace('R', '')

        if red_clean:
            key = f"{date}_{place_clean}_{red_clean}"
        else:
            key = f"{date}_{racename}"

        past_races[key].append(row)

    matchups = defaultdict(list)
    matrix = defaultdict(dict)

    for key, participants in past_races.items():
        if len(participants) < 2: continue

        for p1, p2 in itertools.permutations(participants, 2):
            name1 = p1.get('馬名', '')
            name2 = p2.get('馬名', '')

            h1 = name_map.get(name1)
            h2 = name_map.get(name2)

            if not h1 or not h2: continue

            try:
                h1_num = int(h1)
                h2_num = int(h2)

                r1_str = p1.get('着順', '99')
                r2_str = p2.get('着順', '99')

                r1 = int(r1_str) if r1_str.isdigit() else 999
                r2 = int(r2_str) if r2_str.isdigit() else 999

                res_str = 'Draw'
                if r1 < r2: res_str = 'Win'
                elif r1 > r2: res_str = 'Lose'

                my_pass_str = make_pass_string(p1)
                op_pass_str = make_pass_string(p2)

                matchups[h1_num].append({
                    'opponent_num': h2_num,
                    'opponent_name': name2,
                    'date': p1.get('日付', ''),
                    'race_name': p1.get('レース名', ''),
                    'place': p1.get('場所', ''),
                    'course': p1.get('ｺｰｽ', ''),
                    'dist': p1.get('距離', ''),
                    'time': p1.get('タイム', ''),

                    'my_rank': p1.get('着順', '-'),
                    'my_idx': p1.get('T指数', '-'),
                    'my_pass': my_pass_str,
                    'my_l3f': p1.get('上がり', '-'),
                    'my_rpci': p1.get('RPCI', '-'),

                    'op_rank': p2.get('着順', '-'),
                    'op_time': p2.get('タイム', '-'),
                    'op_idx': p2.get('T指数', '-'),
                    'op_pass': op_pass_str,
                    'op_l3f': p2.get('上がり', '-'),
                    'op_rpci': p2.get('RPCI', '-'),

                    'result': res_str
                })

                cur = matrix[h1_num].get(h2_num, {'w':0, 'l':0, 'd':0})
                if res_str == 'Win': cur['w'] += 1
                elif res_str == 'Lose': cur['l'] += 1
                else: cur['d'] += 1
                matrix[h1_num][h2_num] = cur

            except Exception as e:
                continue

    return matchups, matrix

def create_interactive_charts(source_file: str, output_file: str = None):
    if not os.path.exists(source_file):
        return None

    if not output_file:
        base, _ = os.path.splitext(source_file)
        output_file = f"{base}_interactive.html"

    wb = load_workbook(source_file, data_only=True)

    all_data = {
        'races': {}, 'matchups': {}, 'matrix': {}, 'chart_data': {}, 'race_info': {}
    }

    # === RaceInfoシートからレースメタデータを読み込み ===
    race_meta = {}  # race_num -> {race_name, course_info, venue, date}
    if 'RaceInfo' in wb.sheetnames:
        ri_sheet = wb['RaceInfo']
        ri_headers, ri_rows = read_sheet_data(ri_sheet)
        for ri_row in ri_rows:
            rn = ri_row.get('race_num', '')
            if rn:
                race_meta[rn] = {
                    'race_name': ri_row.get('race_name', ''),
                    'course_info': ri_row.get('course_info', ''),
                    'venue': ri_row.get('venue', ''),
                    'date': ri_row.get('date', ''),
                }

    # === AI分析シートからのフォールバック ===
    for sn in wb.sheetnames:
        if sn.startswith('AI分析_') and sn.endswith('R'):
            ai_sheet = wb[sn]
            race_num_str = sn.replace('AI分析_', '').replace('R', '')
            if race_num_str not in race_meta:
                cell_val = ai_sheet.cell(row=1, column=1).value
                if cell_val:
                    race_meta[race_num_str] = {'race_name': str(cell_val), 'course_info': '', 'venue': '', 'date': ''}

    for sheet_name in wb.sheetnames:
        if sheet_name.startswith('Race_'):
            race_num = sheet_name.replace('Race_', '')
            headers, rows = read_sheet_data(wb[sheet_name])
            meta = race_meta.get(race_num, {})
            target_dt = parse_report_date(meta.get('date'))
            if target_dt is not None:
                before_count = len(rows)
                rows = [r for r in rows if is_pre_target_history(r, target_dt)]
                removed = before_count - len(rows)
                if removed:
                    logger.info(
                        f"Race_{race_num}: removed {removed} same-day/future rows from HTML data"
                    )

            processed_rows = []
            horse_chart_info = {}

            name_map = {}
            waku_map = {}

            # Pass 1: Build Maps
            for row in rows:
                name = row.get('馬名', '')
                umaban = row.get('出走馬番', '')
                waku = row.get('出走枠番', '') or row.get('枠', '')

                if name and umaban and str(umaban).isdigit():
                    u_int = int(umaban)
                    name_map[name] = u_int
                    if waku and str(waku).isdigit():
                        waku_map[u_int] = int(waku)

            # Pass 2: Process Rows
            for row in rows:
                new_row = {}
                for k, v in row.items():
                    nk = COL_MAP.get(k, k)
                    new_row[nk] = v

                if '通過' not in new_row or not new_row['通過']:
                    new_row['通過'] = make_pass_string(new_row)

                processed_rows.append(new_row)

                try:
                    name = row.get('馬名', '')
                    umaban = name_map.get(name, 0)

                    if umaban > 0:
                        if umaban not in horse_chart_info:
                            real_waku = waku_map.get(umaban, get_waku_fallback(umaban))
                            horse_chart_info[umaban] = {
                                'umaban': umaban, 'waku': real_waku, 'name': name, 'records': []
                            }
                        try:
                            def safe_float(r, k_list, d=0.0):
                                for k in k_list:
                                    if k in r and r[k]:
                                        try:
                                            return float(r[k])
                                        except:
                                            m = re.search(r'\d+(?:\.\d+)?', str(r[k]))
                                            if m:
                                                try:
                                                    return float(m.group(0))
                                                except:
                                                    pass
                                return d

                            time_idx = safe_float(row, ['タイム指数', 'T指数'])
                            l3f_idx = safe_float(row, ['上り指数', 'L指数'])
                            idx_conf = safe_float(row, ['指数信頼度'])
                            idx_n = safe_float(row, ['指数N'])
                            heads = safe_float(row, ['頭数'], 1)
                            c1 = safe_float(row, ['1C', '1角'])
                            c2 = safe_float(row, ['2C', '2角'])
                            c3 = safe_float(row, ['3C', '3角'])
                            c4 = safe_float(row, ['4C', '4角'])

                            valid_corners = [c for c in [c1, c2, c3, c4] if c > 0]
                            has_corners = len(valid_corners) > 0
                            corner_pairs = [('1C', c1), ('2C', c2), ('3C', c3), ('4C', c4)]
                            first_pair = next(((lbl, c) for lbl, c in corner_pairs if c > 0), None)
                            last_rank = next((c for c in [c4, c3, c2, c1] if c > 0), None)
                            avg_rank = statistics.mean(valid_corners) if valid_corners else heads / 2
                            initial_corner = first_pair[0] if first_pair else ''
                            initial_rank = first_pair[1] if first_pair else avg_rank

                            def norm_rank(v):
                                if heads > 1:
                                    return max(0.0, min(1.0, (v - 1) / (heads - 1)))
                                return 0.5

                            last_pos = round(norm_rank(last_rank if last_rank is not None else avg_rank), 3)
                            avg_pos = round(norm_rank(avg_rank), 3)
                            initial_pos = round(norm_rank(initial_rank), 3)

                            rpci = safe_float(row, ['RPCI'], 0.0)
                            has_rpci = rpci > 0
                            if not has_rpci: rpci = 50.0

                            if time_idx > 0:
                                horse_chart_info[umaban]['records'].append({
                                    'time_idx': round(time_idx, 1),
                                    'l3f_idx': round(l3f_idx, 1),
                                    'idx_conf': round(idx_conf, 2) if idx_conf else None,
                                    'idx_n': int(idx_n) if idx_n else None,
                                    'correction_note': row.get('補正情報', ''),
                                    'c1_ratio': initial_pos,
                                    'initial_pos': initial_pos,
                                    'initial_corner': initial_corner,
                                    'last_pos': last_pos,
                                    'avg_pos': avg_pos,
                                    'early_pos': initial_pos,
                                    'rpci': round(rpci, 1),
                                    'has_l3f': l3f_idx != 0.0,
                                    'has_corners': has_corners,
                                    'has_rpci': has_rpci,
                                    'rank': row.get('着順', ''),
                                    'date': row.get('日付', ''),
                                    'course': row.get('ｺｰｽ', '')
                                })
                        except: pass
                except: pass

            all_data['races'][race_num] = {'rows': processed_rows}
            all_data['chart_data'][race_num] = {'horses': list(horse_chart_info.values())}

            if rows:
                first = processed_rows[0]
                # RaceInfoシートからレース名を取得 (フォールバック: 最初の行のデータ)
                ri_name = meta.get('race_name', '')
                ri_course_info = meta.get('course_info', '')
                ri_venue = meta.get('venue', '')

                # course_infoからコース・距離を推定 (例: "ダ1800m(左)")
                ri_course = ''
                ri_dist = ''
                if ri_course_info:
                    import re as _re
                    if 'ダ' in ri_course_info:
                        ri_course = 'ダート'
                    elif '芝' in ri_course_info:
                        ri_course = '芝'
                    dm = _re.search(r'(\d{3,4})', ri_course_info)
                    if dm:
                        ri_dist = dm.group(1)

                # 頭数はデータ行のユニーク馬番数から推定
                unique_umaban = set()
                for r in rows:
                    u = r.get('出走馬番', r.get('番', ''))
                    if u and str(u).strip():
                        unique_umaban.add(str(u).strip())
                heads_count = str(len(unique_umaban)) if unique_umaban else first.get('頭数', '')

                all_data['race_info'][race_num] = {
                    'race_name': ri_name if ri_name else first.get('レース名', ''),
                    'date': meta.get('date', first.get('日付', '')),
                    'place': ri_venue if ri_venue else first.get('場所', ''),
                    'course': ri_course if ri_course else first.get('ｺｰｽ', ''),
                    'dist': ri_dist if ri_dist else first.get('距離', ''),
                    'weather': first.get('天気', ''),
                    'cond': first.get('馬場', ''),
                    'heads': heads_count,
                }
            else:
                meta = race_meta.get(race_num, {})
                ri_course_info = meta.get('course_info', '')
                ri_course = ''
                ri_dist = ''
                if ri_course_info:
                    if 'ダ' in ri_course_info:
                        ri_course = 'ダート'
                    elif '芝' in ri_course_info:
                        ri_course = '芝'
                    dm = re.search(r'(\d{3,4})', ri_course_info)
                    if dm:
                        ri_dist = dm.group(1)
                all_data['race_info'][race_num] = {
                    'race_name': meta.get('race_name', ''),
                    'date': meta.get('date', ''),
                    'place': meta.get('venue', ''),
                    'course': ri_course,
                    'dist': ri_dist,
                    'weather': '',
                    'cond': '',
                    'heads': '',
                }

            m_data, matrix_data = calculate_dynamic_matchups(processed_rows, name_map)
            all_data['matchups'][race_num] = m_data
            all_data['matrix'][race_num] = matrix_data

    # --- ADVANCED ANALYZER INJECTION ---
    all_data['pacing'] = {}
    all_data['course_stats'] = {}
    all_data['horse_profiles'] = {}
    all_data['summary_table'] = {}
    all_data['race_dynamics'] = {}
    try:
        try:
            from scripts.reports.advanced_analyzer import AdvancedPaceAnalyzer
        except ImportError:
            from advanced_analyzer import AdvancedPaceAnalyzer

        report_cutoffs = [
            parse_report_date(info.get('date'))
            for info in all_data.get('race_info', {}).values()
            if parse_report_date(info.get('date')) is not None
        ]
        analyzer_cutoff = min(report_cutoffs) if report_cutoffs else None
        analyzer = AdvancedPaceAnalyzer(cutoff_date=analyzer_cutoff)

        for race_num in all_data['races'].keys():
            race_info = all_data['race_info'].get(race_num, {})
            venue = race_info.get('place', '')
            dist_str = race_info.get('dist', '')
            course = race_info.get('course', '')
            cond = race_info.get('cond', '')

            pacing_data = {'clusters': []}

            if venue and dist_str.isdigit() and course:
                dist = int(dist_str)
                pacing_data['clusters'] = analyzer.get_course_lap_clusters(venue, dist, course)

                # ① コースラップ統計
                lap_stats = analyzer.get_course_lap_stats(venue, dist, course)
                # ② 位置取り×着順
                pos_stats = analyzer.get_course_position_stats(venue, dist, course)

                all_data['course_stats'][race_num] = {
                    'lap': lap_stats,
                    'pos': pos_stats,
                    'venue': venue, 'dist': dist, 'surface': course, 'cond': cond,
                }

                # ③④⑤⑥⑦ 馬別プロファイル
                rows = all_data['races'][race_num]['rows']
                # horse_idを馬名→番号のマッピングと共に収集
                horse_id_map = {}  # horse_id -> umaban
                horse_ids = []
                seen_horses = set()
                for r in rows:
                    hid = r.get('horse_id', '')
                    hname = r.get('馬名', '')
                    umaban = r.get('番', '0')
                    if hid and hname and hname not in seen_horses:
                        def _cid(v):
                            s = str(v).strip()
                            return s[:-2] if s.endswith('.0') else s
                        hid_clean = _cid(hid)
                        horse_id_map[hid_clean] = {'umaban': umaban, 'name': hname}
                        horse_ids.append(hid_clean)
                        seen_horses.add(hname)

                if horse_ids:
                    profiles = analyzer.get_horse_profiles(horse_ids, venue, dist, course)
                    # umaban注入
                    for hid, prof in profiles.items():
                        info = horse_id_map.get(hid, {})
                        prof['umaban'] = info.get('umaban', '0')
                    all_data['horse_profiles'][race_num] = profiles
                    all_data['summary_table'][race_num] = analyzer.get_summary_table(profiles)
                    # summary_tableにもumaban注入
                    for hid in all_data['summary_table'][race_num]:
                        info = horse_id_map.get(hid, {})
                        all_data['summary_table'][race_num][hid]['umaban'] = info.get('umaban', '0')

                    # ⑧ 展開予測
                    try:
                        dynamics = analyzer.predict_race_dynamics(
                            profiles, venue, dist, course, cond
                        )
                        all_data['race_dynamics'][race_num] = dynamics
                    except Exception as dyn_e:
                        print(f"Race dynamics prediction failed for {race_num}: {dyn_e}")

            all_data['pacing'][race_num] = pacing_data
    except Exception as e:
        import traceback
        print(f"Failed to run AdvancedPaceAnalyzer: {e}")
        traceback.print_exc()
    # -----------------------------------

    sorted_races = sorted(all_data['races'].keys(), key=lambda x: int(x) if x.isdigit() else 99)
    html = generate_html(source_file, sorted_races, all_data)

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(html)
    return output_file


def generate_html(source_file: str, races: list, all_data: dict) -> str:
    base_name = os.path.splitext(os.path.basename(source_file))[0]
    title = base_name.replace('race_analysis_', '')

    data_json = json.dumps(all_data, ensure_ascii=False)
    races_json = json.dumps(races, ensure_ascii=False)
    waku_color_json = json.dumps(WAKU_COLORS)
    waku_text_json = json.dumps(WAKU_TEXT_COLORS)
    cols_json = json.dumps(DISPLAY_COLUMNS, ensure_ascii=False)

    html = f'''<!DOCTYPE html>
<html lang="ja">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title}</title>
    <script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
    <style>
        :root {{
            --primary: #4F46E5; --bg: #F3F4F6; --card: #FFFFFF;
            --text: #111827; --text-light: #6B7280; --border: #E5E7EB;
            --safe-b: env(safe-area-inset-bottom);
        }}
        body {{
            margin: 0; padding: 0; font-family: sans-serif;
            background: var(--bg); color: var(--text);
            padding-bottom: calc(70px + var(--safe-b));
            -webkit-tap-highlight-color: transparent;
        }}
        .flex {{ display: flex; }} .flex-c {{ display: flex; flex-direction: column; }}
        .flex-ac {{ display: flex; align-items: center; }} .flex-jc {{ display: flex; justify-content: center; }}
        .flex-sb {{ display: flex; justify-content: space-between; }}
        .gap-1 {{ gap: 4px; }} .gap-2 {{ gap: 8px; }}
        .font-bold {{ font-weight: 700; }}
        .text-xs {{ font-size: 10px; }} .text-sm {{ font-size: 12px; }}
        .card {{ background: var(--card); border-radius: 12px; padding: 12px; margin-bottom: 10px; box-shadow: 0 2px 8px rgba(0,0,0,0.08); border: 1px solid rgba(0,0,0,0.04); }}
        .badge {{ background: #E5E7EB; padding: 3px 8px; border-radius: 8px; font-size: 11px; margin-right: 4px; font-weight: 500; }}

        .header {{ position: sticky; top: 0; background: linear-gradient(135deg, #4F46E5 0%, #6366F1 50%, #818CF8 100%); color: white; z-index: 100; padding: 10px; backdrop-filter: blur(10px); }}
        .race-tabs {{ overflow-x: auto; white-space: nowrap; padding-bottom: 2px; -webkit-overflow-scrolling: touch; }}
        .race-tab {{ display: inline-block; padding: 5px 12px; margin-right: 6px; background: rgba(255,255,255,0.15); border-radius: 12px; font-size: 12px; font-weight: bold; cursor: pointer; transition: all 0.2s; }}
        .race-tab:hover {{ background: rgba(255,255,255,0.3); }}
        .race-tab.active {{ background: white; color: var(--primary); box-shadow: 0 2px 6px rgba(0,0,0,0.15); }}
        .main {{ width: min(100% - 20px, 1280px); max-width: 1280px; padding: 10px; margin: 0 auto; box-sizing: border-box; }}

        body.chart-tab {{ padding-bottom: calc(96px + var(--safe-b)); }}
        .chart-card {{ padding: 10px; overflow: hidden; }}
        .chart-grid {{ display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 12px; align-items: stretch; }}
        .chart-panel {{ border: 1px solid var(--border); border-radius: 8px; background: #fff; padding: 8px; min-width: 0; }}
        .chart-panel.wide {{ grid-column: 1 / -1; }}
        .chart-panel-head {{ display: flex; align-items: center; justify-content: space-between; gap: 8px; margin-bottom: 6px; }}
        .chart-title {{ font-size: 12px; font-weight: 700; line-height: 1.35; }}
        .chart-meta {{ font-size: 10px; color: var(--text-light); white-space: nowrap; }}
        .chart-box {{ width: 100%; height: clamp(320px, 32vw, 460px); }}
        .chart-box.short {{ height: clamp(280px, 28vw, 400px); }}
        .chart-controls {{ display: flex; flex-wrap: wrap; gap: 6px; align-items: center; margin: 0 0 8px; }}
        .segmented {{ display: inline-flex; border: 1px solid var(--border); border-radius: 8px; overflow: hidden; background: #fff; }}
        .seg-btn {{ appearance: none; border: 0; border-right: 1px solid var(--border); background: #fff; color: var(--text-light); padding: 6px 9px; font-size: 11px; font-weight: 700; cursor: pointer; }}
        .seg-btn:last-child {{ border-right: 0; }}
        .seg-btn.active {{ background: var(--primary); color: #fff; }}
        .chart-summary {{ display: flex; flex-wrap: wrap; align-items: center; gap: 6px; color: var(--text-light); font-size: 11px; margin-left: auto; }}
        .empty-chart {{ display: flex; flex-direction: column; align-items: center; justify-content: center; min-height: 240px; gap: 10px; color: var(--text-light); font-size: 12px; text-align: center; }}
        .link-btn {{ appearance: none; border: 1px solid var(--primary); border-radius: 8px; color: var(--primary); background: #fff; padding: 7px 12px; font-size: 12px; font-weight: 700; cursor: pointer; }}
        .filter-sticky {{ border-radius: 0 0 10px 10px; box-shadow: 0 4px 12px rgba(15,23,42,0.08); }}
        .filter-chip, .preset-btn, .race-tab, .nav-btn, .seg-btn, .link-btn {{ touch-action: manipulation; }}
        .modebar {{ opacity: 0.72; }}
        .modebar:hover {{ opacity: 1; }}

        @media (max-width: 760px) {{
            body {{ padding-bottom: calc(92px + var(--safe-b)); }}
            .main {{ width: 100%; max-width: none; padding: 8px; }}
            .card {{ border-radius: 10px; padding: 10px; }}
            .chart-card {{ padding: 8px; }}
            .chart-grid {{ grid-template-columns: 1fr; gap: 10px; }}
            .chart-panel {{ padding: 6px; border-radius: 8px; }}
            .chart-panel-head {{ align-items: flex-start; flex-direction: column; gap: 2px; }}
            .chart-title {{ font-size: 12px; }}
            .chart-meta {{ font-size: 10px; white-space: normal; }}
            .chart-box, .chart-box.short {{ height: 340px; }}
            .chart-controls {{ position: sticky; top: 44px; z-index: 91; background: var(--bg); padding: 6px 0; margin-bottom: 6px; }}
            .chart-summary {{ width: 100%; margin-left: 0; }}
            .filter-sticky {{ top: 42px; max-height: 48vh; overflow-y: auto; -webkit-overflow-scrolling: touch; }}
            .filter-row {{ padding-bottom: 5px; }}
            .filter-chip {{ min-height: 26px; padding: 4px 8px; }}
            .preset-btn {{ min-height: 24px; padding: 4px 8px; }}
            .race-tab {{ padding: 7px 13px; font-size: 12px; }}
            .nav-btn {{ padding: 8px 4px 9px; }}
        }}


        .sort-area {{ display: flex; flex-wrap: wrap; gap: 6px; margin-bottom: 8px; }}
        .sort-chip {{ padding: 6px 10px; background: #fff; border: 1px solid var(--border); border-radius: 14px; font-size: 11px; cursor: pointer; transition: all 0.15s; }}
        .sort-chip:hover {{ border-color: var(--primary); }}
        .sort-chip.active {{ background: var(--primary); color: #fff; border-color: var(--primary); }}

        .grid-dense {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(65px, 1fr)); gap: 4px; }}
        .gd-cell {{ background: #F9FAFB; padding: 4px; border-radius: 4px; border: 1px solid #F3F4F6; }}
        .gd-lbl {{ font-size: 8px; color: var(--text-light); }}
        .gd-val {{ font-size: 10px; font-weight: 600; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }}

        .matrix-box {{ overflow-x: auto; -webkit-overflow-scrolling: touch; margin-bottom: 12px; }}
        .matrix-table {{ border-collapse: collapse; font-size: 10px; width: 100%; min-width: 300px; }}
        .matrix-table th, .matrix-table td {{ border: 1px solid #eee; padding: 4px; text-align: center; }}
        .matrix-table th {{ background: #f3f4f6; position: sticky; left: 0; z-index: 10; font-weight: bold; }}
        .mx-win {{ background: #d1fae5; color: #065f46; }}
        .mx-lose {{ background: #fee2e2; color: #991b1b; }}
        .mx-draw {{ background: #f3f4f6; color: #9ca3af; }}

        .filter-sticky {{ position: sticky; top: 42px; background: var(--bg); z-index: 90; padding: 4px 8px; margin: 0 -10px; border-bottom: 1px solid var(--border); }}
        .filter-section {{ margin-bottom: 3px; }}
        .filter-section-title {{ font-size: 8px; font-weight: bold; color: var(--text-light); margin-bottom: 2px; display: flex; align-items: center; gap: 3px; }}
        .filter-row {{
            display: flex;
            flex-wrap: nowrap;
            align-items: center;
            gap: 3px;
            margin-bottom: 2px;
            overflow-x: auto;
            padding-bottom: 1px;
            -webkit-overflow-scrolling: touch;
            scrollbar-width: none;
        }}

        .filter-row::-webkit-scrollbar {{
            display: none;
        }}.filter-label {{ min-width: 34px; font-size: 8px; font-weight: bold; color: var(--text-light); margin-right: 1px; white-space: nowrap; flex-shrink: 0; }}
        .filter-chip {{ display: inline-flex; align-items: center; min-height: 18px; padding: 1px 5px; border: 1px solid #ddd; border-radius: 7px; font-size: 9px; line-height: 1.1; cursor: pointer; background: #fff; transition: all 0.15s; user-select: none; box-sizing: border-box; }}
        .filter-chip:hover {{ border-color: var(--primary); }}
        .filter-chip.on {{ background: var(--primary); color: #fff; border-color: var(--primary); }}
        .filter-chip.on .dot {{ border: 1px solid #fff; }}
        .filter-chip .dot {{ width: 5px; height: 5px; border-radius: 50%; margin-right: 2px; }}
        .preset-btn {{ display: inline-flex; align-items: center; min-height: 18px; padding: 1px 5px; border: 1px solid #ddd; border-radius: 7px; font-size: 9px; line-height: 1.1; cursor: pointer; background: #fff; transition: all 0.15s; user-select: none; box-sizing: border-box; }}
        .preset-btn:hover {{ border-color: var(--primary); background: #f0f0ff; }}
        .preset-btn.on {{ background: var(--primary); color: #fff; border-color: var(--primary); }}
        .date-input {{ border: 1px solid #ccc; padding: 1px 3px; border-radius: 5px; font-size: 9px; width: 58px; height: 18px; line-height: 1.1; text-align: center; box-sizing: border-box; }}
        @media (max-width: 760px) {{
            .filter-sticky {{
                position: static;
                top: auto;
                max-height: none;
                overflow: visible;
                padding: 3px 6px;
                margin: 0 0 5px;
                border-radius: 6px;
                box-shadow: none;
            }}
            .chart-controls {{ position: static; top: auto; padding: 2px 0; margin-bottom: 4px; }}
            .filter-row {{
                flex-wrap: wrap;
                align-items: center;
                gap: 3px;
                overflow-x: visible;
                padding: 2px 0;
                margin-bottom: 0;
                border-bottom: 1px solid #EDF2F7;
                scrollbar-width: auto;
            }}
            .filter-row.flex-ac {{ align-items: center; }}
            .filter-row::-webkit-scrollbar {{ display: none; }}
            .filter-label {{
                flex: 0 0 32px;
                min-width: 32px;
                margin: 0 1px 0 0;
                font-size: 8px;
                line-height: 18px;
                color: #4B5563;
            }}
            .filter-chip {{
                min-height: 18px;
                padding: 1px 5px;
                font-size: 9px;
                line-height: 1.1;
                border-radius: 6px;
            }}
            .preset-btn {{
                min-height: 18px;
                padding: 1px 5px;
                font-size: 9px;
                line-height: 1.1;
                border-radius: 6px;
            }}
            .date-input {{ width: 60px; height: 18px; font-size: 9px; padding: 0 3px; }}
            .chart-box, .chart-box.short {{ touch-action: pan-y pinch-zoom; }}
        }}

        .nav {{ position: fixed; bottom: 0; left: 0; width: 100%; background: rgba(255,255,255,0.95); backdrop-filter: blur(10px); border-top: 1px solid var(--border); display: flex; padding-bottom: var(--safe-b); z-index: 200; }}
        .nav-btn {{ flex: 1; padding: 8px; text-align: center; font-size: 10px; color: var(--text-light); cursor: pointer; transition: color 0.15s; }}
        .nav-btn.active {{ color: var(--primary); font-weight: bold; }}
        .nav-icon {{ font-size: 18px; display: block; margin-bottom: 2px; }}

        .h-num {{ width: 20px; height: 20px; border-radius: 50%; display: flex; align-items: center; justify-content: center; font-size: 12px; font-weight: bold; margin-right: 6px; }}



        .badge-course {{ padding: 4px 10px; border-radius: 12px; font-size: 12px; font-weight: bold; color: #fff; }}
        .badge-venue {{ background: #6366F1; padding: 4px 10px; border-radius: 12px; font-size: 12px; font-weight: bold; color: #fff; }}

        .prediction-card {{ background: linear-gradient(135deg, #fefce8 0%, #fef3c7 100%); border: 1px solid #fde68a; border-radius: 12px; padding: 12px; margin-bottom: 10px; }}
        .verdict-tag {{ display: inline-block; padding: 2px 8px; border-radius: 6px; font-size: 10px; font-weight: bold; }}
        .verdict-有利 {{ background: #d1fae5; color: #065f46; }}
        .verdict-不利 {{ background: #fee2e2; color: #991b1b; }}
        .verdict-中立 {{ background: #f3f4f6; color: #6b7280; }}

        .pace-badge {{ display: inline-block; padding: 4px 12px; border-radius: 8px; font-size: 14px; font-weight: bold; }}
        .pace-high {{ background: #fee2e2; color: #991b1b; }}
        .pace-mid {{ background: #fef3c7; color: #92400e; }}
        .pace-slow {{ background: #dbeafe; color: #1e40af; }}
    </style>
</head>
<body>
<div id="app"></div>
<script>
    const DATA = {data_json};
    const RACES = {races_json};
    const WAKU = {waku_color_json};
    const TXT = {waku_text_json};
    const COLS = {cols_json};

    const CENTRAL_VENUES = ['東京','中山','阪神','京都','中京','小倉','新潟','福島','札幌','函館'];

    let state = {{
        race: RACES[0], tab: 'chart', expanded: {{}}, filters: {{}}, sortCol: null, sortAsc: false, wakuMap: {{}},
        dateFrom: '',
        dateTo: '',
        venueInc: [],
        distInc: [],
        courseInc: [],
        showFilters: window.innerWidth > 760,
        labelMode: 'all'
    }};

    function init() {{
        buildWakuMap();
        render();
    }}

    function setState(s) {{
        state = {{...state, ...s}};
        if(s.race) buildWakuMap();
        render();
    }}

    function buildWakuMap() {{
        const d = DATA.chart_data[state.race];
        const m = {{}};
        if(d && d.horses) {{ d.horses.forEach(h => m[h.umaban] = h.waku); }}
        state.wakuMap = m;
    }}

    function getWaku(u) {{
        return state.wakuMap[u] || 1;
    }}

    function render() {{
        const app = document.getElementById('app');
        document.body.classList.toggle('chart-tab', state.tab === 'chart');
        app.innerHTML = `
            ${{renderHeader()}}
            <div class="main">
                ${{renderInfo()}}
                ${{renderMain()}}
            </div>
            ${{renderNav()}}
        `;
        if(state.tab === 'chart') setTimeout(drawCharts, 50);
        if(state.tab === 'pacing') setTimeout(window.drawPacingCharts, 50);
    }}

    function renderHeader() {{
        return `<div class="header"><div class="race-tabs">${{RACES.map(r =>
            `<div class="race-tab ${{state.race==r?'active':''}}" onclick="setState({{race:'${{r}}'}})">${{r}}R</div>`
        ).join('')}}</div></div>`;
    }}

    function renderInfo() {{
        const i = DATA.race_info[state.race] || {{}};
        const courseBg = (i.course||'').includes('ダ') ? '#FF6D00' : '#16A34A';
        const courseLabel = (i.course||'') + (i.dist||'') + 'm';
        return `<div class="card" style="padding:12px;">
            <div class="font-bold" style="font-size:14px;margin-bottom:8px;">${{state.race}}R ${{i.race_name||''}}</div>
            <div style="display:flex;flex-wrap:wrap;gap:6px;align-items:center;">
                <div class="badge-course" style="background:${{courseBg}};">${{courseLabel}}</div>
                <div class="badge-venue">${{i.place||''}}</div>
                <span class="badge" style="font-size:11px;">${{i.cond||''}}</span>
                <span class="badge" style="font-size:11px;">${{i.heads||''}}頭</span>
            </div>
        </div>`;
    }}

    function renderMain() {{
        if(state.tab === 'chart') return renderChart();
        if(state.tab === 'data') return renderData();
        if(state.tab === 'matchup') return renderMatchup();
        if(state.tab === 'pacing') return renderPacing();
    }}

    // --- PACING (v3: 網羅的分析ダッシュボード) ---
    function renderPacing() {{
        const isMobile = window.innerWidth <= 600;
        const h = isMobile ? '280px' : '360px';
        const noData = '<div style="display:flex;align-items:center;justify-content:center;height:200px;color:#999;font-size:12px;">データなし</div>';

        // ⑦ 総合テーブル
        const st = DATA.summary_table?.[state.race] || {{}};
        let tableHTML = '';
        const stKeys = Object.keys(st).sort((a,b) => parseInt(st[a].umaban||0) - parseInt(st[b].umaban||0));
        if(stKeys.length > 0) {{
            const cs = DATA.course_stats?.[state.race] || {{}};
            const curCond = cs.cond || '';
            tableHTML = `<div style="overflow-x:auto"><table style="border-collapse:collapse;font-size:10px;width:100%;min-width:500px">
            <thead><tr style="background:#f3f4f6">
                <th style="padding:4px;border:1px solid #e5e7eb">番</th>
                <th style="padding:4px;border:1px solid #e5e7eb">馬名</th>
                <th style="padding:4px;border:1px solid #e5e7eb" title="走破タイムの相対評価 (負=速い)">走力</th>
                <th style="padding:4px;border:1px solid #e5e7eb" title="上がり3Fの相対評価 (負=速い)">末脚</th>
                <th style="padding:4px;border:1px solid #e5e7eb" title="最初に取得できる通過順の相対位置 (0=先頭, 1=最後尾)">初期位置</th>
                <th style="padding:4px;border:1px solid #e5e7eb" title="同コースでの走力評価と出走回数">同コース</th>
                <th style="padding:4px;border:1px solid #e5e7eb" title="ペース指数 (低=前傾, 高=後傾)">ペース</th>
                <th style="padding:4px;border:1px solid #e5e7eb">勝率</th>
            </tr></thead><tbody>`;
            stKeys.forEach(hid => {{
                const s = st[hid];
                const zColor = (v) => {{
                    if(v===null||v===undefined) return '#f9fafb';
                    if(v<=-0.5) return '#d1fae5'; if(v<=-0.2) return '#ecfdf5';
                    if(v>=0.5) return '#fee2e2'; if(v>=0.2) return '#fef2f2';
                    return '#f9fafb';
                }};
                const posColor = (v) => {{
                    if(v===null||v===undefined) return '#f9fafb';
                    return `hsl(${{v*120}},60%,90%)`;
                }};
                const wr = s.career?.runs > 0 ? ((s.career.wins/s.career.runs)*100).toFixed(0)+'%' : '-';
                tableHTML += `<tr>
                    <td style="padding:3px;border:1px solid #e5e7eb;text-align:center;font-weight:bold">${{s.umaban}}</td>
                    <td style="padding:3px;border:1px solid #e5e7eb;white-space:nowrap">${{s.name||''}}</td>
                    <td style="padding:3px;border:1px solid #e5e7eb;text-align:center;background:${{zColor(s.tz_med)}}">${{s.tz_med??'-'}}</td>
                    <td style="padding:3px;border:1px solid #e5e7eb;text-align:center;background:${{zColor(s.lz_med)}}">${{s.lz_med??'-'}}</td>
                    <td style="padding:3px;border:1px solid #e5e7eb;text-align:center;background:${{posColor(s.pos3_med)}}">${{s.pos3_med??'-'}}</td>
                    <td style="padding:3px;border:1px solid #e5e7eb;text-align:center;background:${{zColor(s.sc_z)}}">${{s.sc_z!=null?s.sc_z+'('+s.sc_n+')':'-'}}</td>
                    <td style="padding:3px;border:1px solid #e5e7eb;text-align:center">${{s.pi_med??'-'}}</td>
                    <td style="padding:3px;border:1px solid #e5e7eb;text-align:center">${{wr}} (${{s.career?.runs||0}}走)</td>
                </tr>`;
            }});
            tableHTML += '</tbody></table></div>';
        }}

        // === 展開予測 ===
        const dyn = DATA.race_dynamics?.[state.race] || {{}};
        let dynHTML = '';
        if(dyn.pace_prediction && dyn.formation) {{
            const pp = dyn.pace_prediction;
            const paceClass = pp.type === 'high' ? 'pace-high' : pp.type === 'slow' ? 'pace-slow' : 'pace-mid';

            // 想定隊列
            let formationHTML = '<div style="display:flex;flex-wrap:wrap;gap:4px;margin:8px 0;">';
            (dyn.formation||[]).forEach((f,i) => {{
                const posColors = {{'逃げ':'#EF4444','先行':'#F59E0B','中団':'#6B7280','差し':'#3B82F6','追込':'#8B5CF6'}};
                formationHTML += `<div style="display:flex;align-items:center;gap:3px;padding:3px 8px;background:#fff;border-radius:8px;border:1px solid #e5e7eb;font-size:10px;">
                    <span style="color:${{posColors[f.position_label]||'#666'}};font-weight:bold">${{f.position_label}}</span>
                    <span style="font-weight:600">${{f.umaban}}</span>
                    <span>${{f.name}}</span>
                </div>`;
            }});
            formationHTML += '</div>';

            // 有利/不利テーブル
            let advHTML = '<div style="overflow-x:auto"><table style="border-collapse:collapse;font-size:10px;width:100%;">';
            advHTML += '<thead><tr style="background:#f3f4f6"><th style="padding:4px;border:1px solid #e5e7eb">番</th><th style="padding:4px;border:1px solid #e5e7eb">馬名</th><th style="padding:4px;border:1px solid #e5e7eb">脚質</th><th style="padding:4px;border:1px solid #e5e7eb">展開</th><th style="padding:4px;border:1px solid #e5e7eb">理由</th><th style="padding:4px;border:1px solid #e5e7eb">想定T</th><th style="padding:4px;border:1px solid #e5e7eb">想定3F</th></tr></thead><tbody>';
            (dyn.advantages||[]).forEach(a => {{
                const tStr = a.est_time ? (Math.floor(a.est_time/60) + ':' + (a.est_time%60).toFixed(1).padStart(4,'0')) : '-';
                const lStr = a.est_l3f ? a.est_l3f.toFixed(1) : '-';
                advHTML += `<tr>
                    <td style="padding:3px;border:1px solid #e5e7eb;text-align:center;font-weight:bold">${{a.umaban}}</td>
                    <td style="padding:3px;border:1px solid #e5e7eb;white-space:nowrap">${{a.name}}</td>
                    <td style="padding:3px;border:1px solid #e5e7eb;text-align:center">${{a.position_label}}</td>
                    <td style="padding:3px;border:1px solid #e5e7eb;text-align:center"><span class="verdict-tag verdict-${{a.verdict}}">${{a.verdict}}</span></td>
                    <td style="padding:3px;border:1px solid #e5e7eb;font-size:9px;color:#666">${{a.reason}}</td>
                    <td style="padding:3px;border:1px solid #e5e7eb;text-align:center">${{tStr}}</td>
                    <td style="padding:3px;border:1px solid #e5e7eb;text-align:center">${{lStr}}</td>
                </tr>`;
            }});
            advHTML += '</tbody></table></div>';

            dynHTML = `
            <div class="prediction-card">
                <div class="font-bold" style="font-size:13px;margin-bottom:8px;">🔮 展開予測</div>
                <div style="margin-bottom:8px;">
                    <span class="pace-badge ${{paceClass}}">${{pp.label}}</span>
                    <span style="font-size:11px;color:#666;margin-left:8px;">${{pp.reason}}</span>
                </div>
                <div class="font-bold text-xs" style="margin-bottom:4px;">想定隊列（前→後）</div>
                ${{formationHTML}}
                <div class="font-bold text-xs" style="margin-top:8px;margin-bottom:4px;">各馬の展開適性</div>
                ${{advHTML}}
            </div>`;
        }}

        return `
            ${{dynHTML}}
            <div class="card">
                <div class="font-bold text-xs" style="margin-bottom:4px;">📊 出走馬 総合比較</div>
                <div style="font-size:9px;color:var(--text-light);margin-bottom:6px">走力/末脚: 負=速い（緑）, 正=遅い（赤） | 初期位置: 0=先頭, 1=最後尾 | 同コース=同コース評価(出走数)</div>
                ${{tableHTML || noData}}
            </div>
            <div class="card">
                <div class="font-bold text-xs" style="margin-bottom:4px;">🏇 コースのペースパターン</div>
                <div style="font-size:9px;color:var(--text-light);margin-bottom:4px">同コース・同距離の過去レースからペースの傾向を分析。灰色帯=平均±1σ</div>
                <div id="pac1" style="height:${{h}}"></div>
            </div>
            <div class="card">
                <div class="font-bold text-xs" style="margin-bottom:4px;">📍 初期位置 × 着順の関係</div>
                <div style="font-size:9px;color:var(--text-light);margin-bottom:4px">最終コーナーでの位置（0=先頭, 1=最後尾）と着順の関係。赤線=トレンド</div>
                <div id="pac2" style="height:${{h}}"></div>
            </div>
            <div class="card">
                <div class="font-bold text-xs" style="margin-bottom:4px;">🏃 各馬の通過パターン</div>
                <div style="font-size:9px;color:var(--text-light);margin-bottom:4px">各馬の過去走における1角→4角→ゴールの位置推移。太線=直近, 実線=同コース</div>
                <div id="pac3" style="height:${{h}}"></div>
            </div>
            <div class="card">
                <div class="font-bold text-xs" style="margin-bottom:4px;">⏱️ 走破タイム評価</div>
                <div style="font-size:9px;color:var(--text-light);margin-bottom:4px">コース・距離・馬場別に標準化したタイム。0=平均, 負=速い(良い)</div>
                <div id="pac4" style="height:${{h}}"></div>
            </div>
            <div class="card">
                <div class="font-bold text-xs" style="margin-bottom:4px;">🏁 末脚（上がり3F）評価</div>
                <div style="font-size:9px;color:var(--text-light);margin-bottom:4px">コース・距離・馬場別に標準化した上がり3F。0=平均, 負=速い(良い)</div>
                <div id="pac4b" style="height:${{h}}"></div>
            </div>
            <div class="card">
                <div class="font-bold text-xs" style="margin-bottom:4px;">🌧️ 馬場状態別パフォーマンス</div>
                <div style="font-size:9px;color:var(--text-light);margin-bottom:4px">馬場条件ごとのタイム評価。青=速い, 赤=遅い</div>
                <div id="pac5" style="height:${{h}}"></div>
            </div>
            <div class="card">
                <div class="font-bold text-xs" style="margin-bottom:4px;">📈 ペース適性</div>
                <div style="font-size:9px;color:var(--text-light);margin-bottom:4px">ペース指数（低=前傾ハイペース, 高=後傾スロー）vs 着順の関係</div>
                <div id="pac6" style="height:${{h}}"></div>
            </div>`;
    }}

    window.drawPacingCharts = () => {{
        const pData = DATA.pacing?.[state.race];
        const cs = DATA.course_stats?.[state.race];
        const hp = DATA.horse_profiles?.[state.race];
        const isMobile = window.innerWidth <= 600;
        const baseLayout = {{
            margin: isMobile ? {{t:10,b:35,l:40,r:15}} : {{t:10,b:45,l:50,r:20}},
            xaxis:{{zeroline:false, tickfont:{{size:9,color:'#6B7280'}}, gridcolor:'#F3F4F6'}},
            yaxis:{{zeroline:false, tickfont:{{size:9,color:'#6B7280'}}, gridcolor:'#F3F4F6'}},
            showlegend:true, legend:{{orientation:'h',y:-0.25,font:{{size:8}}}},
            autosize:true, paper_bgcolor:'transparent', plot_bgcolor:'transparent'
        }};
        const cfg = {{displayModeBar:false, responsive:true}};

        // ① ラップクラスター + 平均±σ帯
        const pac1 = document.getElementById('pac1');
        if(pac1) {{
            const traces = [];
            if(cs?.lap) {{
                const labels = cs.lap.mean.map((_,i) => (i+1)*200+'m');
                const upper = cs.lap.mean.map((m,i) => m - (cs.lap.std?.[i]||0));
                const lower = cs.lap.mean.map((m,i) => m + (cs.lap.std?.[i]||0));
                traces.push({{x:labels,y:upper,type:'scatter',mode:'lines',line:{{width:0}},showlegend:false,hoverinfo:'skip'}});
                traces.push({{x:labels,y:lower,type:'scatter',mode:'lines',line:{{width:0}},fill:'tonexty',fillcolor:'rgba(0,0,0,0.06)',name:`平均±1σ(N=${{cs.lap.count}})`,hoverinfo:'skip'}});
                traces.push({{x:labels,y:cs.lap.mean,type:'scatter',mode:'lines+markers',name:'平均',line:{{color:'#888',width:2,dash:'dash'}},marker:{{size:5}}}});
                // 馬場別
                const condColors = {{'良':'#3B82F6','稍重':'#F59E0B','重':'#EF4444','不良':'#7C3AED'}};
                Object.entries(cs.lap.by_condition||{{}}).forEach(([cond,d]) => {{
                    traces.push({{x:labels,y:d.mean,type:'scatter',mode:'lines',name:`${{cond}}(N=${{d.count}})`,line:{{color:condColors[cond]||'#999',width:1.5}}}});
                }});
            }}
            if(pData?.clusters?.length > 0) {{
                pData.clusters.forEach((c,i) => {{
                    const labels = c.center_laps.map((_,idx) => (idx+1)*200+'m');
                    traces.push({{x:labels,y:c.center_laps,type:'scatter',mode:'lines+markers',name:`C${{i+1}}(${{c.count}}R)`,marker:{{size:6}},line:{{width:2}},opacity:0.7}});
                }});
            }}
            if(traces.length) Plotly.newPlot('pac1',traces,{{...baseLayout,yaxis:{{...baseLayout.yaxis,title:'秒',autorange:'reversed'}}}},cfg);
        }}

        // ② 位置×着順散布
        const pac2 = document.getElementById('pac2');
        if(pac2 && cs?.pos) {{
            const pts = cs.pos.points;
            const traces2 = [];
            if(pts.length > 0) {{
                const condColors = {{'良':'rgba(59,130,246,0.3)','稍重':'rgba(245,158,11,0.3)','重':'rgba(239,68,68,0.3)','不良':'rgba(124,58,237,0.3)'}};
                const condGroups = {{}};
                pts.forEach(p => {{ (condGroups[p.cond] = condGroups[p.cond]||[]).push(p); }});
                Object.entries(condGroups).forEach(([cond,ps]) => {{
                    traces2.push({{x:ps.map(p=>p.nlc),y:ps.map(p=>p.fp),mode:'markers',type:'scatter',name:cond,marker:{{size:4,color:condColors[cond]||'rgba(0,0,0,0.2)'}},hovertext:ps.map(p=>`位置:${{p.nlc}} 着:${{p.fp}} ${{p.cond}}`),hoverinfo:'text'}});
                }});
            }}
            if(cs.pos.trend?.length > 0) {{
                traces2.push({{x:cs.pos.trend.map(t=>t.x),y:cs.pos.trend.map(t=>t.y_mean),mode:'lines+markers',type:'scatter',name:'平均着順',line:{{color:'#EF4444',width:3}},marker:{{size:6}}}});
            }}
            // 各馬の位置をマーカー
            if(hp) {{
                const blocked = state.filters[state.race] || [];
                const hpArr = Object.values(hp).filter(h => !blocked.includes(h.umaban)).sort((a,b)=>parseInt(a.umaban)-parseInt(b.umaban));
                hpArr.forEach(h => {{
                    const df = state.dateFrom ? state.dateFrom.replaceAll('/','') : '';
                    const dt = state.dateTo ? state.dateTo.replaceAll('/','') : '';
                    const trajs = (df || dt) ? h.trajs.filter(t => {{
                        const dStr = String(t.date).replaceAll('/','');
                        return (!df || dStr >= df) && (!dt || dStr <= dt);
                    }}) : h.trajs;

                    const pos3s = trajs.filter(t=>t.pos[2]!=null).map(t=>t.pos[2]);
                    if(pos3s.length > 0) {{
                        const med = pos3s.sort((a,b)=>a-b)[Math.floor(pos3s.length/2)];
                        const w = getWaku(h.umaban);
                        traces2.push({{x:[med,med],y:[1,18],mode:'lines',type:'scatter',name:`${{h.umaban}}${{h.name.slice(0,3)}}`,line:{{color:WAKU[w],width:2,dash:'dot'}},showlegend:true}});
                    }}
                }});
            }}
            if(traces2.length) Plotly.newPlot('pac2',traces2,{{...baseLayout,xaxis:{{...baseLayout.xaxis,title:'正規化初期位置(0=先頭)'}},yaxis:{{...baseLayout.yaxis,title:'着順',autorange:'reversed'}}}},cfg);
        }}

        // ③ 通過軌跡 (直近2走のみ + 同コース)
        const pac3 = document.getElementById('pac3');
        if(pac3 && hp) {{
            const xLabels = ['1角','2角','3角','4角','ゴール'];
            const traces3 = [];
            const blocked = state.filters[state.race] || [];
            const df = state.dateFrom ? state.dateFrom.replaceAll('/','') : '';
            const dt = state.dateTo ? state.dateTo.replaceAll('/','') : '';

            Object.values(hp).filter(h => !blocked.includes(h.umaban)).sort((a,b)=>parseInt(a.umaban)-parseInt(b.umaban)).forEach(h => {{
                const w = getWaku(h.umaban);
                const color = WAKU[w]||'#888';
                const maxShow = 2;
                let shown = 0;

                const trajs = (df || dt) ? h.trajs.filter(t => {{
                    if(!t.date) return true;
                    const dStr = String(t.date).replaceAll('/','');
                    return (!df || dStr >= df) && (!dt || dStr <= dt);
                }}) : h.trajs;

                trajs.forEach((t,i) => {{
                    if(shown >= maxShow && !t.same) return;
                    const validPos = t.pos.filter(p=>p!=null);
                    const validLabels = t.pos.map((p,j)=>(p!=null?xLabels[j]:null)).filter(l=>l!=null);
                    if(validPos.length < 2) return;
                    const isFirst = shown === 0;
                    shown++;
                    traces3.push({{
                        x:validLabels, y:validPos, mode:'lines+markers', type:'scatter',
                        name: isFirst ? `${{h.umaban}}${{h.name.slice(0,3)}}` : '',
                        showlegend: isFirst,
                        legendgroup: `h${{h.umaban}}`,
                        line:{{color:color, width:isFirst?2.5:1.5, dash:t.same?'solid':'dot'}},
                        marker:{{size:isFirst?6:4}},
                        opacity: isFirst ? 1 : 0.5,
                        hovertext: validPos.map(p=>`${{h.umaban}} ${{h.name.slice(0,3)}}\n${{t.date}} ${{t.venue}}${{t.dist}} ${{t.cond}} ${{t.finish}}着`),
                        hoverinfo:'text'
                    }});
                }});
            }});
            if(traces3.length) Plotly.newPlot('pac3',traces3,{{...baseLayout,xaxis:{{...baseLayout.xaxis,categoryorder:'array',categoryarray:['1角','2角','3角','4角','ゴール']}},yaxis:{{...baseLayout.yaxis,title:'位置(0=先頭)',range:[1.05,-0.05]}}}},cfg);
        }}

        // ④ タイムZ-score Box Plot
        const pac4 = document.getElementById('pac4');
        if(pac4 && hp) {{
            const traces4 = [];
            const blocked = state.filters[state.race] || [];
            const hpArr = Object.values(hp).filter(h => !blocked.includes(h.umaban)).sort((a,b)=>parseInt(a.umaban)-parseInt(b.umaban));
            const df = state.dateFrom ? state.dateFrom.replaceAll('/','') : '';
            const dt = state.dateTo ? state.dateTo.replaceAll('/','') : '';
            hpArr.forEach(h => {{
                const validTimeZ = (df || dt) ? h.time_z.filter(t => {{
                    if(!t.date) return true;
                    const dStr = String(t.date).replaceAll('/','');
                    return (!df || dStr >= df) && (!dt || dStr <= dt);
                }}) : h.time_z;
                const zs = validTimeZ.map(x=>x.z);
                if(zs.length > 0) {{
                    const w = getWaku(h.umaban);
                    traces4.push({{y:zs, type:'box', name:`${{h.umaban}}${{h.name.slice(0,3)}}`, marker:{{color:WAKU[w]}}, boxpoints:'all', jitter:0.3, pointpos:-1.5, line:{{width:1}}, whiskerwidth:0.5}});
                }}
            }});
            if(traces4.length) {{
                Plotly.newPlot('pac4',traces4,{{
                    ...baseLayout,
                    yaxis:{{...baseLayout.yaxis,title:'タイムZ (負=速い)',zeroline:true,zerolinecolor:'#EF4444',zerolinewidth:1}},
                    xaxis:{{...baseLayout.xaxis,tickangle:-45,tickfont:{{size:8}}}},
                    showlegend:false
                }},cfg);
            }}
        }}
        // ④-b 3F Z-score Box Plot
        const pac4b = document.getElementById('pac4b');
        if(pac4b && hp) {{
            const traces4b = [];
            const blocked = state.filters[state.race] || [];
            const hpArr2 = Object.values(hp).filter(h => !blocked.includes(h.umaban)).sort((a,b)=>parseInt(a.umaban)-parseInt(b.umaban));
            const df = state.dateFrom ? state.dateFrom.replaceAll('/','') : '';
            const dt = state.dateTo ? state.dateTo.replaceAll('/','') : '';
            hpArr2.forEach(h => {{
                const validL3fZ = (df || dt) ? h.l3f_z.filter(t => {{
                    if(!t.date) return true;
                    const dStr = String(t.date).replaceAll('/','');
                    return (!df || dStr >= df) && (!dt || dStr <= dt);
                }}) : h.l3f_z;
                const zs = validL3fZ.map(x=>x.z);
                if(zs.length > 0) {{
                    const w = getWaku(h.umaban);
                    traces4b.push({{y:zs, type:'box', name:`${{h.umaban}}${{h.name.slice(0,3)}}`, marker:{{color:WAKU[w]}}, boxpoints:'all', jitter:0.3, pointpos:-1.5, line:{{width:1}}, whiskerwidth:0.5}});
                }}
            }});
            if(traces4b.length) {{
                Plotly.newPlot('pac4b',traces4b,{{
                    ...baseLayout,
                    yaxis:{{...baseLayout.yaxis,title:'3F Z (負=速い)',zeroline:true,zerolinecolor:'#EF4444',zerolinewidth:1}},
                    xaxis:{{...baseLayout.xaxis,tickangle:-45,tickfont:{{size:8}}}},
                    showlegend:false
                }},cfg);
            }}
        }}

        // ⑤ 馬場別ヒートマップ
        const pac5 = document.getElementById('pac5');
        if(pac5 && hp) {{
            const conditions = ['良','稍重','重','不良'];
            const blocked = state.filters[state.race] || [];
            const hpArr = Object.values(hp).filter(h => !blocked.includes(h.umaban)).sort((a,b)=>parseInt(a.umaban)-parseInt(b.umaban));
            const zValues = []; const hoverTexts = []; const yLabels = [];
            hpArr.forEach(h => {{
                const row = []; const hrow = [];
                conditions.forEach(c => {{
                    const cd = h.by_cond[c];
                    if(cd) {{ row.push(cd.z_med); hrow.push(`${{h.umaban}} ${{h.name.slice(0,3)}} ${{c}}: Z=${{cd.z_med}} (N=${{cd.n}})`); }}
                    else {{ row.push(null); hrow.push(`${{h.umaban}} ${{h.name.slice(0,3)}} ${{c}}: データなし`); }}
                }});
                zValues.push(row); hoverTexts.push(hrow); yLabels.push(`${{h.umaban}}${{h.name.slice(0,3)}}`);
            }});
            const curCond = DATA.course_stats?.[state.race]?.cond || '';
            if(zValues.length > 0) {{
                Plotly.newPlot('pac5',[{{
                    z:zValues, x:conditions, y:yLabels, type:'heatmap',
                    colorscale:[[0,'#065f46'],[0.3,'#d1fae5'],[0.5,'#f9fafb'],[0.7,'#fee2e2'],[1,'#991b1b']],
                    zmin:-1.5, zmax:1.5, hovertext:hoverTexts, hoverinfo:'text',
                    colorbar:{{title:'Z',thickness:10,len:0.8,tickfont:{{size:8}}}}
                }}],{{
                    ...baseLayout, margin:{{...baseLayout.margin,l:isMobile?80:100}},
                    yaxis:{{...baseLayout.yaxis,tickfont:{{size:8}}}},
                    annotations: curCond ? [{{x:curCond,y:yLabels.length,text:'▼本走',showarrow:false,font:{{size:9,color:'#EF4444'}}}}] : []
                }},cfg);
            }}
        }}

        // ⑥ ペース適性散布図
        const pac6 = document.getElementById('pac6');
        if(pac6 && hp) {{
            const traces6 = [];
            const blocked = state.filters[state.race] || [];
            const df = state.dateFrom ? state.dateFrom.replaceAll('/','') : '';
            const dt = state.dateTo ? state.dateTo.replaceAll('/','') : '';
            Object.values(hp).filter(h => !blocked.includes(h.umaban)).forEach(h => {{
                if(!h.pace || h.pace.length === 0) return;

                const validPace = (df || dt) ? h.pace.filter(t => {{
                    if(!t.date) return true;
                    const dStr = String(t.date).replaceAll('/','');
                    return (!df || dStr >= df) && (!dt || dStr <= dt);
                }}) : h.pace;
                if(validPace.length === 0) return;

                const w = getWaku(h.umaban);
                traces6.push({{
                    x:validPace.map(p=>p.pi), y:validPace.map(p=>p.fp),
                    mode:'markers', type:'scatter',
                    name:`${{h.umaban}} ${{h.name.slice(0,3)}}`,
                    marker:{{size:7,color:WAKU[w],line:{{width:1,color:'#888'}}}},
                    hovertext:validPace.map(p=>`${{h.umaban}} ${{h.name.slice(0,3)}} PI=${{p.pi}} ${{p.fp}}着`), hoverinfo:'text'
                }});
            }});
            if(traces6.length) Plotly.newPlot('pac6',traces6,{{
                ...baseLayout,
                xaxis:{{...baseLayout.xaxis,title:'ペース指数(低=前傾)',tickangle:0}},
                yaxis:{{...baseLayout.yaxis,title:'着順',autorange:'reversed'}}
            }},cfg);
        }}
    }};

    // --- CHART ---
    function renderChart() {{
        const d = DATA.chart_data[state.race] || {{horses:[]}};
        const blocked = state.filters[state.race] || [];
        const ri = DATA.race_info?.[state.race] || {{}};

        const rows = DATA.races[state.race]?.rows || [];
        const allVenues = [...new Set(rows.map(r => r['場所']).filter(v => v))].sort();
        const allDists = [...new Set(rows.map(r => r['距離']).filter(v => v))].sort((a,b)=>parseInt(a)-parseInt(b));
        const allCourses = [...new Set(rows.map(r => r['ｺｰｽ']).filter(v => v))].sort();

        const vc={{}},dc={{}},cc={{}};
        rows.forEach(r => {{
            const v=r['場所'],dd=r['距離'],c=r['ｺｰｽ'];
            if(v) vc[v]=(vc[v]||0)+1;
            if(dd) dc[dd]=(dc[dd]||0)+1;
            if(c) cc[c]=(cc[c]||0)+1;
        }});

        const horseChips = d.horses.map(h => {{
            const isOn = !blocked.includes(h.umaban);
            return `<div onclick="togFilter(${{h.umaban}})" class="filter-chip ${{isOn?'on':''}}"
                style="${{isOn?'background:'+WAKU[h.waku]+';color:'+TXT[h.waku]+';border-color:'+WAKU[h.waku]:'opacity:0.4'}}">
                <b>${{h.umaban}}</b></div>`;
        }}).join('');

        const venueChips = allVenues.map(v => {{
            const isOn = state.venueInc.length===0 || state.venueInc.includes(v);
            return `<div onclick="togVenue('${{v}}')" class="filter-chip ${{isOn?'on':''}}"
                style="${{isOn?'':'opacity:0.4'}}">${{v}}<span style="font-size:7px;margin-left:2px;opacity:0.6">${{vc[v]||0}}</span></div>`;
        }}).join('');

        const distChips = allDists.map(dd => {{
            const isOn = state.distInc.length===0 || state.distInc.includes(dd);
            return `<div onclick="togDist('${{dd}}')" class="filter-chip ${{isOn?'on':''}}"
                style="${{isOn?'':'opacity:0.4'}}">${{dd}}<span style="font-size:7px;margin-left:2px;opacity:0.6">${{dc[dd]||0}}</span></div>`;
        }}).join('');

        const courseChips = allCourses.map(c => {{
            const isOn = state.courseInc.length===0 || state.courseInc.includes(c);
            const col = c==='芝'?'#00C853':(c==='ダート'||c==='ダ')?'#FF6D00':'#888';
            return `<div onclick="togCourse('${{c}}')" class="filter-chip ${{isOn?'on':''}}"
                style="${{isOn?'border-color:'+col+';background:'+col+';color:#fff':'opacity:0.4'}}">
                ${{c}}<span style="font-size:7px;margin-left:2px;opacity:0.7">${{cc[c]||0}}</span></div>`;
        }}).join('');

        const hasAny = blocked.length>0 || state.venueInc.length>0 || state.distInc.length>0 || state.courseInc.length>0 || state.dateFrom || state.dateTo;

        return `
            <div class="filter-sticky">
                <div class="flex-sb flex-ac" onclick="setState({{showFilters: !state.showFilters}})" style="cursor:pointer; padding:2px 0;">
                    <div class="flex-ac gap-1">
                        <span class="font-bold text-sm">🔎 絞り込み</span>
                        ${{hasAny ? '<span style="background:#EF4444;color:#fff;font-size:8px;padding:1px 5px;border-radius:8px;margin-left:4px">ON</span>' : ''}}
                    </div>
                    <div style="font-size:14px;color:#6B7280">${{state.showFilters ? '▲' : '▼'}}</div>
                </div>
                ${{state.showFilters ? `
                <div style="margin-top:2px;">
                    <div class="filter-row">
                        <div class="filter-label">一括</div>
                        <div class="preset-btn" onclick="presetCentral()">中央</div>
                        <div class="preset-btn" onclick="presetLocal()">地方</div>
                        <div class="preset-btn" onclick="presetSameVenue()">同場所</div>
                        <div class="preset-btn" onclick="presetSameDist()">同距離</div>
                        <div class="preset-btn" onclick="presetReset()">全表示</div>
                    </div>
                    <div class="filter-row flex-ac">
                        <div class="filter-label">期間</div>
                        <input type="text" class="date-input" value="${{state.dateFrom}}" placeholder="YY/MM/DD"
                               onchange="setState({{dateFrom:this.value}})">
                        <span style="margin:0 1px;font-size:9px;color:#999">-</span>
                        <input type="text" class="date-input" value="${{state.dateTo}}" placeholder="YY/MM/DD"
                               onchange="setState({{dateTo:this.value}})">
                        <div class="preset-btn" style="margin-left:2px" onclick="setPeriod(1)">1</div>
                        <div class="preset-btn" onclick="setPeriod(3)">3</div>
                        <div class="preset-btn" onclick="setPeriod(6)">半年</div>
                        <div class="preset-btn" onclick="setPeriod(12)">年</div>
                        <div class="preset-btn" onclick="setPeriod(0)">全</div>
                    </div>
                    <div class="filter-row">
                        <div class="filter-label">場所</div>
                        ${{venueChips}}
                    </div>
                    <div class="filter-row">
                        <div class="filter-label">距離</div>
                        ${{distChips}}
                    </div>
                    <div class="filter-row">
                        <div class="filter-label">コース</div>
                        ${{courseChips}}
                    </div>
                    <div class="filter-row">
                        <div class="filter-label">馬番</div>
                        ${{horseChips}}
                    </div>
                </div>
                ` : ''}}
            </div>
            <div class="card chart-card">
                <div class="chart-controls">
                    <span class="font-bold text-xs">馬番表示</span>
                    <div class="segmented" role="group" aria-label="馬番表示">
                        <button class="seg-btn ${{state.labelMode==='auto'?'active':''}}" onclick="setLabelMode('auto')">自動</button>
                        <button class="seg-btn ${{state.labelMode==='all'?'active':''}}" onclick="setLabelMode('all')">全点</button>
                        <button class="seg-btn ${{state.labelMode==='none'?'active':''}}" onclick="setLabelMode('none')">なし</button>
                    </div>
                    <div class="chart-summary" id="chart-summary">表示準備中</div>
                </div>
                <div id="chart-section"></div>
            </div>`;
    }}


    window.setLabelMode = (mode) => setState({{labelMode: mode}});

    window.togFilter = (u) => {{
        const old = state.filters[state.race] || [];
        const next = old.includes(u) ? old.filter(x=>x!==u) : [...old, u];
        setState({{filters: {{...state.filters, [state.race]: next}} }});
    }};
    window.togVenue = (v) => {{
        let cur = [...state.venueInc];
        if(cur.length===0) cur=[v];
        else if(cur.includes(v)) cur=cur.filter(x=>x!==v);
        else cur.push(v);
        setState({{venueInc: cur}});
    }};
    window.togDist = (d) => {{
        let cur = [...state.distInc];
        if(cur.length===0) cur=[d];
        else if(cur.includes(d)) cur=cur.filter(x=>x!==d);
        else cur.push(d);
        setState({{distInc: cur}});
    }};
    window.togCourse = (c) => {{
        let cur = [...state.courseInc];
        if(cur.length===0) cur=[c];
        else if(cur.includes(c)) cur=cur.filter(x=>x!==c);
        else cur.push(c);
        setState({{courseInc: cur}});
    }};
    window.presetCentral = () => {{
        const allV = [...new Set(DATA.races[state.race]?.rows.map(r=>r['場所']).filter(x=>x))];
        setState({{venueInc: allV.filter(v=>CENTRAL_VENUES.includes(v))}});
    }};
    window.presetLocal = () => {{
        const allV = [...new Set(DATA.races[state.race]?.rows.map(r=>r['場所']).filter(x=>x))];
        setState({{venueInc: allV.filter(v=>!CENTRAL_VENUES.includes(v))}});
    }};
    window.presetSameVenue = () => {{
        const v = (DATA.race_info?.[state.race]||{{}}).place||'';
        setState({{venueInc: v?[v]:[]}});
    }};
    window.presetSameDist = () => {{
        const d = (DATA.race_info?.[state.race]||{{}}).dist||'';
        setState({{distInc: d?[d]:[]}});
    }};
    window.presetReset = () => {{
        setState({{filters:{{...state.filters,[state.race]:[]}},venueInc:[],distInc:[],courseInc:[],dateFrom:'',dateTo:''}});
    }};
    window.setPeriod = (months) => {{
        if(months===0) {{ setState({{dateFrom:'',dateTo:''}}); return; }}

        // 本日日付を基準に期間を設定
        const now = new Date();
        const toStr = now.getFullYear() + '/' + String(now.getMonth()+1).padStart(2,'0') + '/' + String(now.getDate()).padStart(2,'0');
        let fy = now.getFullYear(), fm = now.getMonth()+1 - months;
        while(fm<=0){{ fy--; fm+=12; }}
        const fromStr = fy + '/' + String(fm).padStart(2,'0') + '/' + String(now.getDate()).padStart(2,'0');
        setState({{dateFrom:fromStr, dateTo:toStr}});
    }};

    function drawCharts() {{
        const d = DATA.chart_data[state.race] || {{horses:[]}};
        const blocked = state.filters[state.race] || [];
        const active = d.horses.filter(h => !blocked.includes(h.umaban));
        const rows = DATA.races[state.race]?.rows || [];
        const rowMap = {{}};
        rows.forEach(r => {{ rowMap[r['日付']+'-'+r['番']] = r; }});

        const normalizeDate = (v) => String(v || '').replaceAll('/','');
        const passesFilters = (h, r) => {{
            const df = state.dateFrom ? normalizeDate(state.dateFrom) : '';
            const dt = state.dateTo ? normalizeDate(state.dateTo) : '';
            const rd = normalizeDate(r.date);
            if(df && rd < df) return false;
            if(dt && rd > dt) return false;
            const row = rowMap[r.date+'-'+h.umaban];
            if(state.venueInc.length>0 && row && !state.venueInc.includes(row['場所'])) return false;
            if(state.distInc.length>0 && row && !state.distInc.includes(row['距離'])) return false;
            if(state.courseInc.length>0 && !state.courseInc.includes(r.course)) return false;
            return true;
        }};

        const pts = [];
        active.forEach(h => h.records.forEach((r, idx) => {{
            if(!passesFilters(h, r)) return;
            const row = rowMap[r.date+'-'+h.umaban];
            pts.push({{...r, recordIndex: idx, u:h.umaban, name:h.name, w:h.waku, course:r.course||'', row: row||{{}}}});
        }}));

        const latestByHorse = {{}};
        pts.forEach(p => {{
            const prev = latestByHorse[p.u];
            if(!prev || normalizeDate(p.date) >= normalizeDate(prev.date)) latestByHorse[p.u] = p;
        }});

        const isMobile = window.innerWidth <= 760;
        const dense = pts.length > 90 || active.length > 14;
        const labelMode = state.labelMode || 'all';
        const showLabels = labelMode !== 'none';
        const isLabelPoint = (p) => {{
            if(labelMode === 'all') return true;
            if(labelMode === 'none') return false;
            return pts.length <= 70 || latestByHorse[p.u] === p;
        }};
        const labelFor = (p) => showLabels && isLabelPoint(p) ? String(p.u) : '';
        const chartMeta = `${{active.length}}頭 / ${{pts.length}}点${{dense ? ' / 高密度' : ''}}`;
        const summary = document.getElementById('chart-summary');
        if(summary) {{
            const activeFilters = [
                blocked.length ? `非表示${{blocked.length}}頭` : '',
                state.venueInc.length ? `場所${{state.venueInc.length}}` : '',
                state.distInc.length ? `距離${{state.distInc.length}}` : '',
                state.courseInc.length ? `コース${{state.courseInc.length}}` : '',
                (state.dateFrom || state.dateTo) ? '期間' : ''
            ].filter(Boolean).join(' / ');
            summary.textContent = `${{chartMeta}}${{activeFilters ? ' / '+activeFilters : ''}}`;
        }}

        const chartSection = document.getElementById('chart-section');
        if(!chartSection) return;
        if(!active.length || !pts.length) {{
            chartSection.innerHTML = `<div class="empty-chart">
                <div>該当データがありません。フィルタ条件を緩めるか、馬番表示を戻してください。</div>
                <button class="link-btn" onclick="presetReset()">フィルタを全表示に戻す</button>
            </div>`;
            return;
        }}

        const baseLayout = {{
            margin: isMobile ? {{t:12,b:42,l:42,r:12}} : {{t:12,b:48,l:50,r:18}},
            paper_bgcolor:'#fff',
            plot_bgcolor:'#fff',
            font: {{family:'sans-serif', size:isMobile ? 10 : 11}},
            xaxis:{{zeroline:false, gridcolor:'#EEF2F7', automargin:true}},
            yaxis:{{zeroline:false, gridcolor:'#EEF2F7', automargin:true}},
            showlegend:false,
            hovermode:'closest',
            dragmode:'pan'
        }};
        const cfg = {{
            displayModeBar:true,
            responsive:true,
            scrollZoom:true,
            staticPlot:false,
            displaylogo:false,
            modeBarButtonsToRemove:['lasso2d','select2d','autoScale2d','toggleSpikelines']
        }};
        const courseColor = (c) => c === '芝' ? '#00A651' : (c === 'ダート' || c === 'ダ') ? '#E86C00' : '#64748B';
        const pointOpacity = dense ? 0.68 : 0.86;
        const hasPlotValue = (v) => v !== null && v !== undefined && v !== '' && !Number.isNaN(Number(v));
        const baseMarkerSize = dense ? (isMobile ? 13 : 12) : (isMobile ? 16 : 15);
        const distanceSize = (p) => {{
            const dist = parseInt(p.row?.['距離']) || 0;
            if(!dist) return baseMarkerSize;
            if(dist >= 2600) return baseMarkerSize + 7;
            if(dist >= 2200) return baseMarkerSize + 5;
            if(dist >= 1800) return baseMarkerSize + 3;
            if(dist <= 1200) return Math.max(baseMarkerSize - 1, 10);
            return baseMarkerSize + 1;
        }};

        const getHoverText = (p) => {{
            const r = p.row || {{}};
            const pop = r['人気'] ? `${{r['人気']}}人気(${{r['ｵｯｽﾞ']||'-'}}倍)` : '';
            const rank = r['着順'] ? `${{r['着順']}}着(差${{r['着差']||'-'}})` : '';
            const raceInfo = r['場所'] ? `${{p.course}}${{r['場所']}}${{r['距離']}}m` : p.course;
            const horseName = p.name || r['馬名'] || '';
            const conf = p.idx_conf ? `<br>信頼度:${{p.idx_conf}} N:${{p.idx_n || '-'}}` : '';
            const initPos = p.initial_corner ? `<br>初期位置:${{p.initial_corner}} ${{p.initial_pos ?? '-'}}` : '';
            return `${{p.date}} ${{p.u}}番 ${{horseName}}<br>${{raceInfo}} ${{pop}}<br>${{rank}}<br>T:${{p.time_idx ?? '-'}} 上り:${{p.l3f_idx ?? '-'}} RPCI:${{p.rpci ?? '-'}}${{initPos}}${{conf}}`;
        }};

        const makeMarker = (list) => ({{
            size:list.map(p=>distanceSize(p)),
            sizemode:'diameter',
            color:list.map(p=>WAKU[p.w]),
            opacity: pointOpacity,
            line:{{width: dense ? 2 : 2.4, color:list.map(p=>courseColor(p.course))}}
        }});
        const makeTextFont = (list) => ({{
            size: isMobile ? 9 : 10,
            color:list.map(p=>TXT[p.w])
        }});

        const createPanel = (id, title, meta='', wide=false) => `
            <div class="chart-panel ${{wide?'wide':''}}">
                <div class="chart-panel-head">
                    <div class="chart-title">${{title}}</div>
                    <div class="chart-meta">${{meta || chartMeta}}</div>
                </div>
                <div id="${{id}}" class="chart-box ${{wide?'':'short'}}"></div>
            </div>`;

        let hasL3f = false, hasCorners = false, hasRpci = false;
        pts.forEach(p => {{ if(p.has_l3f) hasL3f=true; if(p.has_corners) hasCorners=true; if(p.has_rpci) hasRpci=true; }});

        let chartHTML = '<div class="chart-grid">';
        chartHTML += createPanel('c_tidx', '① 走力(T指数) 馬別比較', '箱ひげ + 全履歴点', true);
        if(hasL3f) chartHTML += createPanel('c1', '② 走力(T指数) vs 末脚(上がり指数)');
        if(hasCorners) chartHTML += createPanel('c2', `${{hasL3f?'③':'②'}} 走力(T指数) vs 初期位置`);
        if(hasL3f && hasCorners) chartHTML += createPanel('c3', '④ 末脚(上がり指数) vs 初期位置');
        if(hasL3f && hasRpci) chartHTML += createPanel('c4', `${{hasCorners?'⑤':'③'}} 末脚(上がり指数) vs ペース(RPCI)`);
        chartHTML += createPanel('c_rank', `${{hasL3f&&hasCorners?'⑥':'②'}} 走力(T指数) vs 着順`);
        chartHTML += '</div>';
        if(!hasL3f && !hasCorners) {{
            chartHTML += `<div style="margin-top:8px;padding:8px;background:#fef3c7;border-radius:8px;font-size:11px;color:#92400e">
                ⚠ 地方競馬データのため、上がり3F・通過順・RPCIは未提供です。T指数と着順のチャートを表示しています。</div>`;
        }}
        chartSection.innerHTML = chartHTML;

        const createScatterChart = (id, xKey, yKey, titleX, titleY, extraLayout={{}}) => {{
            const list = pts.filter(p => hasPlotValue(p[xKey]) && hasPlotValue(p[yKey]));
            const gd = document.getElementById(id);
            if(!gd) return;
            if(!list.length) {{
                gd.innerHTML = '<div class="empty-chart">この条件で描画できる点がありません。</div>';
                return;
            }}
            const tr = [{{
                x: list.map(p=>Number(p[xKey])), y: list.map(p=>Number(p[yKey])), mode:'markers+text', type:'scatter',
                marker: makeMarker(list),
                text: list.map(p=>labelFor(p)), textposition:'middle center', textfont:makeTextFont(list),
                hovertext: list.map(p=>getHoverText(p)), hoverinfo:'text', showlegend:false
            }}];
            Plotly.newPlot(id, tr, {{...baseLayout, ...extraLayout, xaxis:{{...baseLayout.xaxis,title:titleX,...(extraLayout.xaxis||{{}})}}, yaxis:{{...baseLayout.yaxis,title:titleY,...(extraLayout.yaxis||{{}})}}}}, cfg);
        }};

        const activeHorses = d.horses.filter(h => !blocked.includes(h.umaban)).sort((a,b) => a.umaban - b.umaban);
        const boxTraces = [];
        activeHorses.forEach(h => {{
            const recs = h.records.filter(r => passesFilters(h, r));
            const vals = recs.filter(r => r.time_idx > 0);
            if(vals.length > 0) {{
                const hoverTexts = vals.map(r => {{
                    const row = rowMap[r.date+'-'+h.umaban] || {{}};
                    const pop = row['人気'] ? `${{row['人気']}}人気(${{row['ｵｯｽﾞ']||'-'}}倍)` : '';
                    const rank = row['着順'] ? `${{row['着順']}}着(差${{row['着差']||'-'}})` : '';
                    const raceInfo = row['場所'] ? `${{r.course}}${{row['場所']}}${{row['距離']}}m` : r.course;
                    return `${{r.date}} ${{h.umaban}}番 ${{h.name}}<br>${{raceInfo}} ${{pop}}<br>${{rank}} T:${{r.time_idx}}`;
                }});
                boxTraces.push({{
                    y: vals.map(v=>v.time_idx), type:'box',
                    name: `${{h.umaban}} ${{h.name.slice(0,5)}}`,
                    marker: {{color:WAKU[h.waku], opacity:0.7, size:dense?4:5}},
                    boxpoints: 'all', jitter:dense?0.22:0.3, pointpos:-1.35,
                    line: {{width:1.2, color:WAKU[h.waku]}}, whiskerwidth:0.5,
                    text: hoverTexts, hoverinfo:'text'
                }});
            }}
        }});
        if(boxTraces.length) {{
            Plotly.newPlot('c_tidx', boxTraces, {{
                ...baseLayout,
                margin: isMobile ? {{t:12,b:72,l:42,r:12}} : {{t:12,b:82,l:50,r:18}},
                yaxis:{{...baseLayout.yaxis,title:'T指数',zeroline:true,zerolinecolor:'#ddd'}},
                xaxis:{{...baseLayout.xaxis,tickangle:isMobile?-55:-35,tickfont:{{size:isMobile?8:9}}}}
            }}, cfg);
        }}

        const rankPts = pts.filter(p => p.row['着順'] && String(p.row['着順']).match(/^\\d+$/));
        rankPts.forEach(p => {{ p.rankValue = parseInt(p.row['着順']); }});
        if(rankPts.length > 0) {{
            const tr = [{{
                x: rankPts.map(p => p.rankValue),
                y: rankPts.map(p => p.time_idx),
                mode:'markers+text', type:'scatter',
                marker: makeMarker(rankPts),
                text: rankPts.map(p=>labelFor(p)), textposition:'middle center',
                textfont:makeTextFont(rankPts),
                hovertext: rankPts.map(p=>getHoverText(p)), hoverinfo:'text', showlegend:false
            }}];
            Plotly.newPlot('c_rank', tr, {{...baseLayout, xaxis:{{...baseLayout.xaxis,title:'着順',dtick:1}}, yaxis:{{...baseLayout.yaxis,title:'T指数'}}}}, cfg);
        }}

        if(hasL3f) createScatterChart('c1', 'l3f_idx', 'time_idx', '上がり指数', 'T指数');
        if(hasCorners) createScatterChart('c2', 'initial_pos', 'time_idx', '初期位置(最初に取れる通過順)', 'T指数');
        if(hasL3f && hasCorners) createScatterChart('c3', 'initial_pos', 'l3f_idx', '初期位置(最初に取れる通過順)', '上がり指数');
        if(hasL3f && hasRpci) createScatterChart('c4', 'rpci', 'l3f_idx', 'RPCI', '上がり指数');
    }}


    function renderData() {{
        const rows = DATA.races[state.race]?.rows || [];
        const headers = COLS.map(c =>
            `<div class="sort-chip ${{state.sortCol===c?'active':''}}" onclick="setState({{sortCol:'${{c}}', sortAsc:${{state.sortCol===c?!state.sortAsc:true}} }})">
                ${{c}} ${{state.sortCol===c ? (state.sortAsc?'▲':'▼') : ''}}
             </div>`
        ).join('');

        let content = '';
        if(state.sortCol) {{
            const sorted = [...rows].sort((a,b) => {{
                let va = a[state.sortCol]||'', vb = b[state.sortCol]||'';
                let na = parseFloat(va), nb = parseFloat(vb);
                if(!isNaN(na) && !isNaN(nb)) return state.sortAsc ? na-nb : nb-na;
                return state.sortAsc ? va.localeCompare(vb) : vb.localeCompare(va);
            }});
            content = sorted.map(row => `
                <div class="card" style="padding:8px;">
                    <div class="flex-ac gap-2" style="border-bottom:1px solid #eee; margin-bottom:4px; padding-bottom:4px;">
                        <div class="h-num" style="background:${{WAKU[getWaku(row['番'])] || '#ccc'}}; color:${{TXT[getWaku(row['番'])] || '#fff'}}">${{row['番']}}</div>
                        <div class="text-xs font-bold" style="flex:1">${{row['馬名']}}</div>
                        <div class="text-xs text-light">${{row['日付']}}</div>
                        <div class="font-bold text-xs" style="color:${{state.sortCol}}">${{row[state.sortCol]}}</div>
                    </div>
                    <div class="grid-dense">
                        ${{COLS.map(c => `<div class="gd-cell"><span class="gd-lbl">${{c}}</span><div class="gd-val">${{row[c]||'-'}}</div></div>`).join('')}}
                    </div>
                </div>
            `).join('');
            content = `<div class="card" onclick="setState({{sortCol:null}})" style="text-align:center;padding:8px;color:var(--primary);font-weight:bold;cursor:pointer">↺ グループ表示に戻す</div>` + content;

        }} else {{
            const groups = {{}};
            rows.forEach(r => {{
                const k = r['番']||'0';
                if(!groups[k]) groups[k] = [];
                groups[k].push(r);
            }});
            content = Object.keys(groups).sort((a,b)=>parseInt(a)-parseInt(b)).map(k => {{
                const list = groups[k];
                const meta = list[0];
                const w = getWaku(k);
                const open = state.expanded[state.race+'-'+k];
                const arrow = open ? '▲' : '▼';
                const body = open ? `
                    <div style="margin-top:8px">
                        ${{list.map(row => `
                            <div style="margin-bottom:8px; border-bottom:1px dashed #eee; padding-bottom:4px;">
                                <div class="grid-dense">
                                    ${{COLS.map(c => `<div class="gd-cell"><span class="gd-lbl">${{c}}</span><div class="gd-val">${{row[c]||'-'}}</div></div>`).join('')}}
                                </div>
                            </div>
                        `).join('')}}
                    </div>` : '';
                return `
                <div class="card" style="padding:8px; cursor:pointer" onclick="togExpand('${{k}}')">
                    <div class="flex-ac">
                        <div class="h-num" style="background:${{WAKU[w]}}; color:${{TXT[w]}}">${{k}}</div>
                        <div class="flex-c" style="flex:1">
                            <div class="font-bold text-sm">${{meta['馬名']}}</div>
                            <div class="text-xs text-light">${{meta['騎手']}} (${{list.length}}走)</div>
                        </div>
                        <div class="text-light">${{arrow}}</div>
                    </div>
                    ${{body}}
                </div>`;
            }}).join('');
        }}
        return `<div class="card"><div class="font-bold text-xs mb-2">▼ 並び替え（タップでランキング表示）</div><div class="sort-area">${{headers}}</div></div>${{content}}`;
    }}

    window.togExpand = (k) => {{
        const key = state.race+'-'+k;
        setState({{expanded: {{...state.expanded, [key]:!state.expanded[key]}} }});
    }};

    window.togMatchup = (k) => {{
        const key = state.race+'-m-'+k;
        setState({{expanded: {{...state.expanded, [key]:!state.expanded[key]}} }});
    }};

    // --- MATCHUP ---
    function renderMatchup() {{
        const m = DATA.matchups[state.race];
        const mx = DATA.matrix[state.race];
        if(!m) return '<div class="card">データなし</div>';

        const sortedKeys = Object.keys(m).sort((a,b)=>parseInt(a)-parseInt(b));

        const headRow = sortedKeys.map(k => `<th>${{k}}</th>`).join('');
        const matrixBody = sortedKeys.map(h1 => {{
            const cells = sortedKeys.map(h2 => {{
                if(h1 === h2) return '<td style="background:#ddd">-</td>';
                const rec = (mx[h1] && mx[h1][h2]);
                if(!rec) return '<td></td>';
                // rec: w, l, d
                let cls = 'mx-draw';
                let txt = '';
                if(rec.w > 0) {{ cls='mx-win'; txt=`${{rec.w}}勝`; }}
                else if(rec.l > 0) {{ cls='mx-lose'; txt=`${{rec.l}}敗`; }}
                else {{ txt=`${{rec.d}}分`; }}
                return `<td class="${{cls}}">${{txt}}</td>`;
            }}).join('');
            return `<tr><th style="background:${{WAKU[getWaku(h1)]}};color:${{TXT[getWaku(h1)]}}">${{h1}}</th>${{cells}}</tr>`;
        }}).join('');

        const matrixHTML = `
            <div class="card">
                <div class="font-bold text-xs mb-2">対戦早見表 (縦 vs 横)</div>
                <div class="matrix-box">
                    <table class="matrix-table">
                        <thead><tr><th></th>${{headRow}}</tr></thead>
                        <tbody>${{matrixBody}}</tbody>
                    </table>
                </div>
            </div>`;

        const listHTML = sortedKeys.map(k => {{
            const list = m[k];
            const w = getWaku(k);
            const stats = {{w:0, l:0, d:0}};
            list.forEach(x => {{ if(x.result=='Win')stats.w++; else if(x.result=='Lose')stats.l++; else stats.d++; }});

            const open = state.expanded[state.race+'-m-'+k];
            const arrow = open ? '▲' : '▼';

            const details = open ? list.map(x => {{
                let cls = 'mx-draw';
                if(x.result==='Win') cls='mx-win';
                if(x.result==='Lose') cls='mx-lose';
                return `
                <div style="padding:8px; border-left:4px solid transparent; margin-bottom:6px; font-size:11px; background:#f9fafb; border-radius:4px" class="${{cls}}">
                    <div class="flex-sb" style="margin-bottom:4px;">
                        <div class="font-bold">vs ${{x.opponent_name}} (${{x.opponent_num}}番)</div>
                        <div style="font-weight:bold">${{x.result}}</div>
                    </div>
                    <div class="text-xs text-light" style="margin-bottom:4px;">
                        ${{x.date}} ${{x.place}} ${{x.race_name}} (${{x.course}} ${{x.dist}})
                    </div>
                    <div style="display:grid; grid-template-columns:1fr 1fr; gap:8px;">
                        <div>
                            <div class="text-xs text-light">自分 (${{x.my_rank}}着)</div>
                            <div class="grid-dense" style="grid-template-columns:1fr 1fr; margin-top:2px;">
                                <div class="gd-cell"><span class="gd-lbl">T指数</span><div class="gd-val">${{x.my_idx}}</div></div>
                                <div class="gd-cell"><span class="gd-lbl">通過</span><div class="gd-val">${{x.my_pass}}</div></div>
                                <div class="gd-cell"><span class="gd-lbl">上がり</span><div class="gd-val">${{x.my_l3f}}</div></div>
                                <div class="gd-cell"><span class="gd-lbl">RPCI</span><div class="gd-val">${{x.my_rpci}}</div></div>
                                <div class="gd-cell" style="grid-column:1/3"><span class="gd-lbl">タイム</span><div class="gd-val">${{x.time}}</div></div>
                            </div>
                        </div>
                        <div>
                            <div class="text-xs text-light">相手 (${{x.op_rank}}着)</div>
                            <div class="grid-dense" style="grid-template-columns:1fr 1fr; margin-top:2px;">
                                <div class="gd-cell"><span class="gd-lbl">T指数</span><div class="gd-val">${{x.op_idx}}</div></div>
                                <div class="gd-cell"><span class="gd-lbl">通過</span><div class="gd-val">${{x.op_pass}}</div></div>
                                <div class="gd-cell"><span class="gd-lbl">上がり</span><div class="gd-val">${{x.op_l3f}}</div></div>
                                <div class="gd-cell"><span class="gd-lbl">RPCI</span><div class="gd-val">${{x.op_rpci}}</div></div>
                                <div class="gd-cell" style="grid-column:1/3"><span class="gd-lbl">タイム</span><div class="gd-val">${{x.op_time}}</div></div>
                            </div>
                        </div>
                    </div>
                </div>`;
            }}).join('') : '';

            // border-bottom style adjust
            const bb = open ? '1px solid #eee' : 'none';

            return `
            <div class="card" style="padding:0">
                <div class="flex-ac" style="padding:8px; border-bottom:${{bb}}; cursor:pointer" onclick="togMatchup('${{k}}')">
                    <div class="h-num" style="background:${{WAKU[w]}}; color:${{TXT[w]}}">${{k}}</div>
                    <div class="font-bold text-sm">対戦成績: ${{stats.w}}勝 ${{stats.l}}敗 ${{stats.d}}分</div>
                    <div class="text-light" style="margin-left:auto">${{arrow}}</div>
                </div>
                ${{open ? `<div style="padding:8px">${{details}}</div>` : ''}}
            </div>`;
        }}).join('');

        return matrixHTML + listHTML;
    }}

    function renderNav() {{
        const tabs = [{{id:'chart',l:'チャート',i:'📊'}},{{id:'data',l:'データ',i:'📋'}},{{id:'matchup',l:'比較',i:'⚔️'}},{{id:'pacing',l:'ラップ',i:'⏱️'}}];
        return `<div class="nav">${{tabs.map(t=>
            `<div class="nav-btn ${{state.tab==t.id?'active':''}}" onclick="setState({{tab:'${{t.id}}'}})">
                <div class="nav-icon">${{t.i}}</div>${{t.l}}
             </div>`).join('')}}</div>`;
    }}

    init();
</script>
</body>
</html>
'''
    return html

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="インタラクティブレポート生成")
    parser.add_argument("source", help="元のレポートファイル")
    parser.add_argument("-o", "--output", help="出力ファイル")
    args = parser.parse_args()
    result = create_interactive_charts(args.source, args.output)
    print(f"完了: {result}" if result else "失敗")
