from openpyxl import load_workbook

wb = load_workbook('outputs/reports/race_analysis_20260228_中山.xlsx', data_only=True)
sheets = [s for s in wb.sheetnames if s.startswith('Race_')]
if sheets:
    sheet = wb[sheets[-1]]
    headers = [cell.value for cell in sheet[1]]
    print(f"Sheet: {sheets[-1]}")
    print(f"Headers: {headers}")
    
    # print first row to see if race_id is there
    for row in sheet.iter_rows(min_row=2, max_row=2, values_only=True):
        print("Row 1 data:")
        for k, v in zip(headers, row):
            if k:
                print(f"  {k}: {v}")
